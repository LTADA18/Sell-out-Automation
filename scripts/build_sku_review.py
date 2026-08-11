r"""ตารางไกด์สินค้าที่จับคู่ SKU ไม่ได้ — ดึงจาก Postgres ทั้งฐาน รวมทุกร้าน

ใช้เป็นฐานข้อมูลอ้างอิง เจ้าของงานกรอก SKU/Model ที่ถูกต้องลงไป
แล้วระบบสกรีนเอาไปใช้จับคู่รอบหน้า

การยุบซ้ำ — จาก 308,516 บรรทัดออเดอร์ เหลือประมาณ 14,800 รายการ
  1 แถว = 1 สินค้า ไม่ใช่ 1 ออเดอร์  ไกด์ครั้งเดียวใช้ได้กับทุกออเดอร์ของสินค้านั้น
  รวมข้ามร้านและข้ามแพลตฟอร์มแล้ว สินค้าตัวเดียวกันที่ 5 ร้านขายจะเหลือแถวเดียว

⚠️ ยุบทิ้ง variation ไม่ได้ — จะเหลือ 8,635 แถวก็จริง แต่ variation คือตัวที่แยก
   แบต 4Ah ออกจาก 2Ah ยุบแล้วของคนละรุ่นจะปนกันโดยไม่มีอะไรเตือน

    .\.venv\Scripts\python.exe -u scripts\build_sku_review.py
"""
from __future__ import annotations

import argparse
import os
import subprocess
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
PSQL = r"C:\Program Files\PostgreSQL\18\bin\psql.exe"
PGPASS = r"C:\Users\tada.p\Postgres\pgpass.conf"

# แกนชื่อ = ตัดวงเล็บ คำโปรโมท อีโมจิ ออก เหลือแกนจริงของชื่อสินค้า
# ทำให้ชื่อเดียวกันที่แต่ละร้านแต่งต่างกันยุบมารวมเป็นแถวเดียวได้
NORM = r"""
  btrim(regexp_replace(
    regexp_replace(
      regexp_replace(lower(product_name), '\([^)]*\)|\[[^]]*\]', ' ', 'g'),
      '(ถูกสุด|ราคาถูก|สินค้าใหม่|ใหม่ล่าสุด|พร้อมส่ง|ส่งฟรี|ลดราคา|แถม|โปรโมชั่น|ของแท้|รับประกัน)',
      ' ', 'g'),
    '[^0-9a-zก-๙]+', ' ', 'g'))
"""

SQL = f"""
WITH v AS (
  SELECT product_name, coalesce(variation,'') AS variation, sku, shop_name, platform,
         quantity, revenue_thb, counts_as_sale, ordered_at, osuka_model_code, product_brand,
         {NORM} AS norm_name
  FROM   intel.mp_order_line
  WHERE  coalesce(osuka_sml_id,'') = ''
    -- เฉพาะสินค้า OSUKA ตามที่เจ้าของงานสั่ง แบรนด์อื่นไม่เอา
    --
    -- ⚠️ กรองด้วยคำว่า osuka อย่างเดียวไม่พอ ตกหล่นเพราะสะกดเพี้ยนและเพราะบางแถว
    --    ระบุแต่รหัสรุ่นไม่ระบุแบรนด์ ตรวจกับฐานแล้วได้กติกา 4 ข้อนี้
    --
    --   1. แบรนด์เป็น OSUKA หรือตัวสะกดเพี้ยนที่ยืนยันแล้วจากฐาน
    --      OSK 144 · OSID 13 · OSUKAX 4 · OSUKS 3 · OSLUKA 1
    --      (TOSLI / TROSLI / TOSAKI ไม่นับ เป็นแบรนด์อื่นจริง)
    --   2. ชื่อหรือ sku มีคำว่า osuka หรือคำเพี้ยน
    --   3. ชื่อสินค้ามีรหัสรุ่นขึ้นต้น OC/OS เช่น OCHD802 OSID832 OSAF3211
    --      ตรวจแล้ว 289,804 แถวเป็น OSUKA ล้วน มี OSUKAX ปนแค่ 4 แถว
    --   4. sku ขึ้นต้น OSK — แต่ต้องไม่ใช่แบรนด์อื่นที่ระบุมาชัด
    --      ⚠️ ข้อนี้ดักกับดัก: OSK0088 ถูก BRAVO ใช้ด้วย 3,812 แถว
    --         ถ้ากรองด้วยคำนำหน้า sku อย่างเดียวจะลาก BRAVO/BONCHI/RSK/VERGIN เข้ามา
    AND (
          product_brand ~* '^(osuka|osk|osuks|osluka|osukax|oslika|osid)$'
       OR product_name  ~* 'osuka|oslika|osuks|osluka|osukax'
       OR sku           ~* 'osuka'
       OR product_name  ~* '\\m(OC|OS)[A-Z]{{2,}}[0-9]'
       OR (upper(regexp_replace(sku,'[^0-9A-Za-z]','','g')) LIKE 'OSK%'
           AND (coalesce(product_brand,'') = ''
                OR product_brand ~* '^(osuka|osk|osuks|osluka|osukax|oslika|osid)$'))
        )
),
g AS (
  SELECT norm_name, variation,
         round(coalesce(sum(revenue_thb) FILTER (WHERE counts_as_sale),0))::bigint AS revenue,
         count(*)::bigint                       AS lines,
         coalesce(sum(quantity),0)::bigint      AS units,
         count(DISTINCT shop_name)              AS shops,
         count(DISTINCT sku)                    AS skus,
         string_agg(DISTINCT shop_name, ', ')   AS shop_names,
         string_agg(DISTINCT platform, ', ')    AS platforms,
         string_agg(DISTINCT coalesce(nullif(product_brand,''),'(ว่าง)'), ', ') AS brands,
         string_agg(DISTINCT sku, ' | ')        AS all_skus,
         string_agg(DISTINCT coalesce(nullif(osuka_model_code,''),''), ' | ') AS model_guess,
         min(ordered_at)::date                  AS first_seen,
         max(ordered_at)::date                  AS last_seen,
         (array_agg(product_name ORDER BY revenue_thb DESC NULLS LAST))[1] AS sample_name
  FROM   v GROUP BY norm_name, variation
)
SELECT variation, sample_name, revenue, lines, units, shops, skus,
       shop_names, platforms, brands, left(all_skus, 300), btrim(model_guess, ' |'),
       first_seen, last_seen
FROM   g ORDER BY revenue DESC NULLS LAST
"""

# ⚠️ ชื่อคอลัมน์คงที่ — ระบบสกรีนกับตัวตรวจจะอ้างอิงชื่อพวกนี้ ห้ามเปลี่ยนโดยไม่แจ้ง
# 5 คอลัมน์แรกคือตารางไกด์ตามที่เจ้าของงานกำหนด ที่เหลือเป็นข้อมูลประกอบการตัดสิน
COLS = [
    ("Variation",                  30, "key"),
    ("Product Name",               62, "key"),
    ("✍ SKU ที่ถูกต้อง",             22, "fill"),
    ("✍ Model seller ที่ถูกต้อง",    24, "fill"),
    ("✍ ข้อเสนอแนะ",                40, "fill"),
    ("สะสม % ยอดขาย",              14, "info"),
    ("ยอดขาย",                     13, "info"),
    ("บรรทัดออเดอร์",               13, "info"),
    ("ชิ้น",                        9,  "info"),
    ("จำนวนร้าน",                   10, "info"),
    ("ร้านที่ขาย",                   30, "info"),
    ("แพลตฟอร์ม",                  14, "info"),
    ("แบรนด์ที่บันทึกไว้",             16, "info"),
    ("sku ที่ร้านใช้ตอนนี้",           34, "info"),
    ("รหัสรุ่นที่ระบบเดา",             20, "info"),
    ("ขายวันแรก",                   12, "info"),
    ("ขายวันล่าสุด",                 12, "info"),
    ("สถานะรีวิว",                   12, "fill"),
]
TEXT_COLS = {"✍ SKU ที่ถูกต้อง", "✍ Model seller ที่ถูกต้อง",
             "sku ที่ร้านใช้ตอนนี้", "รหัสรุ่นที่ระบบเดา", "Variation"}


def run_sql(sql: str) -> list[list[str]]:
    """⚠️ ต้องส่งผ่านไฟล์ ห้ามใช้ -c

    SQL ก้อนนี้มีช่วงอักขระไทยใน regex (`[^0-9a-zก-๙]+`) ถ้าส่งผ่าน -c บน Windows
    ตัวอักษรไทยจะเพี้ยนจน character class พัง แล้ว Postgres ฟ้อง
    invalid regular expression: quantifier operand invalid

    และ path ของไฟล์ต้องเป็น ASCII ล้วนด้วย — โฟลเดอร์ชื่อไทยส่งเข้า psql แล้ว
    กลายเป็นเครื่องหมายคำถาม หาไฟล์ไม่เจอ
    """
    env = dict(os.environ, PGPASSFILE=PGPASS, PGCLIENTENCODING="UTF8")
    tmp = PROJECT_ROOT / "output" / "_sku_review_tmp" / "query.sql"
    tmp.parent.mkdir(parents=True, exist_ok=True)
    tmp.write_text(sql, encoding="utf-8")          # UTF-8 ไม่มี BOM
    p = subprocess.run(
        [PSQL, "service=osuka", "-w", "-A", "-t", "-F", "\x1f", "-f", str(tmp)],
        capture_output=True, env=env,
    )
    out = p.stdout.decode("utf-8", errors="replace")
    if p.returncode != 0:
        print(out, p.stderr.decode("utf-8", errors="replace"))
        raise SystemExit(f"❌ psql ไม่ผ่าน (exit {p.returncode})")
    return [ln.split("\x1f") for ln in out.splitlines() if ln.strip()]


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--top", type=int, default=0, help="เอาแค่ N อันดับแรก (0 = ทั้งหมด)")
    args = ap.parse_args()

    print("ดึงจาก Postgres ...")
    rows = run_sql(SQL)
    total_rev = sum(int(r[2] or 0) for r in rows)
    total_lines = sum(int(r[3] or 0) for r in rows)
    print(f"  {total_lines:,} บรรทัดออเดอร์  ->  {len(rows):,} รายการที่ต้องไกด์")
    if args.top:
        rows = rows[: args.top]

    wb = Workbook()
    ws = wb.active
    ws.title = "ไกด์ SKU"

    fills = {
        "key":  PatternFill("solid", fgColor="1F4E79"),
        "fill": PatternFill("solid", fgColor="C55A11"),
        "info": PatternFill("solid", fgColor="595959"),
    }
    for i, (name, width, kind) in enumerate(COLS, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fills[kind]
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 34

    cum = 0
    marks = [50, 100, 200, 500, 1000, 2000, 5000, len(rows)]
    acc: dict[int, int] = {}
    for n, r in enumerate(rows, start=1):
        (variation, name, revenue, lines, units, shops, skus,
         shop_names, platforms, brands, all_skus, model_guess, first_seen, last_seen) = r
        rev = int(revenue or 0)
        cum += rev
        if n in marks:
            acc[n] = cum
        vals = [
            variation, name, "", "", "",
            round(cum / total_rev * 100, 2) if total_rev else 0,
            rev, int(lines or 0), int(units or 0), int(shops or 0),
            shop_names, platforms, brands, all_skus, model_guess,
            first_seen, last_seen, "รอไกด์",
        ]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=n + 1, column=i, value=v)
            head = COLS[i - 1][0]
            if head in TEXT_COLS:
                c.number_format = "@"     # กันรหัสที่มีเลข 0 นำหน้าโดนตัดทิ้ง
            elif head == "ยอดขาย":
                c.number_format = "#,##0"
            elif head == "สะสม % ยอดขาย":
                c.number_format = "0.00"
            if head in ("Product Name", "ร้านที่ขาย", "sku ที่ร้านใช้ตอนนี้", "✍ ข้อเสนอแนะ"):
                c.alignment = Alignment(wrap_text=True, vertical="top")
    acc[len(rows)] = cum

    ws.freeze_panes = "C2"          # ตรึง Variation + Product Name ไว้ให้เห็นตลอด
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(rows) + 1}"

    # ---------- ชีทสรุป ----------
    s = wb.create_sheet("สรุป")
    s["A1"], s["B1"], s["C1"] = "ไกด์ถึงอันดับที่", "ครอบคลุมยอดขาย", "คิดเป็น %"
    for c in ("A1", "B1", "C1"):
        s[c].font = Font(bold=True)
    for i, m in enumerate([m for m in marks if m in acc], start=2):
        s.cell(row=i, column=1, value=m)
        s.cell(row=i, column=2, value=acc[m]).number_format = "#,##0"
        s.cell(row=i, column=3, value=round(acc[m] / total_rev * 100, 1) if total_rev else 0)
    for col, w in (("A", 18), ("B", 20), ("C", 12)):
        s.column_dimensions[col].width = w

    # ---------- ชีทวิธีใช้ ----------
    h = wb.create_sheet("วิธีใช้")
    guide = [
        "ไฟล์นี้คืออะไร",
        f"  สินค้าที่ระบบจับคู่กับตัวกลาง SML ไม่ได้ ดึงจากฐาน intel.mp_order_line ทั้งก้อน",
        f"  ยุบจาก {total_lines:,} บรรทัดออเดอร์ เหลือ {len(rows):,} รายการ",
        "  1 แถว = 1 สินค้า รวมทุกร้านทุกแพลตฟอร์มแล้ว ไกด์ครั้งเดียวใช้ได้กับทุกออเดอร์",
        "",
        "ไกด์ยังไงให้คุ้มแรง",
        "  เรียงตามยอดขายมากไปน้อยแล้ว ดูคอลัมน์ สะสม % ยอดขาย",
        "  ไล่จากบนลงล่าง พอ % ถึงระดับที่พอใจก็หยุดได้ ตัวเลขจริงอยู่ในชีท สรุป",
        "",
        "ช่องสีส้มคือช่องที่ต้องกรอก",
        "  ✍ SKU ที่ถูกต้อง          SKU ในตัวกลาง SML ที่สินค้านี้ควรจับคู่ด้วย",
        "  ✍ Model seller ที่ถูกต้อง  รหัสรุ่นที่ถูกต้อง",
        "  ✍ ข้อเสนอแนะ             เหตุผล ข้อสังเกต หรือใส่ ไม่ใช่สินค้าเรา ถ้าเป็นของแถม/ค่าส่ง",
        "  สถานะรีวิว               รอไกด์ -> ไกด์แล้ว -> ยืนยันแล้ว",
        "",
        "คอลัมน์สีเทาคือข้อมูลประกอบ ไม่ต้องแก้",
        "  sku ที่ร้านใช้ตอนนี้   รหัสที่ร้านตั้งเอง อาจมีหลายอันคั่นด้วย |",
        "  รหัสรุ่นที่ระบบเดา     ระบบสกรีนเดาไว้แต่ยังจับคู่ไม่สำเร็จ ใช้เป็นจุดตั้งต้นได้",
        "  ร้านที่ขาย            ถ้าหลายร้านขายตัวเดียวกัน แปลว่ายิ่งคุ้มที่จะไกด์",
        "",
        "สำหรับ AI agent 2 ตัวที่จะทำต่อ",
        "  ตัวสกรีน  อ่าน Variation / Product Name / sku ที่ร้านใช้ตอนนี้",
        "            เติม ✍ SKU ที่ถูกต้อง กับ ✍ Model seller ที่ถูกต้อง แล้วตั้งสถานะเป็น ไกด์แล้ว",
        "  ตัวตรวจ   ตรวจกับตัวกลาง SML แล้วตั้งเป็น ยืนยันแล้ว",
        "            ถ้าไม่เห็นด้วย เขียนเหตุผลใน ✍ ข้อเสนอแนะ แล้วตั้งกลับเป็น รอไกด์",
        "  ชื่อคอลัมน์คงที่ ห้ามเปลี่ยนโดยไม่แจ้ง เพราะทั้งสองตัวอ้างอิงชื่อ",
        "",
        "ข้อควรระวัง",
        "  ช่อง SKU กับรหัสรุ่นตั้งเป็นข้อความไว้แล้ว อย่าเปลี่ยนเป็นตัวเลข",
        "  ไม่งั้นรหัสที่มีเลข 0 นำหน้าจะโดนตัดทิ้งเงียบ ๆ",
        "  variation ไม่ได้ถูกยุบทิ้ง เพราะเป็นตัวแยกแบต 4Ah ออกจาก 2Ah",
        "  ยุบแล้วจะเหลือ 8,635 แถวก็จริง แต่ของคนละรุ่นจะปนกันโดยไม่มีอะไรเตือน",
    ]
    for i, line in enumerate(guide, start=1):
        c = h.cell(row=i, column=1, value=line)
        if line and not line.startswith("  "):
            c.font = Font(bold=True)
    h.column_dimensions["A"].width = 100

    out_dir = PROJECT_ROOT / "output" / "_รีวิว_SKU"
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / "ไกด์ SKU สินค้าที่จับคู่ไม่ได้.xlsx"
    wb.save(out)

    print(f"\n✅ {out.relative_to(PROJECT_ROOT)}  ({out.stat().st_size/1024/1024:,.1f} MB)")
    print("\n   ไกด์ถึงอันดับ   ครอบคลุมยอดขาย")
    for m in marks:
        if m in acc:
            print(f"     {m:>6,}        {acc[m]/total_rev*100:>5.1f}%   ({acc[m]:,.0f} บาท)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
