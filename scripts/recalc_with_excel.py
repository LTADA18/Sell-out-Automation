r"""คำนวณสูตรในไฟล์ Excel ด้วย Excel ตัวจริง แล้วตรวจว่ามีสูตรพังไหม

ใช้แทน recalc.py ของ skill ซึ่งรันบน Windows ไม่ได้ (พึ่ง socket แบบ Unix)
เครื่องนี้มี Excel 16.0 อยู่แล้ว จึงสั่งผ่าน COM ได้ตรง ๆ

openpyxl เขียนสูตรเป็นข้อความโดยไม่มีค่าที่คำนวณไว้ ถ้าไม่ผ่านขั้นนี้
ตัวที่อ่านค่าจาก cache (pandas, openpyxl data_only=True, ตัวพรีวิวส่วนใหญ่)
จะเห็นเป็นค่าว่างทั้งหมด

    .\.venv\Scripts\python.exe -u scripts\recalc_with_excel.py <ไฟล์>
"""
from __future__ import annotations

import sys
from pathlib import Path

import win32com.client as win32
from openpyxl import load_workbook

ERRORS = ("#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#N/A", "#NULL!", "#NUM!")


def main() -> int:
    if len(sys.argv) < 2:
        print("ใช้: recalc_with_excel.py <ไฟล์.xlsx>")
        return 2
    path = Path(sys.argv[1]).resolve()
    if not path.exists():
        print(f"ไม่พบไฟล์ {path}")
        return 2

    app = win32.DispatchEx("Excel.Application")
    app.Visible = False
    app.DisplayAlerts = False
    try:
        wb = app.Workbooks.Open(str(path))
        app.CalculateFullRebuild()
        wb.Save()
        wb.Close(SaveChanges=True)
    finally:
        app.Quit()

    # อ่านกลับด้วยค่าที่คำนวณแล้ว — ถ้ายังเป็น None แปลว่าไม่ได้คำนวณจริง
    wb2 = load_workbook(path, data_only=True)
    n_err = 0
    n_val = 0
    try:
        for ws in wb2.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if c.value is None:
                        continue
                    if isinstance(c.value, str) and c.value.strip() in ERRORS:
                        print(f"  ❌ {ws.title}!{c.coordinate} = {c.value}")
                        n_err += 1
                    n_val += 1
    finally:
        wb2.close()

    print(f"เซลล์ที่มีค่า {n_val:,} · สูตรพัง {n_err}")
    print("✅ ไม่มีสูตรพัง" if n_err == 0 else f"❌ ต้องแก้ {n_err} เซลล์")
    return 1 if n_err else 0


if __name__ == "__main__":
    raise SystemExit(main())
