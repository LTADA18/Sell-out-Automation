r"""รวมไฟล์รายวันหลายวันเป็นไฟล์เดียว — ทุกร้านทุกแพลตฟอร์มในชีทเดียว

แบบเดียวกับไฟล์รวมของงานย้อนหลัง 7 เดือน แต่เลือกช่วงวันได้

ค่าเริ่มต้นรวม **ไฟล์ที่ผ่านการสกรีนแล้ว (63 คอลัมน์)** เพราะเป็นของส่งมอบจริง
ใส่ --raw ถ้าอยากได้ไฟล์ดิบ 32 คอลัมน์แทน

⚠️ โฟลเดอร์ output ใช้ "วันที่รัน" = วันของข้อมูล + 1 วัน
   ช่วง 1-9 ส.ค. จึงอยู่ในโฟลเดอร์ 2026-08-02 ถึง 2026-08-10

    .\.venv\Scripts\python.exe -u scripts\merge_range.py --from 2026-08-01 --to 2026-08-09
"""
from __future__ import annotations

import argparse
import sys
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from src.core.config import load_config                  # noqa: E402

MAX_ROWS = 1_000_000                                      # Excel รับ 1,048,576 เผื่อหัวตาราง
TEXT_COLS = {"order_id", "sku", "tracking_no"}

# ไฟล์สรุปของระบบสกรีน — ไม่ใช่ข้อมูลรายบรรทัด ห้ามเอามารวม
SKIP_NAMES = {"brand_summary.xlsx", "data_issues.xlsx", "missing_models.xlsx"}


def read_rows(path: Path) -> tuple[list[str], list[list]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        # ไฟล์สกรีนใช้ชีท data · ไฟล์ดิบใช้ Orders
        for cand in ("data", "Orders"):
            if cand in wb.sheetnames:
                ws = wb[cand]
                break
        else:
            ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = [str(c) if c is not None else "" for c in next(it)]
        rows = [list(r) for r in it if any(v is not None for v in r)]
        return header, rows
    finally:
        wb.close()


def write_sheet(ws, header: list[str], rows: list[list], text_idx: set[int]) -> None:
    ws.append(header)
    for r in rows:
        cells = []
        for i, v in enumerate(r):
            c = WriteOnlyCell(ws, value=v)
            if i in text_idx:
                # เลข 19 หลักของ TikTok ถ้าหลุดเป็นตัวเลข Excel จะปัดหลักท้ายทิ้งเงียบ ๆ
                c.number_format = "@"
            cells.append(c)
        ws.append(cells)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--from", dest="d_from", required=True)
    ap.add_argument("--to", dest="d_to", required=True)
    ap.add_argument("--raw", action="store_true", help="รวมไฟล์ดิบ 32 คอลัมน์แทนไฟล์สกรีน")
    ap.add_argument("--out", help="โฟลเดอร์ปลายทาง")
    args = ap.parse_args()

    d_from = date.fromisoformat(args.d_from)
    d_to = date.fromisoformat(args.d_to)
    cfg = load_config()
    out_root = PROJECT_ROOT / cfg.settings.paths.output_dir

    files: list[Path] = []
    day = d_from
    while day <= d_to:
        run_date = (day + timedelta(days=1)).isoformat()   # โฟลเดอร์ = วันข้อมูล + 1
        folder = out_root / run_date
        if args.raw:
            found = sorted(folder.glob("*.xlsx"))
        else:
            found = sorted((folder / "screened").glob("*.xlsx"))
        found = [f for f in found if f.name not in SKIP_NAMES]
        if not found:
            print(f"  ⚠️ {day} (โฟลเดอร์ {run_date}) ไม่มีไฟล์")
        files += found
        day += timedelta(days=1)

    if not files:
        print("❌ ไม่พบไฟล์เลย — ดึงหรือสกรีนยังไม่เสร็จ?")
        return 1

    kind = "ไฟล์ดิบ 32 คอลัมน์" if args.raw else "ไฟล์สกรีน 63 คอลัมน์"
    print(f"รวม {len(files)} ไฟล์ ({kind})\n")

    header: list[str] | None = None
    all_rows: list[list] = []
    per_day: dict[str, int] = {}
    for f in files:
        h, rows = read_rows(f)
        if header is None:
            header = h
        elif h != header:
            # คอลัมน์ไม่ตรง = เอามารวมแล้วข้อมูลจะเลื่อนช่อง ต้องรู้ ไม่ใช่รวมมั่ว
            print(f"  ❌ {f.name[:52]}: คอลัมน์ไม่ตรงกับไฟล์แรก ({len(h)} vs {len(header)}) — ข้าม")
            continue
        all_rows += rows
        per_day[f.parent.parent.name if not args.raw else f.parent.name] = \
            per_day.get(f.parent.parent.name if not args.raw else f.parent.name, 0) + len(rows)
        print(f"  {f.name[:56]:<58} {len(rows):>7,} แถว")

    assert header is not None
    text_idx = {i for i, c in enumerate(header) if c in TEXT_COLS}

    print(f"\nรวมทั้งหมด {len(all_rows):,} แถว · {len(header)} คอลัมน์")
    print("แยกตามโฟลเดอร์ (วันที่รัน):")
    for k in sorted(per_day):
        print(f"  {k}  {per_day[k]:>7,}")

    out_dir = Path(args.out) if args.out else out_root / f"_report_{d_from}_{d_to}"
    out_dir.mkdir(parents=True, exist_ok=True)
    suffix = "ดิบ" if args.raw else "สกรีนแล้ว"
    path = out_dir / f"ทุกร้านทุกแพลตฟอร์ม_{d_from}_ถึง_{d_to}_{suffix}.xlsx"

    wb = Workbook(write_only=True)
    for i in range(0, max(len(all_rows), 1), MAX_ROWS):
        chunk = all_rows[i:i + MAX_ROWS]
        name = "ALL" if i == 0 else f"ALL_{i // MAX_ROWS + 1}"
        write_sheet(wb.create_sheet(name), header, chunk, text_idx)
        print(f"  ชีท {name}: {len(chunk):,} แถว")
    wb.save(path)

    print(f"\n✅ {path.relative_to(PROJECT_ROOT)}  ({path.stat().st_size / 1024 / 1024:.1f} MB)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
