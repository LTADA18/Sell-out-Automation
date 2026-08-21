r"""ดึงข้อมูลร้านเดียวจากไฟล์สกรีน ออกเป็น Excel ที่เปิดดูไหว

    .\.venv\Scripts\python.exe -u scripts\build_shop_review.py --file <ไฟล์ _matched.xlsx>

⚠️ ไฟล์สกรีนตัวเต็มมี 104 คอลัมน์ และของร้านใหญ่หนักถึง 82 MB
   เปิดดูด้วยตาจริง ๆ ใช้ไม่กี่คอลัมน์ ตัวนี้จึงคัดเฉพาะที่ใช้ตอบคำถามว่า
   "สินค้าตัวนี้จับคู่รหัสรุ่น OSUKA ได้ไหม เพราะอะไร"

ออกมา 2 ชีท
  ทุกบรรทัด   — ทุกแถวพร้อมเหตุผลการจับคู่
  สรุป        — นับตามแบรนด์ / สถานะการจับคู่ / เหตุผล
"""
from __future__ import annotations

import argparse
import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]

# (ชื่อคอลัมน์ในไฟล์สกรีน, ชื่อที่จะแสดง, ความกว้าง)
PICK = [
    ("order_id",            "เลขออเดอร์",        22),
    ("ordered_at",          "วันที่สั่ง",          18),
    ("order_status",        "สถานะ",            14),
    ("sku",                 "SKU",              20),
    ("product_name",        "ชื่อสินค้า",          58),
    ("variation",           "ตัวเลือกสินค้า",       26),
    ("quantity",            "ชิ้น",               7),
    ("item_price",          "ราคา/ชิ้น",          11),
    ("revenue_thb",         "ยอดขาย",           12),
    ("product_brand",       "แบรนด์",            15),
    ("brand_status",        "ความมั่นใจแบรนด์",     14),
    ("is_osuka_brand",      "เป็น OSUKA",        11),
    ("osuka_model_code",    "รหัสรุ่น OSUKA",      17),
    ("osuka_product_name",  "ชื่อสินค้าฝั่ง master",  40),
    ("match_status",        "ผลการจับคู่",         16),
    ("mapping_status",      "สถานะจับคู่",         13),
    ("match_confidence",    "ความมั่นใจ",          11),
    ("accuracy_matching_%", "ความแม่น %",         11),
    ("matched_by",          "จับคู่ด้วยกฎ",         46),
    ("review_reason",       "เหตุผลที่ต้องตรวจ",     34),
    ("review_question",     "คำถามถึงคนตรวจ",      44),
]

TEXT_COLS = {"เลขออเดอร์", "SKU", "รหัสรุ่น OSUKA"}
HDR = PatternFill("solid", fgColor="1F4E79")
SUB = PatternFill("solid", fgColor="DDEBF7")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True, help="ไฟล์ *_matched.xlsx")
    ap.add_argument("--out", default="", help="ไฟล์ปลายทาง (ไม่ใส่ = ตั้งชื่อเอง)")
    args = ap.parse_args()

    src = Path(args.file)
    if not src.exists():
        print(f"❌ ไม่พบไฟล์ {src}")
        return 1

    wb_in = load_workbook(src, read_only=True, data_only=True)
    ws_in = wb_in["data"]
    it = ws_in.iter_rows(values_only=True)
    hdr = [str(c) if c is not None else "" for c in next(it)]

    have = [(c, label, w) for c, label, w in PICK if c in hdr]
    miss = [c for c, _l, _w in PICK if c not in hdr]
    if miss:
        print(f"⚠️ ไฟล์นี้ไม่มีคอลัมน์: {miss}")
    idx = {c: hdr.index(c) for c, _l, _w in have}

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("ทุกบรรทัด")
    head = []
    for i, (_c, label, w) in enumerate(have, start=1):
        cell = WriteOnlyCell(ws, value=label)
        cell.font, cell.fill = Font(bold=True, color="FFFFFF"), HDR
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        head.append(cell)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.append(head)

    n = 0
    by_brand: Counter = Counter()
    by_match: Counter = Counter()
    by_reason: Counter = Counter()
    for r in it:
        if not any(v is not None for v in r):
            continue
        n += 1
        row = []
        for c, label, _w in have:
            v = r[idx[c]]
            v = "" if v is None else v
            if label in TEXT_COLS:
                cell = WriteOnlyCell(ws, value=str(v))
                cell.number_format = "@"      # กฎเหล็กข้อ 2 — ห้ามหลุดเป็นตัวเลข
                row.append(cell)
            else:
                row.append(v)
        ws.append(row)
        by_brand[str(r[idx["product_brand"]] or "(ไม่มีชื่อแบรนด์)")] += 1
        if "match_status" in idx:
            by_match[str(r[idx["match_status"]] or "(ว่าง)")] += 1
        if "review_reason" in idx:
            rr = str(r[idx["review_reason"]] or "").strip()
            if rr and rr.lower() != "null":
                by_reason[rr[:60]] += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(have))}{n + 1}"
    wb_in.close()

    # ── ชีทสรุป ───────────────────────────────────────────────
    ws2 = wb.create_sheet("สรุป")
    for w, i in ((46, 1), (14, 2), (12, 3)):
        ws2.column_dimensions[get_column_letter(i)].width = w

    def block(title: str, counter: Counter, limit: int = 25) -> None:
        c = WriteOnlyCell(ws2, value=title)
        c.font, c.fill = Font(bold=True, color="FFFFFF"), HDR
        ws2.append([c, "", ""])
        h = []
        for t in ("ค่า", "บรรทัด", "% ของทั้งหมด"):
            x = WriteOnlyCell(ws2, value=t)
            x.font, x.fill = Font(bold=True), SUB
            h.append(x)
        ws2.append(h)
        for k, v in counter.most_common(limit):
            pct = WriteOnlyCell(ws2, value=round(100.0 * v / n, 1))
            pct.number_format = '0.0"%"'
            ws2.append([k, v, pct])
        ws2.append([])

    block(f"แบรนด์ที่ขาย (ทั้งหมด {n:,} บรรทัด)", by_brand)
    if by_match:
        block("ผลการจับคู่", by_match)
    if by_reason:
        block("เหตุผลที่ต้องให้คนตรวจ", by_reason)

    out = Path(args.out) if args.out else (
        PROJECT_ROOT / "output" / "_reports" / f"ตรวจสอบ_{src.stem.replace('_matched','')}.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    mb = out.stat().st_size / 1024 / 1024
    print(f"✅ {out}")
    print(f"   {n:,} บรรทัด · {len(have)} คอลัมน์ · {mb:,.1f} MB")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
