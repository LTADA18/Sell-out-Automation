r"""วัดว่าคอลัมน์ไหน "มีค่าจริงกี่ %" แยกตามแพลตฟอร์ม

ใช้ทำ Data Dictionary ให้ฝั่งฐานข้อมูล — ต้องบอกตามจริงว่าคอลัมน์ไหนใช้ได้
คอลัมน์ไหนว่างทั้งแถบ ไม่ใช่ประกาศ schema สวย ๆ แล้วปล่อยให้เขาไปเจอ NULL เอง

    .\.venv\Scripts\python.exe -u scripts\profile_columns.py
"""
from __future__ import annotations

import json
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

SRC = PROJECT_ROOT / "output" / "_report_2026h1"
PERIOD = "2026-01-01_ถึง_2026-07-31"
CAP = 60_000                                              # อ่านต่อแพลตฟอร์มเท่านี้พอ

# เลือกไฟล์ที่มีแต่ละแพลตฟอร์มอยู่จริง
SOURCES = {
    "lazada": ("กัปตัน เอกสตีล", "Lazada"),
    "tiktok": ("toolsdee1", "TikTok"),
    "shopee": ("Smarttooltech", "Shopee"),
}


def profile(path: Path, sheet: str) -> tuple[list[str], dict[str, int], int]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        header = [str(c) if c is not None else "" for c in next(it)]
        filled: dict[str, int] = defaultdict(int)
        n = 0
        for row in it:
            if not any(v is not None for v in row):
                continue
            n += 1
            for col, v in zip(header, row):
                if v is not None and str(v).strip() not in ("", "Null", "None"):
                    filled[col] += 1
            if n >= CAP:
                break
        return header, dict(filled), n
    finally:
        wb.close()


result: dict[str, dict[str, float]] = {}
header_ref: list[str] = []

for plat, (brand, sheet) in SOURCES.items():
    path = SRC / f"{brand}_{PERIOD}.xlsx"
    if not path.exists():
        print(f"⚠️  ไม่มีไฟล์ {path.name} — ข้าม {plat}")
        continue
    header, filled, n = profile(path, sheet)
    header_ref = header_ref or header
    result[plat] = {c: round(100 * filled.get(c, 0) / n, 1) if n else 0.0 for c in header}
    print(f"{plat:<8} อ่าน {n:,} แถว จาก {path.name[:40]}")

print(f"\n{'คอลัมน์':<22} {'Lazada':>8} {'TikTok':>8} {'Shopee':>8}   สรุป")
print("-" * 68)
for c in header_ref:
    vals = [result.get(p, {}).get(c, 0.0) for p in ("lazada", "tiktok", "shopee")]
    if all(v == 0 for v in vals):
        note = "ว่างทุกแพลตฟอร์ม"
    elif any(v == 0 for v in vals):
        have = [p for p, v in zip(("Lazada", "TikTok", "Shopee"), vals) if v > 0]
        note = f"มีเฉพาะ {'+'.join(have)}"
    elif all(v > 99.9 for v in vals):
        note = "ครบทุกแถว"
    else:
        note = "มีบ้างไม่มีบ้าง"
    print(f"{c:<22} {vals[0]:>7.1f}% {vals[1]:>7.1f}% {vals[2]:>7.1f}%   {note}")

out = PROJECT_ROOT / "docs" / "column_fill_rate.json"
out.parent.mkdir(exist_ok=True)
out.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"\nบันทึกตัวเลขดิบไว้ที่ {out.relative_to(PROJECT_ROOT)}")
