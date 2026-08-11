r"""ลองใช้กฎที่เจ้าของงานไกด์ไว้ จับคู่สินค้า OSUKA ที่ยังจับไม่ได้ — ยังไม่เขียนฐาน

อ่านกฎจาก config/sku_rules.yaml แล้วไล่จับกับรายการที่ยังค้างรีวิว
รายงานว่าลดคิวรีวิวได้เท่าไหร่ ครอบคลุมยอดขายเท่าไหร่

⚠️ สคริปต์นี้อ่านฐานอย่างเดียว ไม่เขียนอะไรลงไป
   ผลที่ได้เอาไปให้เจ้าของงานตรวจก่อนว่าจับถูกจริงไหม ค่อยตัดสินใจกันต่อ

    .\.venv\Scripts\python.exe -u scripts\apply_sku_guide.py
    .\.venv\Scripts\python.exe -u scripts\apply_sku_guide.py --show 40
"""
from __future__ import annotations

import argparse
import os
import re
import subprocess
from collections import defaultdict
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
PSQL = r"C:\Program Files\PostgreSQL\18\bin\psql.exe"
PGPASS = r"C:\Users\tada.p\Postgres\pgpass.conf"
RULES_FILE = PROJECT_ROOT / "config" / "sku_rules.yaml"

_WS = re.compile(r"\s+")

# ตัวกรอง OSUKA ชุดเดียวกับ build_sku_review.py — ถ้าแก้ที่นั่นต้องแก้ที่นี่ด้วย
SQL = r"""
WITH v AS (
  SELECT product_name, coalesce(variation,'') AS variation, sku, shop_name,
         quantity, revenue_thb, counts_as_sale,
         btrim(regexp_replace(
           regexp_replace(
             regexp_replace(lower(product_name), '\([^)]*\)|\[[^]]*\]', ' ', 'g'),
             '(ถูกสุด|ราคาถูก|สินค้าใหม่|ใหม่ล่าสุด|พร้อมส่ง|ส่งฟรี|ลดราคา|แถม|โปรโมชั่น|ของแท้|รับประกัน)',
             ' ', 'g'),
           '[^0-9a-zก-๙]+', ' ', 'g')) AS norm_name
  FROM   intel.mp_order_line
  WHERE  coalesce(osuka_sml_id,'') = ''
    AND (
          product_brand ~* '^(osuka|osk|osuks|osluka|osukax|oslika|osid)$'
       OR product_name  ~* 'osuka|oslika|osuks|osluka|osukax'
       OR sku           ~* 'osuka'
       OR product_name  ~* '\m(OC|OS)[A-Z]{2,}[0-9]'
       OR (upper(regexp_replace(sku,'[^0-9A-Za-z]','','g')) LIKE 'OSK%'
           AND (coalesce(product_brand,'') = ''
                OR product_brand ~* '^(osuka|osk|osuks|osluka|osukax|oslika|osid)$'))
        )
)
SELECT variation,
       (array_agg(product_name ORDER BY revenue_thb DESC NULLS LAST))[1],
       string_agg(DISTINCT sku, ' | '),
       round(coalesce(sum(revenue_thb) FILTER (WHERE counts_as_sale),0))::bigint,
       count(*)::bigint,
       count(DISTINCT shop_name)
FROM   v GROUP BY norm_name, variation
ORDER  BY 4 DESC NULLS LAST
"""


STRIP: list[str] = []          # วลีปฏิเสธ ตั้งค่าจาก sku_rules.yaml


def norm(s: str) -> str:
    """ยุบช่องว่าง ทำเป็นตัวพิมพ์เล็ก และตัดวลีปฏิเสธทิ้งก่อนเทียบ

    ⚠️ วลีปฏิเสธสำคัญมาก — "ไม่มีไอน้ำ" มีคำว่า "ไอน้ำ" อยู่ข้างใน
    ถ้าไม่ตัดก่อน พัดลม OCF762 ที่เขียนว่าไม่มีไอน้ำ จะถูกจับเป็นรุ่นไอน้ำ OCF763
    """
    t = _WS.sub(" ", (s or "").replace("\xa0", " ")).strip().lower()
    for phrase in STRIP:
        t = t.replace(phrase, " ")
    return _WS.sub(" ", t).strip()


def run_sql(sql: str) -> list[list[str]]:
    """ส่งผ่านไฟล์เสมอ — regex มีช่วงอักขระไทย ส่งผ่าน -c แล้วเพี้ยน
    และ path ต้องเป็น ASCII ล้วน โฟลเดอร์ชื่อไทย psql หาไม่เจอ"""
    env = dict(os.environ, PGPASSFILE=PGPASS, PGCLIENTENCODING="UTF8")
    tmp = PROJECT_ROOT / "output" / "_sku_review_tmp" / "guide_query.sql"
    tmp.parent.mkdir(parents=True, exist_ok=True)
    tmp.write_text(sql, encoding="utf-8")
    p = subprocess.run(
        [PSQL, "service=osuka", "-w", "-A", "-t", "-F", "\x1f", "-f", str(tmp)],
        capture_output=True, env=env,
    )
    if p.returncode != 0:
        print(p.stdout.decode("utf-8", "replace"), p.stderr.decode("utf-8", "replace"))
        raise SystemExit("❌ psql ไม่ผ่าน")
    return [ln.split("\x1f") for ln in
            p.stdout.decode("utf-8", "replace").splitlines() if ln.strip()]


def cond_ok(cond: dict, name: str, var: str, sku: str) -> bool:
    for key, words in cond.items():
        w = [norm(x) for x in words]
        if key == "name_has" and not any(x in name for x in w):
            return False
        if key == "name_not" and any(x in name for x in w):
            return False
        if key == "var_has" and not any(x in var for x in w):
            return False
        if key == "var_not" and any(x in var for x in w):
            return False
        if key == "sku_has" and not any(x in sku for x in w):
            return False
        if key == "any_of_words_in_name_or_var" and not any(
                x in name or x in var for x in w):
            return False
    return True


def match(rules: list[dict], name: str, var: str, sku: str) -> dict | None:
    n, v, s = norm(name), norm(var), norm(sku)
    for r in rules:
        if "all" in r and all(cond_ok(c, n, v, s) for c in r["all"]):
            return r
        if "any" in r and any(cond_ok(c, n, v, s) for c in r["any"]):
            return r
    return None


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--show", type=int, default=25, help="โชว์ตัวอย่างกี่แถว")
    args = ap.parse_args()

    cfg = yaml.safe_load(RULES_FILE.read_text(encoding="utf-8"))
    rules = cfg["rules"]
    STRIP[:] = [p.lower() for p in cfg.get("strip_phrases", [])]
    print(f"อ่านกฎ {len(rules)} ข้อ · วลีปฏิเสธ {len(STRIP)} วลี "
          f"จาก {RULES_FILE.relative_to(PROJECT_ROOT)}")

    rows = run_sql(SQL)
    total_rev = sum(int(r[3] or 0) for r in rows)
    total_lines = sum(int(r[4] or 0) for r in rows)
    print(f"รายการที่ยังค้างรีวิว {len(rows):,} · {total_lines:,} บรรทัด · "
          f"ยอดขาย {total_rev:,.0f} บาท\n")

    hits: dict[str, list] = defaultdict(list)
    matched = miss = 0
    m_rev = m_lines = 0
    detail = []
    for variation, name, skus, revenue, lines, shops in rows:
        r = match(rules, name, variation, skus)
        if r is None:
            miss += 1
            continue
        matched += 1
        m_rev += int(revenue or 0)
        m_lines += int(lines or 0)
        hits[r["id"]].append((int(revenue or 0), int(lines or 0)))
        detail.append((int(revenue or 0), int(lines or 0), int(shops or 0),
                       r["id"], r.get("sku", ""), r.get("model", ""),
                       r.get("verdict", ""), variation, name))

    print("=" * 74)
    print(f"จับได้ {matched:,} รายการ จาก {len(rows):,}  ({matched/len(rows)*100:.1f}%)")
    print(f"คิดเป็น {m_lines:,} บรรทัดออเดอร์  ยอดขาย {m_rev:,.0f} บาท "
          f"({m_rev/total_rev*100:.1f}% ของที่ค้าง)")
    print(f"ยังจับไม่ได้ {miss:,} รายการ")
    print("=" * 74)

    print("\nแยกตามกฎ")
    for rid, lst in sorted(hits.items(), key=lambda kv: -sum(x[0] for x in kv[1])):
        r = next(x for x in rules if x["id"] == rid)
        tag = r.get("verdict") or f"{r.get('sku','')}/{r.get('model','')}"
        print(f"  {len(lst):>5,} รายการ · {sum(x[1] for x in lst):>7,} บรรทัด · "
              f"{sum(x[0] for x in lst):>12,} บาท   {rid:<26} {tag}")

    detail.sort(key=lambda x: -x[0])
    print(f"\nตัวอย่าง {min(args.show, len(detail))} อันดับแรกที่จับได้")
    for rev, lines, shops, rid, sku, model, verdict, var, name in detail[: args.show]:
        print(f"  {rev:>11,} | {rid:<24} | {(verdict or model):<26} | "
              f"{var[:22]:<22} | {name[:44]}")

    # ---- ออกไฟล์ให้ตรวจ ----
    wb = Workbook()
    ws = wb.active
    ws.title = "ผลจับคู่ด้วยกฎ"
    cols = [("ยอดขาย", 13), ("บรรทัด", 9), ("ร้าน", 7), ("กฎที่ใช้", 24),
            ("SKU ที่ได้", 15), ("Model ที่ได้", 26), ("ผลพิเศษ", 18),
            ("Variation", 30), ("Product Name", 66), ("✍ ถูกไหม", 12)]
    for i, (n2, w) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i, value=n2)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="C55A11" if n2.startswith("✍") else "1F4E79")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    for n2, (rev, lines, shops, rid, sku, model, verdict, var, name) in enumerate(detail, start=2):
        for i, v in enumerate([rev, lines, shops, rid, sku, model, verdict, var, name, ""], start=1):
            c = ws.cell(row=n2, column=i, value=v)
            if i == 1:
                c.number_format = "#,##0"
            if i in (5, 6):
                c.number_format = "@"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(detail)+1}"

    out = PROJECT_ROOT / "output" / "_รีวิว_SKU" / "ผลลองจับคู่ด้วยกฎที่ไกด์ไว้.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"\n✅ {out.relative_to(PROJECT_ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
