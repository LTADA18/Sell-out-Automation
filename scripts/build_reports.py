r"""สร้างรายงานย้อนหลัง 1 ม.ค. – 31 ก.ค. 2026 จากไฟล์ที่ดึงมาแล้ว

ออก 2 ชุด:
  1. ไฟล์รวม 1 ไฟล์  — ทุกร้านทุกแพลตฟอร์มอยู่ชีทเดียว
                       (ถ้าเกิน 1,048,576 แถว แตกเป็น ALL_2, ALL_3 ... ต่อกันไป)
  2. ไฟล์รายแบรนด์   — แบรนด์ละไฟล์ 4 ชีท: รวม / Lazada / TikTok / Shopee

การจับกลุ่มแบรนด์อ่านจาก config/brands.yaml (ประกาศไว้ชัดเจน ไม่เดาจากชื่อ)

อ่านจากไฟล์ที่ export ไว้แล้วทั้ง 2 ที่ ซึ่ง normalize เป็น schema เดียวกันแล้ว
จึงไม่ต้องแตะ adapter หรือดึงข้อมูลใหม่

    python scripts\build_reports.py
"""
from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from pathlib import Path

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

SOURCES = [
    PROJECT_ROOT / "output" / "_shopee_backfill_2026h1" / "2026-01_ถึง_2026-07",
    PROJECT_ROOT / "output" / "_backfill_2026h1_all" / "2026-01_ถึง_2026-07",
]
OUT = PROJECT_ROOT / "output" / "_report_2026h1"
PERIOD = "2026-01-01_ถึง_2026-07-31"

# Excel รับได้ 1,048,576 แถว/ชีท เผื่อหัวตารางกับกันชนไว้
MAX_ROWS = 1_000_000

# ⚠️ 3 คอลัมน์นี้ต้องเป็น text เสมอ — เลข 19 หลักของ TikTok ถ้าหลุดเป็นตัวเลข
#    Excel จะปัดหลักท้ายทิ้งแบบเงียบ ๆ (กฎข้อ 2 ของโปรเจกต์)
TEXT_COLS = {"order_id", "sku", "tracking_no"}
PLATFORMS = ["lazada", "tiktok", "shopee"]
SHEET_OF = {"lazada": "Lazada", "tiktok": "TikTok", "shopee": "Shopee"}


def read_orders(path: Path) -> tuple[list[str], list[list]]:
    """อ่านเฉพาะชีท Orders — ไฟล์มีชีทรายวันอีก 200+ ชีทที่ไม่ต้องใช้"""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Orders"] if "Orders" in wb.sheetnames else wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = [str(c) if c is not None else "" for c in next(it)]
        rows = [list(r) for r in it if any(v is not None for v in r)]
        return header, rows
    finally:
        wb.close()


def write_sheet(ws, header: list[str], rows: list[list], text_idx: set[int]) -> None:
    ws.append(header)
    for r in rows:
        cells = []
        for i, v in enumerate(r):
            c = WriteOnlyCell(ws, value=v)
            if i in text_idx:
                c.number_format = "@"
            cells.append(c)
        ws.append(cells)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--only-brands", action="store_true",
                    help="ทำเฉพาะไฟล์รายแบรนด์ ไม่สร้างไฟล์รวมใหม่ "
                         "(ใช้ตอนแก้การจับกลุ่มใน brands.yaml — ไฟล์รวมไม่เปลี่ยน "
                         "เพราะทุกร้านอยู่ชีทเดียวอยู่แล้ว ไม่ได้แบ่งตามแบรนด์)")
    args = ap.parse_args()

    OUT.mkdir(parents=True, exist_ok=True)

    header: list[str] | None = None
    by_shop: dict[str, list[list]] = defaultdict(list)
    files = [p for d in SOURCES if d.exists() for p in sorted(d.glob("*.xlsx"))]
    if not files:
        print("ไม่พบไฟล์ต้นทางเลย — ยังไม่ได้ดึงข้อมูลหรือ path เปลี่ยน")
        return 1

    print(f"=== อ่านไฟล์ต้นทาง {len(files)} ไฟล์ ===")
    for p in files:
        h, rows = read_orders(p)
        if header is None:
            header = h
        elif h != header:
            print(f"  ⚠️ {p.name}: คอลัมน์ไม่ตรงกับไฟล์แรก ข้ามไฟล์นี้")
            continue
        i_shop = header.index("shop_id")
        for r in rows:
            by_shop[str(r[i_shop])].append(r)
        print(f"  {p.name[:58]:<60} {len(rows):>8,} แถว")

    assert header is not None
    text_idx = {i for i, c in enumerate(header) if c in TEXT_COLS}
    i_plat = header.index("platform")

    total = sum(len(v) for v in by_shop.values())
    print(f"\nรวมทั้งหมด {total:,} แถว จาก {len(by_shop)} ร้าน\n")

    # ── ไฟล์รวมทุกร้าน ────────────────────────────────────────
    p_all = OUT / f"ทุกร้านทุกแพลตฟอร์ม_{PERIOD}.xlsx"
    if args.only_brands and p_all.exists():
        print(f"ข้ามไฟล์รวม — ใช้ของเดิม {p_all.name} "
              f"({p_all.stat().st_size / 1024 / 1024:.1f} MB)\n")
    else:
        all_rows = [r for rows in by_shop.values() for r in rows]
        wb = Workbook(write_only=True)
        for i in range(0, max(len(all_rows), 1), MAX_ROWS):
            chunk = all_rows[i:i + MAX_ROWS]
            name = "ALL" if i == 0 else f"ALL_{i // MAX_ROWS + 1}"
            write_sheet(wb.create_sheet(name), header, chunk, text_idx)
            print(f"  ชีท {name}: {len(chunk):,} แถว")
        wb.save(p_all)
        print(f"✅ {p_all.name}  ({p_all.stat().st_size / 1024 / 1024:.1f} MB)\n")

    # ── ไฟล์รายแบรนด์ ────────────────────────────────────────
    brands = yaml.safe_load((PROJECT_ROOT / "config" / "brands.yaml")
                            .read_text(encoding="utf-8"))["brands"]
    print(f"=== ไฟล์รายแบรนด์ {len(brands)} ไฟล์ ===")
    keep = {p_all.name}
    for b in brands:
        rows = [r for sid in b["shops"] for r in by_shop.get(sid, [])]
        wb = Workbook(write_only=True)

        # ชีทรวมของแบรนด์ต้องมาก่อน แล้วค่อยแยกรายแพลตฟอร์ม
        write_sheet(wb.create_sheet("รวมทุกแพลตฟอร์ม"), header, rows, text_idx)
        counts = []
        for plat in PLATFORMS:
            sub = [r for r in rows if str(r[i_plat]).lower() == plat]
            ws = wb.create_sheet(SHEET_OF[plat])
            if sub:
                write_sheet(ws, header, sub, text_idx)
            else:
                # ต้องมีครบ 4 ชีทตามที่เจ้าของงานกำหนด แม้แบรนด์นี้ไม่ได้ขายบนเจ้านั้น
                ws.append(header)
                ws.append([f"ไม่มีข้อมูล — แบรนด์นี้ไม่ได้ขายบน {SHEET_OF[plat]}"])
            counts.append(f"{SHEET_OF[plat]} {len(sub):,}")

        safe = "".join(ch for ch in b["name"] if ch not in '\\/:*?"<>|').strip()
        path = OUT / f"{safe}_{PERIOD}.xlsx"
        wb.save(path)
        keep.add(path.name)
        print(f"  {b['name'][:24]:<26} รวม {len(rows):>7,} · {' · '.join(counts)}")

    # ลบไฟล์แบรนด์เก่าที่ไม่ตรงกับ brands.yaml ปัจจุบัน
    # ถ้าไม่ลบ พอยุบ 2 แบรนด์เป็นแบรนด์เดียว ไฟล์เดิมทั้ง 2 จะค้างอยู่
    # เจ้าของงานจะเห็นไฟล์เกินและมีข้อมูลซ้ำซ้อนโดยไม่รู้ว่าอันไหนของจริง
    stale = [f for f in OUT.glob("*.xlsx") if f.name not in keep]
    for f in stale:
        f.unlink()
        print(f"  🗑  ลบไฟล์เก่าที่ไม่ตรงกับการจับกลุ่มปัจจุบัน: {f.name}")

    print(f"\nไฟล์ทั้งหมดอยู่ที่ {OUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
