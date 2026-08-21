r"""ดูว่าไฟล์ Export ของ Lazada แต่ละไฟล์ครอบคลุมช่วงวันไหน

    .\.venv\Scripts\python.exe scripts\inspect_lazada_dumps.py output\_manual_downloads\lazada_02_ready_*.xlsx

⚠️ ไฟล์ Lazada ไม่ประกาศ dimension ต้อง reset_dimensions() ก่อน
   ไม่งั้น openpyxl เห็นแค่คอลัมน์เดียว (กฎที่เจอมาแล้วใน CLAUDE.md)
"""
from __future__ import annotations

import glob
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

DATE_KEYS = ("createTime", "orderCreateTime", "วันที่สั่งซื้อ", "created_at")


def look(p: Path) -> None:
    wb = load_workbook(p, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        ws.reset_dimensions()
        it = ws.iter_rows(values_only=True)
        hdr = [str(c).strip() if c is not None else "" for c in next(it)]

        col = None
        for k in DATE_KEYS:
            if k in hdr:
                col = hdr.index(k)
                break
        if col is None:
            for i, h in enumerate(hdr):
                if "time" in h.lower() or "date" in h.lower():
                    col = i
                    break
        oid = hdr.index("orderNumber") if "orderNumber" in hdr else None

        n = 0
        months: Counter = Counter()
        orders = set()
        lo = hi = None
        for r in it:
            if not any(v is not None for v in r):
                continue
            n += 1
            if col is not None and col < len(r) and r[col]:
                d = str(r[col])[:10]
                if len(d) == 10 and d[4] == "-":
                    months[d[:7]] += 1
                    lo = d if lo is None or d < lo else lo
                    hi = d if hi is None or d > hi else hi
            if oid is not None and oid < len(r) and r[oid]:
                orders.add(str(r[oid]))

        print(f"\n📄 {p.name}  ({p.stat().st_size/1024/1024:,.1f} MB)")
        print(f"   คอลัมน์ {len(hdr)} · บรรทัด {n:,} · ออเดอร์ไม่ซ้ำ {len(orders):,}")
        print(f"   ช่วงวัน {lo} ถึง {hi}")
        if months:
            line = "  ".join(f"{m}:{c:,}" for m, c in sorted(months.items()))
            print(f"   รายเดือน {line}")
    finally:
        wb.close()


def main() -> int:
    pats = sys.argv[1:] or ["output/_manual_downloads/lazada_02_ready_*.xlsx"]
    files: list[Path] = []
    for pat in pats:
        files.extend(Path(f) for f in glob.glob(pat))
    if not files:
        print("ไม่พบไฟล์")
        return 1
    for p in sorted(files):
        try:
            look(p)
        except Exception as exc:                         # noqa: BLE001
            print(f"\n📄 {p.name} — อ่านไม่ได้: {type(exc).__name__}: {str(exc)[:90]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
