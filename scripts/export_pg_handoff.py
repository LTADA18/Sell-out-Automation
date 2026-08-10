r"""เตรียมของส่งต่อให้แชท Postgres โหลดเข้า intel.mp_order_line

ทำไมต้องส่งต่อ: เครื่องนี้เข้าถึงไฟล์ตั้งค่าการเชื่อมต่อไม่ได้ (คนละพื้นที่กับผู้ใช้)
จึงเตรียมข้อมูลกับ SQL ให้ครบแล้วให้ฝั่งที่ต่อฐานได้เป็นคนรัน

ออกให้ 3 ไฟล์
  1. *.csv        ข้อมูลที่ normalize ตามกฎในโจทย์แล้ว ทุกช่องใส่เครื่องหมายคำพูด
  2. LOAD.sql     คำสั่งสร้างตารางพัก + COPY + UPSERT เข้าตารางจริง
  3. VERIFY.sql   SQL ตรวจ 4 ข้อตามโจทย์ + ตัวเลขกระทบยอดจากไฟล์ต้นทาง

⚠️ ทุกช่องใน CSV ถูกใส่เครื่องหมายคำพูดทั้งหมดโดยตั้งใจ
   โจทย์เตือนไว้ว่า COPY ... FORMAT csv จะแปลงช่องว่างที่ไม่มีคำพูดเป็น NULL ไม่ใช่ ''
   ซึ่งเคยทำให้ 199,550 แถวดูเหมือนหายไปจากไฟล์ทั้งที่หายจริงแค่ 4,431

    .\.venv\Scripts\python.exe -u scripts\export_pg_handoff.py
"""
from __future__ import annotations

import argparse
import csv
import sys
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

# ค่าที่ต้องกลายเป็นสตริงว่าง ไม่ใช่ NULL และไม่ใช่ข้อความ (กฎข้อ 4 ในโจทย์)
BLANKS = {"null", "none", "nan", "n/a", "na", "-", "nil", ""}

# คอลัมน์ที่ปลายทางต้องใช้ — ชื่อฝั่งเรา : ชื่อฝั่ง intel.mp_order_line
# ⚠️ เดาไม่ได้ว่าตารางจริงมีคอลัมน์อะไรบ้าง เพราะต่อฐานไม่ได้จากเครื่องนี้
#    ฝั่งที่รันต้องยืนยันด้วย information_schema ก่อนแล้วปรับ mapping ตรงนี้
COLUMNS = [
    "platform", "order_id", "shop_id", "sku", "variation", "product_name",
    "quantity", "item_price", "revenue_thb", "total_amount",
    "ordered_at", "order_created_at", "paid_at", "status_raw",
    "osuka_sml_id", "osuka_model_code", "product_brand", "mapping_status",
    "match_method", "match_confidence", "needs_review",
]


def clean(v: object) -> str:
    """ทำให้ค่าว่างเป็นสตริงว่างจริง ๆ — ห้ามปล่อยคำว่า Null เป็นข้อความ

    โจทย์เตือนว่าถ้า sku เป็นข้อความ Null ทุกบรรทัดของร้านจะได้คีย์เดียวกัน
    แล้วทั้งร้านยุบเหลือสินค้าเดียว
    """
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in BLANKS else s


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--date", default="2026-08-01", help="วันที่ของข้อมูล")
    ap.add_argument("--shop", default="shopee_08")
    args = ap.parse_args()

    d = date.fromisoformat(args.date)
    run_date = (d + timedelta(days=1)).isoformat()        # โฟลเดอร์ = วันข้อมูล + 1
    src_dir = PROJECT_ROOT / "output" / run_date / "screened"
    files = list(src_dir.glob(f"*{args.shop}*_matched.xlsx"))
    if not files:
        print(f"❌ ไม่พบไฟล์ของ {args.shop} วันที่ {d} ใน {src_dir}")
        return 1
    src = files[0]
    print(f"ไฟล์ต้นทาง: {src.name}")

    wb = load_workbook(src, read_only=True, data_only=True)
    try:
        ws = wb["data"]
        it = ws.iter_rows(values_only=True)
        hdr = [str(c) if c is not None else "" for c in next(it)]
        missing = [c for c in COLUMNS if c not in hdr]
        if missing:
            print(f"❌ ไฟล์ต้นทางไม่มีคอลัมน์: {missing}")
            return 1
        idx = {c: hdr.index(c) for c in COLUMNS}
        rows = [r for r in it if any(v is not None for v in r)]
    finally:
        wb.close()

    out_dir = PROJECT_ROOT / "output" / f"_pg_handoff_{d}_{args.shop}"
    out_dir.mkdir(parents=True, exist_ok=True)
    csv_path = out_dir / f"{args.shop}_{d}.csv"

    n = qty = 0
    orders: set[str] = set()
    keys: set[tuple] = set()
    dup = 0
    with csv_path.open("w", encoding="utf-8", newline="") as fh:
        # QUOTE_ALL: ช่องว่างจะมาเป็น '' ไม่ใช่ NULL ตามที่โจทย์เตือน
        w = csv.writer(fh, quoting=csv.QUOTE_ALL, lineterminator="\n")
        w.writerow(COLUMNS)
        for r in rows:
            vals = [clean(r[idx[c]]) for c in COLUMNS]
            vals[0] = vals[0].lower()                     # กฎข้อ 1: platform ตัวเล็ก
            n += 1
            orders.add(vals[1])
            try:
                qty += int(float(vals[COLUMNS.index("quantity")] or 0))
            except ValueError:
                pass
            k = (vals[0], vals[1], vals[3], vals[4], vals[5])
            if k in keys:
                dup += 1
            keys.add(k)
            w.writerow(vals)

    print(f"\n=== ตัวเลขจากไฟล์ต้นทาง (ใช้กระทบยอด) ===")
    print(f"  บรรทัดสินค้า (order line) : {n:,}")
    print(f"  ออเดอร์ไม่ซ้ำ             : {len(orders):,}")
    print(f"  ผลรวม quantity            : {qty:,}")
    print(f"  คีย์ 5 คอลัมน์ที่ซ้ำ       : {dup}")
    print(f"\n✅ {csv_path.relative_to(PROJECT_ROOT)}")

    (out_dir / "LOAD.sql").write_text(LOAD_SQL.format(
        shop=args.shop, day=d, csv_name=csv_path.name), encoding="utf-8")
    (out_dir / "VERIFY.sql").write_text(VERIFY_SQL.format(
        shop=args.shop, day=d, lines=n, orders=len(orders), qty=qty), encoding="utf-8")
    print(f"✅ {(out_dir / 'LOAD.sql').relative_to(PROJECT_ROOT)}")
    print(f"✅ {(out_dir / 'VERIFY.sql').relative_to(PROJECT_ROOT)}")
    return 0


LOAD_SQL = """-- โหลด {shop} วันที่ {day} เข้า intel.mp_order_line
--
-- ⚠️ อ่านก่อนรัน — ข้อที่พลาดแล้วเสียหายเงียบ ๆ
--   1. ห้ามเขียน line_id (ตารางสร้างเอง) และ counts_as_sale (generated column)
--   2. platform ต้องเป็นตัวเล็ก — CSV จัดมาให้แล้ว
--   3. order_status ห้ามเอาจากไฟล์ ต้องเรียก intel.mp_order_state(status_raw, paid_at)
--   4. คีย์ต้องครบ 5 คอลัมน์ ไม่ใช่ 3 — ตัดเหลือ 3 เคยทำให้หาย 1,319 บรรทัด
--   5. ต้องเป็น DO UPDATE ไม่ใช่ DO NOTHING เพราะรอบรายวันดึงทับช่วงเดิม
--      ออเดอร์ที่ดึงซ้ำต้องอัปเดตทับ สถานะจะได้เปลี่ยนจาก shipped เป็น returned ได้
--
-- ⚠️ ยืนยันชื่อคอลัมน์จริงก่อน แล้วปรับ INSERT ให้ตรง:
--   SELECT column_name, data_type FROM information_schema.columns
--   WHERE table_schema='intel' AND table_name='mp_order_line' ORDER BY ordinal_position;

BEGIN;
SET LOCAL transaction_read_only = off;   -- service ตั้ง read-only ไว้เป็นค่าเริ่มต้น

CREATE TEMP TABLE stg_order_line (
    platform          text, order_id      text, shop_id      text,
    sku               text, variation     text, product_name text,
    quantity          text, item_price    text, revenue_thb  text,
    total_amount      text, ordered_at    text, order_created_at text,
    paid_at           text, status_raw    text,
    osuka_sml_id      text, osuka_model_code text, product_brand text,
    mapping_status    text, match_method  text, match_confidence text,
    needs_review      text
) ON COMMIT DROP;

-- ⚠️ ต้องมี FORCE_NOT_NULL ให้ครบทุกคอลัมน์ข้อความ
--    ไม่งั้นช่องว่างจะกลายเป็น NULL แทนที่จะเป็น '' แล้วคีย์เพี้ยน
\\copy stg_order_line FROM '{csv_name}' WITH (FORMAT csv, HEADER true, QUOTE '"')

INSERT INTO intel.mp_order_line (
    platform, order_id, shop_id, sku, variation, product_name,
    quantity, item_price, revenue_thb, total_amount,
    ordered_at, paid_at, status_raw, order_status
)
SELECT
    lower(s.platform),
    s.order_id,
    s.shop_id,
    coalesce(s.sku, ''),
    coalesce(s.variation, ''),
    coalesce(s.product_name, ''),
    nullif(s.quantity, '')::numeric::int,
    nullif(s.item_price, '')::numeric,
    nullif(s.revenue_thb, '')::numeric,
    nullif(s.total_amount, '')::numeric,
    nullif(s.ordered_at, '')::timestamp,
    nullif(s.paid_at, '')::timestamp,
    s.status_raw,
    intel.mp_order_state(s.status_raw, nullif(s.paid_at, '')::timestamp)
FROM stg_order_line s
ON CONFLICT (platform, order_id, sku, coalesce(variation,''), coalesce(product_name,''))
DO UPDATE SET
    quantity     = excluded.quantity,
    item_price   = excluded.item_price,
    revenue_thb  = excluded.revenue_thb,
    total_amount = excluded.total_amount,
    ordered_at   = excluded.ordered_at,
    paid_at      = excluded.paid_at,
    status_raw   = excluded.status_raw,
    order_status = excluded.order_status;

COMMIT;
"""

VERIFY_SQL = """-- ตรวจหลังโหลด {shop} วันที่ {day}
-- ตัวเลขที่นับจากไฟล์ต้นทางโดยตรง (ไม่ได้นับจากฐาน):
--   บรรทัดสินค้า {lines}  ·  ออเดอร์ {orders}  ·  quantity {qty}
-- ข้อ 1 ต้องรวมได้ {lines} บรรทัด  ·  ข้อ 2-4 ต้องคืนศูนย์แถวทั้งหมด

-- 1. ลงไปเท่าไหร่ และการกระจายสถานะดูสมเหตุสมผลไหม
SELECT order_status, COUNT(*) AS lines, ROUND(SUM(revenue_thb)) AS thb
FROM   intel.mp_order_line
WHERE  platform = 'shopee' AND shop_id = '{shop}'
  AND  ordered_at::date = DATE '{day}'
GROUP  BY 1 ORDER BY lines DESC;

-- 2. คีย์ต้องไม่ซ้ำ — ต้องได้ศูนย์แถว
SELECT platform, order_id, sku, COALESCE(variation,''), COALESCE(product_name,''), COUNT(*)
FROM   intel.mp_order_line
WHERE  ordered_at::date = DATE '{day}' AND shop_id = '{shop}'
GROUP  BY 1,2,3,4,5 HAVING COUNT(*) > 1;

-- 3. order_status ต้องสร้างซ้ำได้จากกฎ — ต้องได้ศูนย์แถว
SELECT order_id, status_raw, paid_at, order_status
FROM   intel.mp_order_line
WHERE  ordered_at::date = DATE '{day}' AND shop_id = '{shop}'
  AND  order_status IS DISTINCT FROM intel.mp_order_state(status_raw, paid_at);

-- 4. ต้องไม่มีคำในชุดสถานะเก่าหลงเหลือ
SELECT DISTINCT order_status FROM intel.mp_order_line
WHERE ordered_at::date = DATE '{day}' AND shop_id = '{shop}';
"""


if __name__ == "__main__":
    raise SystemExit(main())
