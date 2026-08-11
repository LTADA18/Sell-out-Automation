r"""รวมไฟล์ Excel ที่ดึงมาทั้งหมดไว้ที่เดียว ใช้เป็นคลังข้อมูลอ้างอิง

⚠️ จุดที่สับสนแล้วหยิบไฟล์ผิดวัน: โฟลเดอร์ `output/2026-08-02/` เก็บข้อมูลของ
   **วันที่ 1 ส.ค.** เพราะรอบรายวันดึงข้อมูลของ "เมื่อวาน" คลังนี้จึงตั้งชื่อโฟลเดอร์
   ตามวันของข้อมูลจริง อ่านจาก `ordered_at` ในไฟล์ ไม่ได้เดาจากชื่อโฟลเดอร์

เก็บเฉพาะไฟล์สกรีน 63 คอลัมน์ (`*_matched.xlsx`) เพราะเป็นตัวที่ใช้งานจริง
ไฟล์รวมช่วงยาวไม่ก๊อปมา (ใหญ่หลายร้อย MB) แต่ทำดัชนีชี้ไว้ให้ใน README

    .\.venv\Scripts\python.exe -u scripts\build_archive.py
    .\.venv\Scripts\python.exe -u scripts\build_archive.py --only 2026-08-11
"""
from __future__ import annotations

import argparse
import shutil
import sys
from collections import Counter, defaultdict
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

OUTPUT = PROJECT_ROOT / "output"
ARCHIVE = OUTPUT / "_คลังข้อมูล"


def data_date(xlsx: Path, fallback: str) -> str:
    """วันของข้อมูลจริง อ่านจาก ordered_at แถวแรก ไม่ใช่จากชื่อโฟลเดอร์"""
    try:
        wb = load_workbook(xlsx, read_only=True, data_only=True)
        try:
            ws = wb["data"]
            it = ws.iter_rows(values_only=True)
            hdr = [str(c) if c is not None else "" for c in next(it)]
            if "ordered_at" not in hdr:
                return fallback
            i = hdr.index("ordered_at")
            for r in it:
                if r[i] is not None:
                    return str(r[i])[:10]
        finally:
            wb.close()
    except Exception as exc:                      # ไฟล์เสีย/ชีทหาย — อย่าเงียบ
        print(f"    ⚠️ อ่าน {xlsx.name} ไม่ได้ ({exc.__class__.__name__}) ใช้ {fallback}")
    return fallback


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--only", help="ทำเฉพาะโฟลเดอร์รันวันนี้ เช่น 2026-08-11")
    args = ap.parse_args()

    run_dirs = sorted(
        d for d in OUTPUT.iterdir()
        if d.is_dir() and len(d.name) == 10 and d.name[4] == "-"
        and (args.only is None or d.name == args.only)
    )
    if not run_dirs:
        print("❌ ไม่พบโฟลเดอร์รอบรันที่ตรงเงื่อนไข")
        return 1

    ARCHIVE.mkdir(parents=True, exist_ok=True)
    per_day: dict[str, list[Path]] = defaultdict(list)
    copied = skipped = 0

    for run_dir in run_dirs:
        src_dir = run_dir / "screened"
        if not src_dir.is_dir():
            print(f"  {run_dir.name}  ยังไม่มีโฟลเดอร์ screened ข้าม")
            continue

        files = sorted(src_dir.glob("*_matched.xlsx"))
        if not files:
            print(f"  {run_dir.name}  ไม่มีไฟล์สกรีน ข้าม")
            continue

        # เดาวันจากชื่อโฟลเดอร์ไว้เป็น fallback (รอบรายวันดึงของเมื่อวาน)
        try:
            guess = (date.fromisoformat(run_dir.name) - timedelta(days=1)).isoformat()
        except ValueError:
            guess = run_dir.name

        day = data_date(files[0], guess)
        dest = ARCHIVE / day
        dest.mkdir(parents=True, exist_ok=True)

        for f in files:
            target = dest / f.name
            if target.exists() and target.stat().st_size == f.stat().st_size:
                skipped += 1
            else:
                shutil.copy2(f, target)
                copied += 1
            per_day[day].append(target)

        note = "" if day == guess else f"  ← ชื่อโฟลเดอร์บอก {guess}"
        print(f"  {run_dir.name} → คลัง/{day}  {len(files)} ไฟล์{note}")

    # ---- ดัชนี ----
    lines = [
        "# คลังข้อมูล — ไฟล์สกรีน 63 คอลัมน์",
        "",
        "โฟลเดอร์ตั้งชื่อตาม **วันของข้อมูล** ไม่ใช่วันที่รัน",
        "(รอบรายวันดึงข้อมูลของเมื่อวาน โฟลเดอร์ `output/2026-08-02/` จึงเป็นข้อมูลวันที่ 1 ส.ค.)",
        "",
        "| วันของข้อมูล | ไฟล์ | ร้าน |",
        "|---|---:|---|",
    ]
    for day in sorted(per_day):
        shops = sorted({p.name.split("_")[1] for p in per_day[day] if "_" in p.name})
        lines.append(f"| {day} | {len(per_day[day])} | {len(shops)} ร้าน |")

    lines += [
        "",
        "## ไฟล์รวมช่วงยาว — ไม่ได้ก๊อปมาที่นี่เพราะใหญ่หลายร้อย MB",
        "",
        "| ช่วง | ที่อยู่ |",
        "|---|---|",
        "| 1 ม.ค. – 31 ก.ค. 2026 | `output/_report_2026h1/` |",
        "| 1 – 9 ส.ค. 2026 | `output/_report_2026-08-01_2026-08-09/` |",
        "",
        "## ไฟล์ที่เตรียมส่งเข้า Postgres",
        "",
        "`output/_pg_day_<วันที่>/` — CSV ที่ normalize แล้ว ยุบบรรทัดซ้ำแล้ว",
    ]
    (ARCHIVE / "README.md").write_text("\n".join(lines) + "\n", encoding="utf-8")

    total = sum(len(v) for v in per_day.values())
    size = sum(p.stat().st_size for v in per_day.values() for p in v)
    print(f"\n✅ {ARCHIVE.relative_to(PROJECT_ROOT)}")
    print(f"   {len(per_day)} วัน · {total} ไฟล์ · {size/1024/1024:,.1f} MB")
    print(f"   ก๊อปใหม่ {copied} · มีอยู่แล้วข้าม {skipped}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
