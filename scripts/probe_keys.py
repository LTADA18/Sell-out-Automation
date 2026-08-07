r"""หาว่าอะไรใช้เป็น primary key ได้ — ฝั่งฐานข้อมูลต้องรู้ก่อนออกแบบตาราง

ทดสอบผู้สมัคร 3 แบบกับข้อมูลจริง แล้วบอกว่าซ้ำกี่แถว
เดาไม่ได้ ต้องวัด
"""
from __future__ import annotations

import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

SRC = PROJECT_ROOT / "output" / "_report_2026h1"
FILE = SRC / "ทุกร้านทุกแพลตฟอร์ม_2026-01-01_ถึง_2026-07-31.xlsx"
CAP = 10_000_000                                          # อ่านให้หมด ไม่สุ่ม

CANDIDATES = {
    "order_id": ("order_id",),
    "platform + order_id": ("platform", "order_id"),
    "platform + order_id + sku": ("platform", "order_id", "sku"),
    "shop_id + order_id + sku": ("shop_id", "order_id", "sku"),
}

wb = load_workbook(FILE, read_only=True, data_only=True)
try:
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(it)]
    idx = {c: i for i, c in enumerate(header)}
    counters = {name: Counter() for name in CANDIDATES}
    n = 0
    for row in it:
        if not any(v is not None for v in row):
            continue
        n += 1
        for name, cols in CANDIDATES.items():
            counters[name][tuple(str(row[idx[c]]) for c in cols)] += 1
        if n >= CAP:
            break
finally:
    wb.close()

print(f"อ่าน {n:,} แถว\n")
print(f"{'ผู้สมัครเป็นคีย์':<30} {'ค่าไม่ซ้ำ':>12} {'แถวที่ซ้ำ':>12}   ผล")
print("-" * 74)
for name, ctr in counters.items():
    uniq = len(ctr)
    dup_rows = n - uniq
    verdict = "✅ ใช้เป็นคีย์ได้" if dup_rows == 0 else f"❌ ซ้ำ {dup_rows:,} แถว"
    print(f"{name:<30} {uniq:>12,} {dup_rows:>12,}   {verdict}")

worst = counters["platform + order_id + sku"].most_common(3)
print("\nตัวอย่างที่ซ้ำมากที่สุดของ platform+order_id+sku:")
for k, c in worst:
    print(f"  {c:>3} ครั้ง  {k}")

# sku ว่างได้ ~3% ถ้าออเดอร์เดียวมี 2 บรรทัดที่ sku ว่างทั้งคู่ คีย์จะชนกัน
# ต้องรู้ว่าเกิดจริงไหม ไม่ใช่เดาว่า "น่าจะไม่เกิด"
null_sku = {k: c for k, c in counters["platform + order_id + sku"].items()
            if k[2] in ("Null", "None", "")}
print(f"\nแถวที่ sku ว่าง: {sum(null_sku.values()):,} "
      f"· ในนั้นมีคีย์ที่ชนกัน {sum(c - 1 for c in null_sku.values() if c > 1):,} แถว")
