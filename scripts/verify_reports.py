r"""ตรวจว่ารายงานที่สร้างออกมาถูกต้อง ก่อนส่งให้เจ้าของงาน

ตอบ 4 คำถามที่ถ้าผิดแล้วจะไม่มีอะไรเตือน:
  1. แถวในไฟล์รวม = ผลบวกของทุกร้านจริงไหม (ตกหล่นหรือซ้ำ)
  2. ทุกร้านอยู่ในไฟล์รวมครบ 13 ร้านไหม
  3. ทุกร้านถูกจัดเข้าแบรนด์ครบไหม ไม่มีร้านตกหล่น ไม่มีร้านซ้ำ 2 แบรนด์
  4. order_id ยังเป็น text อยู่ไหม (เลข 19 หลักของ TikTok ถ้าเป็นตัวเลขจะโดนปัดหลักท้าย)

    .\.venv\Scripts\python.exe scripts\verify_reports.py
"""
from __future__ import annotations

import sys
from collections import Counter
from pathlib import Path

import yaml
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

OUT = PROJECT_ROOT / "output" / "_report_2026h1"
PERIOD = "2026-01-01_ถึง_2026-07-31"
ALL_FILE = OUT / f"ทุกร้านทุกแพลตฟอร์ม_{PERIOD}.xlsx"


def scan_all() -> tuple[Counter, Counter, int, int]:
    """นับแถวในไฟล์รวม แยกตามร้านและแพลตฟอร์ม + นับ order_id ที่ไม่ใช่ text"""
    wb = load_workbook(ALL_FILE, read_only=True, data_only=True)
    per_shop: Counter = Counter()
    per_plat: Counter = Counter()
    total = numeric_ids = 0
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            it = ws.iter_rows(values_only=True)
            header = [str(c) if c is not None else "" for c in next(it)]
            i_shop = header.index("shop_id")
            i_plat = header.index("platform")
            i_oid = header.index("order_id")
            for r in it:
                if not any(v is not None for v in r):
                    continue
                total += 1
                per_shop[str(r[i_shop])] += 1
                per_plat[str(r[i_plat]).lower()] += 1
                if not isinstance(r[i_oid], str) and r[i_oid] is not None:
                    numeric_ids += 1
    finally:
        wb.close()
    return per_shop, per_plat, total, numeric_ids


def count_sheet(path: Path, sheet: str) -> int:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            return -1
        it = wb[sheet].iter_rows(values_only=True)
        next(it, None)                                    # ข้ามหัวตาราง
        return sum(1 for r in it if any(v is not None for v in r))
    finally:
        wb.close()


def main() -> int:
    if not ALL_FILE.exists():
        print("ยังไม่มีไฟล์รวม — ยังสร้างไม่เสร็จ")
        return 1

    problems: list[str] = []

    print("=== 1. ไฟล์รวม ===")
    per_shop, per_plat, total, numeric_ids = scan_all()
    print(f"  {total:,} แถว · {len(per_shop)} ร้าน")
    for plat, n in sorted(per_plat.items(), key=lambda kv: -kv[1]):
        print(f"    {plat:<8} {n:>9,}")

    shops = yaml.safe_load((PROJECT_ROOT / "config" / "shops.yaml")
                           .read_text(encoding="utf-8"))["shops"]
    enabled = [s["shop_id"] for s in shops if s.get("enabled", True)]
    missing = [s for s in enabled if s not in per_shop]
    if missing:
        problems.append(f"ไฟล์รวมไม่มีข้อมูลของร้าน {missing}")

    print("\n=== 2. order_id เป็น text ไหม ===")
    if numeric_ids:
        problems.append(f"order_id เป็นตัวเลข {numeric_ids:,} แถว — เสี่ยงโดนปัดหลักท้าย")
        print(f"  ❌ เป็นตัวเลข {numeric_ids:,} แถว")
    else:
        print("  ✅ เป็น text ทุกแถว")

    print("\n=== 3. ไฟล์รายแบรนด์ ===")
    brands = yaml.safe_load((PROJECT_ROOT / "config" / "brands.yaml")
                            .read_text(encoding="utf-8"))["brands"]
    seen: Counter = Counter()
    brand_total = 0
    for b in brands:
        safe = "".join(ch for ch in b["name"] if ch not in '\\/:*?"<>|').strip()
        path = OUT / f"{safe}_{PERIOD}.xlsx"
        want = sum(per_shop.get(sid, 0) for sid in b["shops"])
        for sid in b["shops"]:
            seen[sid] += 1
        if not path.exists():
            problems.append(f"ไม่มีไฟล์ของแบรนด์ {b['name']}")
            print(f"  ⬜ {b['name'][:26]:<28} ยังไม่มีไฟล์")
            continue
        got = count_sheet(path, "รวมทุกแพลตฟอร์ม")
        brand_total += got
        ok = got == want
        if not ok:
            problems.append(f"แบรนด์ {b['name']}: ในไฟล์ {got:,} แถว แต่ควรเป็น {want:,}")
        mb = path.stat().st_size / 1024 / 1024
        print(f"  {'✅' if ok else '❌'} {b['name'][:26]:<28} {got:>8,} แถว "
              f"(ควรได้ {want:,}) {mb:>7.1f} MB")

    print("\n=== 4. ร้านถูกจัดเข้าแบรนด์ครบไหม ===")
    unassigned = [s for s in enabled if seen[s] == 0]
    doubled = [s for s, n in seen.items() if n > 1]
    if unassigned:
        problems.append(f"ร้านที่ไม่ได้อยู่ในแบรนด์ไหนเลย: {unassigned}")
    if doubled:
        problems.append(f"ร้านที่อยู่มากกว่า 1 แบรนด์ (ข้อมูลจะถูกนับซ้ำ): {doubled}")
    if not unassigned and not doubled:
        print(f"  ✅ ครบ {len(enabled)} ร้าน ไม่ซ้ำ ไม่ขาด")

    print(f"\n=== 5. ยอดรวมตรงกันไหม ===")
    print(f"  ไฟล์รวม        {total:,} แถว")
    print(f"  ผลบวกทุกแบรนด์ {brand_total:,} แถว")
    if total != brand_total:
        problems.append(f"ยอดไม่ตรง: ไฟล์รวม {total:,} vs รายแบรนด์ {brand_total:,}")
    else:
        print("  ✅ ตรงกัน")

    print("\n" + "=" * 56)
    if problems:
        print(f"❌ พบปัญหา {len(problems)} ข้อ")
        for p in problems:
            print(f"   · {p}")
        return 1
    print("✅ ตรวจผ่านทั้งหมด — ไฟล์พร้อมส่ง")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
