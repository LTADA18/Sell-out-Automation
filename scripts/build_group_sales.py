r"""ไฟล์ Excel หลักสำหรับดูยอดขาย — 1 ร้าน 1 ชีท รายละเอียดระดับออเดอร์

    .\.venv\Scripts\python.exe -u scripts\build_group_sales.py
    .\.venv\Scripts\python.exe -u scripts\build_group_sales.py --to 2026-08-31

⚠️ ไฟล์นี้คือ "ไฟล์หลัก" ที่เจ้าของงานเปิดดู ชื่อไฟล์จึงคงที่ ไม่มีวันที่ต่อท้าย
   (สั่งไว้ 2026-08-17: "ให้ยึด excel นี้เป็น excel หลักไว้ให้ฉันดูข้อมูลแบบ excel")

โครงไฟล์ — ชีทแรกเป็นสรุปรายเดือน ที่เหลือแยกร้านละชีท 1 บรรทัด = 1 ออเดอร์
   ⚠️ ห้ามยุบทุกร้านมารวมชีทเดียว (สั่งไว้ 2026-08-17)

วิธีอัปเดต: สร้างใหม่ทั้งไฟล์จาก Postgres ทุกครั้ง ไม่ใช่ append ต่อท้ายของเดิม
เหตุผล — ยอดของเดือนที่ผ่านมายัง "ขยับ" ได้อีกหลังปิดเดือน เพราะออเดอร์ที่ยัง
"รอส่ง" ตอนดึงครั้งแรก อาจถูกยกเลิก/คืนของทีหลัง ถ้า append อย่างเดียวเลขเก่า
จะแช่แข็งผิดค้างอยู่ในไฟล์ตลอดไป สร้างใหม่จากฐานทุกครั้งได้เลขที่ตรงเสมอ
ของเดิมถูกสำรองเป็น .bak.xlsx ก่อนเขียนทับทุกครั้ง

⚠️ ไฟล์ใหญ่ (7 แสนออเดอร์) จึงเขียนแบบ write_only — ห้ามย้อนกลับไปแก้เซลล์ที่
   เขียนไปแล้ว ต้องเตรียมค่าให้ครบก่อน append ทีละแถว
"""
from __future__ import annotations

import argparse
import os
import re
import shutil
import subprocess
import sys
import time
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

PSQL = r"C:\Program Files\PostgreSQL\18\bin\psql.exe"
PGPASS = r"C:\Users\tada.p\Postgres\pgpass.conf"
SEP = "\x1f"          # ข้อมูลจริงมี | และ , อยู่ในชื่อสินค้า ใช้ตัวคั่นที่พิมพ์ไม่ได้แทน

# ⚠️ ต้องใช้กติกาเดียวกับที่ฐานใช้ติดธง ไม่งั้น Excel กับ Dashboard จะไม่ตรงกัน
#
#    เทียบชื่อแบรนด์ให้ "เท่ากับ OSUKA พอดี" ห้ามใช้ ILIKE '%osuka%'
#    เพราะจะกวาด 'OSUKAX' เข้ามาด้วย ซึ่งคือเสื้อบอล
#    "OSUKAx100อัน1000อย่าง Limited Edition" ของที่ระลึกจากคอลแลบ ไม่ใช่เครื่องมือ OSUKA
#    (ตรวจเจอ 2026-08-19 ตอนเติมธงย้อนหลัง — ธง yes เดิม 416,719 บรรทัด
#     เป็นแบรนด์ OSUKA เป๊ะทุกแถว ไม่มี OSUKAX เลย และ OSUKAX ถูกติดธง no ไว้แล้ว)
#
#    ตั้งแต่ 2026-08-19 ธงในฐานถูกเติมย้อนหลังครบแล้ว (FIX_osuka_flag_backfill.sql)
#    อ่านจากธงตรง ๆ ได้ แต่ยังคงเทียบชื่อแบรนด์ไว้เพื่อครอบคลุมข้อมูลใหม่
#    ที่โหลดเข้ามาก่อนขั้นตอนติดธงจะทำงาน
#
# ⚠️ เจ้าของงานตัดสินแล้ว 2026-08-19 — ห้ามเปลี่ยนเองโดยไม่ถาม
#      · OKURA  ไม่ใช่ OSUKA (คนละแบรนด์ 2,447 บรรทัด ฿1.25M)
#      · OSUKAX ไม่ใช่ OSUKA (เสื้อบอลของที่ระลึก)
#      · สินค้า OSUKA ที่จับคู่รหัสรุ่นไม่ได้ (ไม่มีในถังข้อมูล) ยังนับเป็น OSUKA
#    ตัวลวง: สินค้า "Okura-Osuka ... รุ่น BSG-135" 5,579 บรรทัดถูกตีเป็นแบรนด์ OSUKA
#    ส่วนรุ่นเดียวกันที่ชื่อขึ้นว่า OKURA เฉย ๆ ไม่ถูกนับ — ดูขัดกันแต่ถูกแล้ว
IS_OSUKA = ("(is_osuka_brand = 'yes' OR "
            "(coalesce(is_osuka_brand,'') = '' AND upper(btrim(product_brand)) = 'OSUKA'))")

MASTER = PROJECT_ROOT / "output" / "_reports" / "ยอดขายรายร้าน_ไฟล์หลัก.xlsx"
START = "2026-01-01"

# กลุ่มร้าน -> shop_id — 10 กลุ่มแรกตามรายชื่อที่เจ้าของงานส่งมา 2026-08-17
# 2 กลุ่มท้ายเป็นร้านที่มีข้อมูลในฐานแต่ไม่ได้อยู่ในรายชื่อ ใส่เพิ่มตามคำสั่ง "ให้ทำทุกร้านเลย"
#
# ⚠️ lazada_01 เป็นร้านเดียวกับ shopee_06 (เอกสตีล) คนละแพลตฟอร์ม จึงอยู่กลุ่มเดียวกัน
GROUPS: list[tuple[str, list[str]]] = [
    ("เอกสตีล",         ["shopee_06", "lazada_01"]),
    # lazada_02 คือหน้าร้าน Lazada ของแบรนด์เดียวกัน — เปิดใช้กลับมา 2026-08-18
    ("TNLTOOLSTORE",   ["shopee_08", "shopee_03", "lazada_02"]),
    ("ร้อยอันพันอย่าง",   ["tiktok_04"]),
    ("toolsdee",       ["tiktok_02"]),
    ("ฝ้ายการช่าง",      ["tiktok_03"]),
    ("PowerS",         ["shopee_04", "tiktok_01"]),
    ("เฮียคิม",          ["shopee_01"]),
    ("เฮียเก๋า",         ["shopee_05", "tiktok_05"]),
    ("DIY Tools",      ["shopee_10"]),
    ("JumboA",         ["shopee_11"]),
    ("Smarttooltech",  ["shopee_02"]),
    ("นาดา",            ["shopee_09"]),
    # ✅ เจ้าของงานยืนยัน 2026-08-19: Metool เป็นร้านเดี่ยว ห้ามรวมกับแบรนด์อื่น
    ("Metool",         ["tiktok_06"]),
]
SHOPS = [s for _, ids in GROUPS for s in ids]
GROUP_OF = {s: g for g, ids in GROUPS for s in ids}

HDR = PatternFill("solid", fgColor="1F4E79")
SUB = PatternFill("solid", fgColor="DDEBF7")
TOT = PatternFill("solid", fgColor="FCE4D6")
WHITE_BOLD = Font(bold=True, color="FFFFFF")

MONEY = "#,##0.00"
INT = "#,##0"
TEXT = "@"            # กฎเหล็กข้อ 2 — order_id/sku/tracking ต้องเป็นข้อความเสมอ

# หัวตารางชีทร้าน: (ชื่อ, กว้าง, number_format)
#
# ⚠️ 1 บรรทัด = 1 สินค้าในออเดอร์ ไม่ใช่ 1 ออเดอร์ (เปลี่ยนเมื่อ 2026-08-17 ตอนเจ้าของงาน
#    ขอ sku / ชื่อสินค้า / ตัวเลือก / แบรนด์ / รหัสรุ่น OSUKA เพิ่ม) ฟิลด์พวกนี้อยู่ระดับ
#    บรรทัดสินค้า ออเดอร์เดียวมีได้หลาย SKU จึงยุบเป็นออเดอร์ละแถวไม่ได้
#    บรรทัดของออเดอร์เดียวกันเรียงติดกันเสมอ (ORDER BY ordered_at, order_id)
ORDER_COLS = [
    ("เลขออเดอร์", 24, TEXT),
    ("วันที่สั่ง", 12, None),
    ("เดือน", 9, None),
    ("สถานะ", 16, None),
    ("นับเป็นยอดขาย", 13, None),
    ("SKU", 22, TEXT),
    ("ชื่อสินค้า", 52, None),
    ("ตัวเลือกสินค้า", 26, None),
    ("แบรนด์", 16, None),
    ("รหัสรุ่น OSUKA", 18, TEXT),
    ("ชิ้น", 8, INT),
    ("ยอดขาย (บาท)", 14, MONEY),
    ("ยอดขาย OSUKA (บาท)", 17, MONEY),
    ("ยอดไม่นับ (บาท)", 14, MONEY),
    ("จังหวัด", 16, None),
    ("ขนส่ง", 18, None),
    ("ช่องทางจ่ายเงิน", 16, None),
]

SUM_COLS = [
    ("กลุ่มร้าน", 18, None), ("ร้าน", 24, None), ("แพลตฟอร์ม", 11, None),
    ("shop_id", 11, None), ("ชีท", 22, None), ("เดือน", 9, None),
    ("ออเดอร์", 10, INT), ("ชิ้น", 10, INT), ("ยอดขาย (บาท)", 16, INT),
    ("ยอดขาย OSUKA (บาท)", 18, INT), ("% OSUKA", 9, None),
    ("ยอดไม่นับ (บาท)", 15, INT),
]


def run_sql(sql: str) -> list[list[str]]:
    env = dict(os.environ, PGPASSFILE=PGPASS, PGCLIENTENCODING="UTF8")
    tmp = PROJECT_ROOT / "output" / "_group_sales.sql"   # ⚠️ path ต้อง ASCII ล้วน psql ถึงอ่านออก
    tmp.parent.mkdir(parents=True, exist_ok=True)
    tmp.write_text(sql, encoding="utf-8")
    p = subprocess.run([PSQL, "service=osuka", "-w", "-A", "-t", "-F", SEP,
                        "-v", "ON_ERROR_STOP=1", "-f", str(tmp)],
                       capture_output=True, env=env)
    out = p.stdout.decode("utf-8", "replace")
    if p.returncode != 0:
        print(out, p.stderr.decode("utf-8", "replace"))
        raise SystemExit("❌ psql ไม่ผ่าน")
    return [ln.split(SEP) for ln in out.splitlines() if ln.strip()]


def num(s: str) -> float:
    try:
        return float(s)
    except (TypeError, ValueError):
        return 0.0


def sheet_name(sid: str, name: str, used: set[str]) -> str:
    """ชื่อชีท Excel ห้ามเกิน 31 ตัว และห้ามมี : \\ / ? * [ ]"""
    base = f"{sid} {re.sub(r'[:\\\\/?*\[\]]', '-', name or '')}".strip()[:31].rstrip()
    nm, i = base, 2
    while nm in used:
        nm = f"{base[:28]}~{i}"
        i += 1
    used.add(nm)
    return nm


def write_header(ws, cols):
    row = []
    for name, _w, _f in cols:
        c = WriteOnlyCell(ws, value=name)
        c.font, c.fill = WHITE_BOLD, HDR
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row.append(c)
    for i, (_n, w, fmt) in enumerate(cols, start=1):
        dim = ws.column_dimensions[get_column_letter(i)]
        dim.width = w
        if fmt:
            dim.number_format = fmt
    ws.append(row)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--from", dest="d_from", default=START)
    ap.add_argument("--to", dest="d_to", default=None, help="ไม่ใส่ = เมื่อวาน")
    args = ap.parse_args()
    d_to = args.d_to or (date.today() - timedelta(days=1)).isoformat()
    t0 = time.time()

    where_date = f"ordered_at >= DATE '{args.d_from}' AND ordered_at < DATE '{d_to}' + 1"

    # ---- ชื่อร้าน/แพลตฟอร์มจริงจากฐาน ไม่เดาเอง ----
    meta = {r[0]: (r[1], r[2]) for r in run_sql(f"""
        SELECT shop_id, max(shop_name), max(platform)
        FROM   intel.mp_order_line
        WHERE  shop_id IN ({",".join(f"'{s}'" for s in SHOPS)}) AND {where_date}
        GROUP  BY shop_id
    """)}
    live = [s for s in SHOPS if s in meta]
    if not live:
        print("❌ ไม่มีข้อมูลในช่วงที่ขอ — ไม่เขียนทับไฟล์หลัก")
        return 1

    wb = Workbook(write_only=True)
    ws_sum = wb.create_sheet("สรุปรายเดือน")
    write_header(ws_sum, SUM_COLS)

    sum_rows: list[list] = []
    grand = [0.0] * 5
    names: set[str] = set()

    for n, sid in enumerate(live, start=1):
        shop_name, platform = meta[sid]
        sname = sheet_name(sid, shop_name, names)
        ws = wb.create_sheet(sname)
        write_header(ws, ORDER_COLS)

        # ---- สรุปรายเดือน: ต้องนับออเดอร์แยกจากตัวรายละเอียด ----
        # ⚠️ นับ DISTINCT order_id ไม่ใช่นับบรรทัด ออเดอร์ที่มี 5 SKU ยังเป็น 1 ออเดอร์
        by_month = {r[0]: [num(x) for x in r[1:]] for r in run_sql(f"""
            SELECT to_char(ordered_at,'YYYY-MM'),
                   count(DISTINCT order_id) FILTER (WHERE counts_as_sale),
                   coalesce(sum(quantity)    FILTER (WHERE counts_as_sale),0),
                   coalesce(sum(revenue_thb) FILTER (WHERE counts_as_sale),0),
                   coalesce(sum(revenue_thb) FILTER (WHERE counts_as_sale
                            AND {IS_OSUKA}),0),
                   coalesce(sum(revenue_thb) FILTER (WHERE NOT counts_as_sale),0)
            FROM   intel.mp_order_line
            WHERE  shop_id = '{sid}' AND {where_date}
            GROUP  BY 1 ORDER BY 1
        """)}
        s = [sum(m[i] for m in by_month.values()) for i in range(5)]

        # ---- รายละเอียด 1 บรรทัด = 1 สินค้าในออเดอร์ ----
        # ⚠️ counts_as_sale อยู่ระดับบรรทัด ออเดอร์เดียวอาจมีบางชิ้นถูกคืนของ
        #    จึงแยกยอด "ที่นับ" กับ "ที่ไม่นับ" คนละคอลัมน์ ห้ามเหมาทั้งออเดอร์ไปข้างใดข้างหนึ่ง
        # กฎเหล็กข้อ 1 — ไม่มีข้อมูลให้เขียน 'Null' ห้ามปล่อยว่างจนแยกไม่ออกว่าไม่มีหรือลืมใส่
        rows = run_sql(f"""
            SELECT order_id, ordered_at::date::text, to_char(ordered_at,'YYYY-MM'),
                   coalesce(nullif(order_status,''),'Null'),
                   counts_as_sale,
                   coalesce(nullif(sku,''),'Null'),
                   coalesce(nullif(product_name,''),'Null'),
                   coalesce(nullif(variation,''),'Null'),
                   coalesce(nullif(product_brand,''),'Null'),
                   coalesce(nullif(osuka_model_code,''),'Null'),
                   coalesce(quantity,0),
                   CASE WHEN counts_as_sale THEN coalesce(revenue_thb,0) ELSE 0 END,
                   CASE WHEN counts_as_sale AND {IS_OSUKA}
                        THEN coalesce(revenue_thb,0) ELSE 0 END,
                   CASE WHEN counts_as_sale THEN 0 ELSE coalesce(revenue_thb,0) END,
                   coalesce(nullif(province,''),'Null'),
                   coalesce(nullif(shipping_carrier,''),'Null'),
                   coalesce(nullif(payment_method,''),'Null')
            FROM   intel.mp_order_line
            WHERE  shop_id = '{sid}' AND {where_date}
            ORDER  BY ordered_at, order_id, sku
        """)

        for (oid, d, ym, status, is_sale, sku, pname, varia, brand, model,
             qty, rev, osk, lost, prov, carrier, pay) in rows:
            # กฎเหล็กข้อ 2 — order_id / sku / รหัสรุ่น ต้องเป็นข้อความ ห้ามหลุดเป็นตัวเลข
            tc = []
            for v in (oid, sku, model):
                c = WriteOnlyCell(ws, value=v)
                c.number_format = TEXT
                tc.append(c)
            ws.append([tc[0], d, ym, status, "ใช่" if is_sale == "t" else "ไม่",
                       tc[1], pname, varia, brand, tc[2],
                       int(num(qty)), num(rev), num(osk), num(lost),
                       prov, carrier, pay])

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(ORDER_COLS))}{len(rows) + 1}"

        for ym in sorted(by_month):
            m = by_month[ym]
            sum_rows.append([GROUP_OF[sid], shop_name, platform, sid, sname, ym,
                             int(m[0]), round(m[1]), round(m[2]), round(m[3]),
                             round(m[3] / m[2] * 100, 1) if m[2] else 0, round(m[4])])
        sum_rows.append(["__SUB__", GROUP_OF[sid], shop_name, platform, sid, sname,
                         int(s[0]), round(s[1]), round(s[2]), round(s[3]),
                         round(s[3] / s[2] * 100, 1) if s[2] else 0, round(s[4])])
        grand = [a + b for a, b in zip(grand, s)]
        print(f"  [{n:2}/{len(live)}] {sname:<32} {len(rows):>8,} บรรทัด "
              f"{int(s[0]):>8,} ออเดอร์ {s[2]:>15,.0f} บาท  ({time.time()-t0:.0f}s)")

    # ---- เติมชีทสรุป ----
    for row in sum_rows:
        if row[0] == "__SUB__":
            body = row[1:]
            cells = []
            for i, v in enumerate(body, start=1):
                c = WriteOnlyCell(ws_sum, value=v)
                c.fill, c.font = SUB, Font(bold=True)
                if i == 11:
                    c.number_format = '0.0"%"'
                cells.append(c)
            cells.insert(5, WriteOnlyCell(ws_sum, value=""))   # ช่องเดือนว่างในแถวรวม
            cells[5].fill = SUB
            ws_sum.append(cells[:len(SUM_COLS)])
        else:
            c = WriteOnlyCell(ws_sum, value=row[10])
            c.number_format = '0.0"%"'
            ws_sum.append(row[:10] + [c, row[11]])

    tot = ["รวมทุกร้าน", "", "", "", "", "", int(grand[0]), round(grand[1]),
           round(grand[2]), round(grand[3]),
           round(grand[3] / grand[2] * 100, 1) if grand[2] else 0, round(grand[4])]
    cells = []
    for i, v in enumerate(tot, start=1):
        c = WriteOnlyCell(ws_sum, value=v)
        c.fill, c.font = TOT, Font(bold=True, size=11)
        if i == 11:
            c.number_format = '0.0"%"'
        cells.append(c)
    ws_sum.append(cells)
    ws_sum.freeze_panes = "A2"
    ws_sum.auto_filter.ref = f"A1:{get_column_letter(len(SUM_COLS))}{len(sum_rows) + 2}"

    # ---- เซฟลงไฟล์ชั่วคราวก่อนแล้วค่อยสลับ ----
    # ⚠️ ห้าม wb.save(MASTER) ตรง ๆ — ถ้าเจ้าของงานเปิดไฟล์หลักค้างไว้ใน Excel
    #    Windows จะล็อกไฟล์ไว้ save พังกลางคัน แล้วไฟล์หลักเดิมพังไปด้วย
    #    เขียนลงไฟล์ชั่วคราวก่อน ของเดิมจึงปลอดภัยเสมอแม้สร้างไม่สำเร็จ
    MASTER.parent.mkdir(parents=True, exist_ok=True)
    tmp = MASTER.with_name(MASTER.stem + ".tmp.xlsx")
    wb.save(tmp)

    out = MASTER
    if MASTER.exists():
        try:
            shutil.copy2(MASTER, MASTER.with_name(MASTER.stem + ".bak.xlsx"))
        except PermissionError:
            pass
    try:
        os.replace(tmp, MASTER)
    except PermissionError:
        out = MASTER.with_name(MASTER.stem + "_ฉบับใหม่.xlsx")
        os.replace(tmp, out)
        print(f"\n⚠️ ไฟล์หลักถูกเปิดค้างอยู่ใน Excel เขียนทับไม่ได้")
        print(f"   ออกให้เป็นไฟล์แยกแทน: {out.name}")
        print(f"   ปิดไฟล์หลักแล้วรันใหม่ ถึงจะเขียนทับได้")

    MASTER_ = out
    mb = MASTER_.stat().st_size / 1024 / 1024
    print(f"\n✅ {MASTER_}")
    print(f"   {len(live)} ชีทร้าน + 1 ชีทสรุป · {mb:,.1f} MB · ใช้เวลา {time.time()-t0:.0f} วินาที")
    print(f"   ช่วง {args.d_from} ถึง {d_to}")
    miss = [s for s in SHOPS if s not in meta]
    if miss:
        print(f"   ⚠️ ไม่มีข้อมูล: {miss}")
    print(f"   ออเดอร์ {grand[0]:>13,.0f}")
    print(f"   ชิ้น    {grand[1]:>13,.0f}")
    print(f"   ยอดขาย  {grand[2]:>13,.0f} บาท")
    print(f"   OSUKA   {grand[3]:>13,.0f} บาท ({grand[3]/grand[2]*100:.1f}%)")
    print(f"   ไม่นับ   {grand[4]:>13,.0f} บาท")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
