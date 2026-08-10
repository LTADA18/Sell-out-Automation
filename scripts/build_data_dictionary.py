r"""สร้าง Data Dictionary เป็นไฟล์ Excel ให้ฝั่งฐานข้อมูล

ตัวเลข % อ่านจาก docs/column_fill_rate.json ซึ่งวัดจากข้อมูลจริง
ไม่ฝังตัวเลขไว้ในสคริปต์ — ถ้าข้อมูลเปลี่ยน รันใหม่แล้วตัวเลขอัปเดตตาม

    .\.venv\Scripts\python.exe -u scripts\build_data_dictionary.py
"""
from __future__ import annotations

import json
import os
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from src.core.models import EXCEL_COLUMNS, OrderStatus     # noqa: E402

OUT = PROJECT_ROOT / "docs" / "Data_Dictionary_คำสั่งซื้อ.xlsx"
FILL_RATE = PROJECT_ROOT / "docs" / "column_fill_rate.json"

# ── Data Dictionary ของระบบสกรีน (คนละโปรเจกต์) ──────────────
#
# ⚠️ อ่านจากไฟล์ของเขาตอน build ไม่คัดลอกคำอธิบายมาเก็บซ้ำในโค้ดนี้
#    ถ้าคัดลอกมา วันหนึ่งเขาแก้ของเขาแล้วของเราจะค้างอยู่กับข้อมูลเก่า
#    กลายเป็น 2 แหล่งความจริงที่ขัดกันเอง ซึ่งเป็นสิ่งที่โปรเจกต์นี้เลี่ยงมาตลอด
#    เขาแก้เมื่อไหร่ รันไฟล์นี้ใหม่แล้วตรงกันทันที
SCREEN_DD = Path(os.environ.get(
    "OSUKA_SKU_DIR", r"C:\Users\tada.p\Clean data\osuka-sku")
) / "docs" / "osuka_sku_matching_data_dictionary_Screening System.xlsx"


def read_screen_dd(sheet: str) -> list[list[str]]:
    """อ่านตาราง Field/Type/Description/Example จาก DD ของระบบสกรีน

    คืน [] ถ้าอ่านไม่ได้ — เอกสารฝั่งเราต้องสร้างได้เสมอ
    ไม่ใช่พังทั้งไฟล์เพราะอีกโปรเจกต์ไม่อยู่ที่เดิม
    """
    # ⚠️ ห้ามกลืน error เงียบ ๆ — ถ้าอ่านไม่ได้ต้องบอกว่าเพราะอะไร
    #    ของเดิมคืน [] เฉย ๆ ทำให้เอกสารออกมาโดยไม่มีชั้นที่ 2 แล้วดูเหมือนสำเร็จ
    if not SCREEN_DD.exists():
        print(f"  ⚠️ ไม่พบไฟล์ DD ของระบบสกรีน: {SCREEN_DD}")
        return []
    try:
        wb = load_workbook(SCREEN_DD, read_only=True, data_only=True)
    except Exception as exc:                             # noqa: BLE001
        print(f"  ⚠️ เปิดไฟล์ DD ของระบบสกรีนไม่ได้: {type(exc).__name__}: {exc}")
        return []
    try:
        if sheet not in wb.sheetnames:
            print(f"  ⚠️ ไม่มีชีท {sheet!r} — ที่มีคือ {wb.sheetnames}")
            return []
        rows = [[str(c).strip() if c is not None else "" for c in r]
                for r in wb[sheet].iter_rows(values_only=True)]
    finally:
        wb.close()

    out: list[list[str]] = []
    started = False
    for r in rows:
        if not started:
            if r and r[0] == "Field":                    # หัวตารางจริงเริ่มตรงนี้
                started = True
            continue
        if not r or not r[0]:
            continue
        out.append((r + ["", "", "", ""])[:4])
    return out

FONT = "Arial"
NAVY = "1F3864"
GREY = "F2F2F2"
WARN = "FFF2CC"
DANGER = "FCE4E4"
OKBG = "E2EFDA"

thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

rates: dict[str, dict[str, float]] = json.loads(FILL_RATE.read_text(encoding="utf-8"))


def pct(col: str, plat: str) -> float | None:
    v = rates.get(plat, {}).get(col)
    return None if v is None else round(v / 100, 3)        # เก็บเป็นเศษส่วน ให้ format เป็น %


# ── นิยามคอลัมน์: (ชนิด, null ได้ไหม, กลุ่ม, คำอธิบาย, หมายเหตุสำคัญ) ──
DEFS: dict[str, tuple[str, str, str, str, str]] = {
    "order_id": ("VARCHAR(64)", "ไม่", "ตัวตน", "เลขออเดอร์จากแพลตฟอร์ม",
                 "ต้องเป็น TEXT เท่านั้น — TikTok ใช้เลข 19 หลัก ถ้ารับเป็นตัวเลขจะโดนปัดหลักท้ายแบบเงียบ ๆ"),
    "platform": ("VARCHAR(16)", "ไม่", "ตัวตน", "แพลตฟอร์ม", "ค่าที่เป็นไปได้: lazada / tiktok / shopee"),
    "shop_id": ("VARCHAR(32)", "ไม่", "ตัวตน", "รหัสหน้าร้านในระบบเรา เช่น shopee_03", ""),
    "shop_name": ("VARCHAR(128)", "ไม่", "ตัวตน", "ชื่อร้านมาตรฐาน",
                  "1 ชื่อร้านมีได้หลาย shop_id — ดูชีท 'รายชื่อร้าน'"),

    "order_created_at": ("DATETIME", "ไม่", "เวลา", "เวลาที่สร้างออเดอร์",
                         "ใช้คอลัมน์นี้แบ่งช่วงเวลาในรายงาน"),
    "order_updated_at": ("DATETIME", "ได้", "เวลา", "เวลาที่แก้ล่าสุด", ""),
    "paid_at": ("DATETIME", "ได้", "เวลา", "เวลาชำระเงิน", "Lazada ไม่มีข้อมูลนี้"),

    "status_raw": ("VARCHAR(64)", "ได้", "สถานะ", "ค่าสถานะดิบจากแพลตฟอร์ม",
                   "คนละคำในแต่ละเจ้า เก็บไว้ย้อนตรวจว่า map ถูกไหม"),
    "order_status": ("VARCHAR(16)", "ไม่", "สถานะ", "สถานะกลางที่แปลงแล้ว",
                     "ดูค่าที่เป็นไปได้ในชีท 'ค่าที่เป็นไปได้'"),
    "payment_method": ("VARCHAR(64)", "ได้", "สถานะ", "วิธีชำระเงิน เช่น COD, SPayLater", ""),

    "sku": ("VARCHAR(128)", "ได้", "สินค้า", "รหัสสินค้าของผู้ขาย",
            "ต้องเป็น TEXT / ว่างได้ประมาณ 3% ซึ่งกระทบการใช้เป็นคีย์"),
    "product_name": ("VARCHAR(512)", "ได้", "สินค้า", "ชื่อสินค้า", ""),
    "variation": ("VARCHAR(256)", "ได้", "สินค้า", "ตัวเลือกสินค้า เช่น สี/ขนาด", ""),
    "quantity": ("INT", "ได้", "สินค้า", "จำนวนชิ้น",
                 "Lazada ไม่มีคอลัมน์นี้ในไฟล์ต้นทาง ระบบยุบ 1 แถวต่อชิ้นให้แล้ว"),

    "item_price": ("DECIMAL(12,2)", "ได้", "เงิน-สินค้า", "ราคาต่อรายการ", ""),
    "item_discount": ("DECIMAL(12,2)", "ได้", "เงิน-สินค้า", "ส่วนลดระดับรายการ",
                      "ว่างทุกแพลตฟอร์ม — ไม่มีในรายงานต้นทาง"),
    "seller_discount": ("DECIMAL(12,2)", "ได้", "เงิน-สินค้า", "ส่วนลดที่ร้านออกเอง",
                        "Lazada มีน้อยมาก (2%)"),
    "platform_discount": ("DECIMAL(12,2)", "ได้", "เงิน-สินค้า", "ส่วนลดที่แพลตฟอร์มออก", ""),

    "shipping_fee": ("DECIMAL(12,2)", "ได้", "ขนส่ง", "ค่าส่ง (ระดับออเดอร์)",
                     "เป็นค่าของทั้งออเดอร์ ซ้ำทุกแถวในออเดอร์เดียวกัน"),
    "shipping_carrier": ("VARCHAR(64)", "ได้", "ขนส่ง", "ผู้ให้บริการขนส่ง", ""),
    "tracking_no": ("VARCHAR(64)", "ได้", "ขนส่ง", "เลขพัสดุ",
                    "ต้องเป็น TEXT / ว่างเมื่อยังไม่ได้จัดส่ง"),

    "commission_fee": ("DECIMAL(12,2)", "ได้", "ค่าธรรมเนียม", "ค่าคอมมิชชั่น",
                       "ยังไม่มีข้อมูล — อยู่ในเมนูการเงินซึ่งยังไม่ได้รับอนุญาตให้เข้าถึง"),
    "transaction_fee": ("DECIMAL(12,2)", "ได้", "ค่าธรรมเนียม", "ค่าธรรมเนียมธุรกรรม",
                        "ยังไม่มีข้อมูล เหตุผลเดียวกับ commission_fee"),
    "service_fee": ("DECIMAL(12,2)", "ได้", "ค่าธรรมเนียม", "ค่าบริการ",
                    "ยังไม่มีข้อมูล เหตุผลเดียวกับ commission_fee"),

    "total_amount": ("DECIMAL(12,2)", "ได้", "เงิน-ออเดอร์", "ยอดที่ลูกค้าจ่าย (ทั้งออเดอร์)",
                     "ห้าม SUM ตรง ๆ จะได้ยอดเกินจริง เพราะซ้ำทุกแถวในออเดอร์ — ดูชีท 'ข้อควรระวัง'"),
    "settlement_amount": ("DECIMAL(12,2)", "ได้", "เงิน-ออเดอร์", "ยอดที่ร้านได้รับจริง",
                          "ยังไม่มีข้อมูล — ห้ามนำไปคำนวณกำไร"),

    "buyer_username": ("VARCHAR(64)", "ได้", "ผู้ซื้อ", "ชื่อผู้ใช้ของผู้ซื้อ (ถูกปกปิดแล้ว)",
                       "PDPA: เก็บแค่ตัวแรกกับตัวท้าย เช่น somchai123 → s********3 / Lazada ไม่มี"),
    "province": ("VARCHAR(64)", "ได้", "ผู้ซื้อ", "จังหวัดผู้รับ",
                 "Lazada ไม่มี — อย่าทำรายงานรายจังหวัดจากข้อมูลทั้งหมด ตัวเลขจะเอียง"),

    "cancel_reason": ("VARCHAR(256)", "ได้", "ยกเลิก/คืน", "เหตุผลการยกเลิก",
                      "ว่างเมื่อออเดอร์ไม่ถูกยกเลิก / Lazada ไม่มี"),
    "return_status": ("VARCHAR(64)", "ได้", "ยกเลิก/คืน", "สถานะการคืนสินค้า", "Lazada ไม่มี"),

    "notes": ("VARCHAR(512)", "ได้", "meta", "เหตุผลที่บางคอลัมน์ว่าง",
              "กฎของระบบคือห้ามสร้างตัวเลขขึ้นเอง ไม่มีข้อมูลจะเว้นว่างแล้วเขียนเหตุผลที่นี่"),
    "fetched_at": ("DATETIME", "ไม่", "meta", "เวลาที่ระบบดึงข้อมูล", "ไม่ใช่เวลาของออเดอร์"),
}

wb = Workbook()


def style_header(ws, row: int, ncol: int) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BOX
    ws.row_dimensions[row].height = 30


def title(ws, text: str, sub: str = "") -> int:
    ws["A1"] = text
    ws["A1"].font = Font(name=FONT, bold=True, size=14, color=NAVY)
    r = 2
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")
        r = 3
    return r + 1


# ══════════ ชีท 1: ภาพรวม ══════════
ws = wb.active
ws.title = "ภาพรวม"
r = title(ws, "Data Dictionary — ข้อมูลคำสั่งซื้อร้านค้าออนไลน์",
          "ทุกตัวเลข % วัดจากข้อมูลจริง 660,181 แถว (ม.ค.–ก.ค. 2026) ไม่ใช่ค่าที่คาดเดา · "
          "ปรับตามคำตอบจากฝั่งฐานข้อมูล 2026-08-10")

META = [
    ("เจ้าของข้อมูล", "ทีม Marketplace"),
    ("แหล่งที่มา", "หลังบ้านร้านค้าของเราเอง — Lazada / TikTok Shop / Shopee Seller Centre"),
    ("ขอบเขต", "13 หน้าร้าน · 9 ชื่อร้าน · 3 แพลตฟอร์ม"),
    ("ความถี่", "รายวัน 08:30 น. ดึงข้อมูลของเมื่อวาน"),
    ("Time zone", "Asia/Bangkok ทุกคอลัมน์เวลา (ไม่มี timezone offset ต่อท้าย)"),
    ("สกุลเงิน", "บาท (THB) ทุกคอลัมน์เงิน"),
    ("รูปแบบส่งมอบ", "Excel .xlsx — sheet ชื่อ Orders"),
    ("จำนวนคอลัมน์", len(EXCEL_COLUMNS)),
    ("Grain (1 แถวคืออะไร)", "1 รายการสินค้าในออเดอร์ (order line) — ไม่ใช่ 1 ออเดอร์"),
    ("การปกปิดข้อมูลส่วนบุคคล", "เปิดใช้ (include_pii = false) — ไม่มีชื่อ-นามสกุล เบอร์โทร ที่อยู่"),
    ("ปลายทางที่ตกลงกัน",
     "Postgres — ส่งเป็น 2 ตาราง order_header + order_line (รอรายละเอียดการเชื่อมต่อ)"),
    ("ช่วงข้อมูลที่ตกลงกัน", "1 ม.ค. 2026 ถึง D-1 (เมื่อวาน) และต่อเนื่องทุกวัน"),
]
for k, v in META:
    ws.cell(row=r, column=1, value=k).font = Font(name=FONT, bold=True, size=10)
    ws.cell(row=r, column=2, value=v).font = Font(name=FONT, size=10)
    ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="center")
    r += 1

r += 1
ws.cell(row=r, column=1, value="สรุปความพร้อมของคอลัมน์").font = Font(
    name=FONT, bold=True, size=11, color=NAVY)
r += 1
head_row = r
for i, h in enumerate(["สถานะ", "จำนวนคอลัมน์"], start=1):
    ws.cell(row=r, column=i, value=h)
style_header(ws, r, 2)
r += 1

# ใช้สูตรนับจากชีท Data Dictionary — ตัวเลขต้องขยับตามเมื่อข้อมูลเปลี่ยน
n = len(EXCEL_COLUMNS)
last = 4 + n - 1
SUMMARY = [
    ("ใช้ได้ทุกแพลตฟอร์ม", f"=COUNTIFS('Data Dictionary'!$J$4:$J${last},\"ครบทุกแถว\")"),
    ("มีบ้างไม่มีบ้าง", f"=COUNTIFS('Data Dictionary'!$J$4:$J${last},\"มีบ้างไม่มีบ้าง\")"),
    ("ขาดบางแพลตฟอร์ม", f"=COUNTIFS('Data Dictionary'!$J$4:$J${last},\"ขาดบางแพลตฟอร์ม\")"),
    ("ยังไม่มีข้อมูลเลย", f"=COUNTIFS('Data Dictionary'!$J$4:$J${last},\"ยังไม่มีข้อมูล\")"),
]
for label, formula in SUMMARY:
    ws.cell(row=r, column=1, value=label).font = Font(name=FONT, size=10)
    c = ws.cell(row=r, column=2, value=formula)
    c.font = Font(name=FONT, size=10)
    c.alignment = Alignment(horizontal="center")
    for col in (1, 2):
        ws.cell(row=r, column=col).border = BOX
    r += 1

ws.cell(row=r, column=1, value="รวม").font = Font(name=FONT, bold=True, size=10)
tot = ws.cell(row=r, column=2, value=f"=SUM(B{head_row + 1}:B{r - 1})")
tot.font = Font(name=FONT, bold=True, size=10)
tot.alignment = Alignment(horizontal="center")
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 76

# ══════════ ชีท 2: Data Dictionary ══════════
ws = wb.create_sheet("Data Dictionary")
ws["A1"] = "Data Dictionary — 32 คอลัมน์ เรียงตามลำดับในไฟล์"
ws["A1"].font = Font(name=FONT, bold=True, size=13, color=NAVY)
ws["A2"] = ("คอลัมน์ % = สัดส่วนแถวที่มีค่าจริง วัดจากข้อมูลจริง · "
            "0% = ว่างทั้งแพลตฟอร์มนั้น ไม่ใช่ข้อมูลหาย")
ws["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")

HEAD = ["#", "คอลัมน์", "ชนิดข้อมูล", "Null ได้", "กลุ่ม", "คำอธิบาย",
        "Lazada", "TikTok", "Shopee", "สถานะ", "หมายเหตุสำคัญ"]
for i, h in enumerate(HEAD, start=1):
    ws.cell(row=3, column=i, value=h)
style_header(ws, 3, len(HEAD))
ws.freeze_panes = "C4"

TEXT_COLS = {"order_id", "sku", "tracking_no"}
row = 4
for i, col in enumerate(EXCEL_COLUMNS, start=1):
    dtype, nullable, group, desc, note = DEFS[col]
    la, ti, sh = pct(col, "lazada"), pct(col, "tiktok"), pct(col, "shopee")
    vals = [v for v in (la, ti, sh) if v is not None]

    if all(v == 0 for v in vals):
        status, bg = "ยังไม่มีข้อมูล", DANGER
    elif any(v == 0 for v in vals):
        status, bg = "ขาดบางแพลตฟอร์ม", WARN
    elif all(v > 0.999 for v in vals):
        status, bg = "ครบทุกแถว", OKBG
    else:
        status, bg = "มีบ้างไม่มีบ้าง", None

    ws.cell(row=row, column=1, value=i)
    c_name = ws.cell(row=row, column=2, value=col)
    c_name.font = Font(name=FONT, bold=True, size=10,
                       color="C00000" if col in TEXT_COLS else "000000")
    ws.cell(row=row, column=3, value=dtype)
    ws.cell(row=row, column=4, value=nullable)
    ws.cell(row=row, column=5, value=group)
    ws.cell(row=row, column=6, value=desc)
    for j, v in enumerate((la, ti, sh), start=7):
        c = ws.cell(row=row, column=j, value=v)
        c.number_format = "0.0%"
        c.alignment = Alignment(horizontal="center")
    c_status = ws.cell(row=row, column=10, value=status)
    c_status.alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=11, value=note)

    for j in range(1, len(HEAD) + 1):
        cell = ws.cell(row=row, column=j)
        cell.border = BOX
        if cell.font.name != FONT or not cell.font.size:
            cell.font = Font(name=FONT, size=10)
        cell.alignment = Alignment(
            wrap_text=(j in (6, 11)), vertical="center",
            horizontal=cell.alignment.horizontal or "left")
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)

    if col in TEXT_COLS:
        c_name.comment = Comment(
            "ต้องเก็บเป็น TEXT เท่านั้น\n"
            "TikTok ใช้เลข 19 หลัก ถ้ารับเป็นตัวเลข\n"
            "ระบบจะปัดหลักท้ายทิ้งโดยไม่มี error",
            "Data Team")
    row += 1

for col, w in zip("ABCDEFGHIJK", (5, 22, 16, 9, 14, 42, 10, 10, 10, 18, 60)):
    ws.column_dimensions[col].width = w

# ══════════ ชีท 2.5: ชั้นที่ 2 — หลังสกรีน SKU (63 คอลัมน์) ══════════
ws = wb.create_sheet("หลังสกรีน 63 คอลัมน์")
screen_rows = read_screen_dd("sku_match_data")
r = title(
    ws, "ชั้นที่ 2 — ไฟล์หลังสกรีน SKU (63 คอลัมน์) = ของที่ส่งมอบจริง",
    f"คำอธิบายอ่านสดจาก DD ของระบบ osuka-sku ตอนสร้างไฟล์นี้ ไม่ได้คัดลอกมาเก็บซ้ำ · "
    f"{'อ่านได้ ' + str(len(screen_rows)) + ' คอลัมน์' if screen_rows else 'อ่านไม่ได้ — ดูหมายเหตุ'}")

if not screen_rows:
    ws.cell(row=r, column=1,
            value=f"อ่าน DD ของระบบสกรีนไม่ได้ที่ {SCREEN_DD}").font = Font(
        name=FONT, size=10, color="C00000")
    ws.cell(row=r + 1, column=1,
            value="ตั้งตัวแปร OSUKA_SKU_DIR ให้ชี้ไปที่โปรเจกต์ osuka-sku แล้วสร้างใหม่").font = Font(
        name=FONT, size=10)
    ws.column_dimensions["A"].width = 100
else:
    ws.cell(row=r, column=1,
            value="63 คอลัมน์ = 32 ของชั้นที่ 1 (ยกมาทั้งดุ้น ไม่ถูกแก้) + 31 ที่ระบบสกรีนเติมให้"
            ).font = Font(name=FONT, size=10, bold=True)
    r += 2
    for i, h in enumerate(["#", "คอลัมน์", "ชนิด", "มาจากชั้นไหน", "คำอธิบาย", "ตัวอย่าง"], start=1):
        ws.cell(row=r, column=i, value=h)
    style_header(ws, r, 6)
    ws.freeze_panes = f"C{r + 1}"
    r += 1

    layer1 = set(EXCEL_COLUMNS)
    n_new = 0
    for i, (fld, typ, desc, ex) in enumerate(screen_rows, start=1):
        name = fld.lstrip("★ ").strip()
        from_l1 = name in layer1
        if not from_l1:
            n_new += 1
        ws.cell(row=r, column=1, value=i)
        c = ws.cell(row=r, column=2, value=fld)
        c.font = Font(name=FONT, bold=True, size=10,
                      color="000000" if from_l1 else "1F6F1F")
        ws.cell(row=r, column=3, value=typ)
        ws.cell(row=r, column=4, value="ชั้น 1 (ดึง)" if from_l1 else "ชั้น 2 (สกรีน)")
        ws.cell(row=r, column=5, value=desc)
        ws.cell(row=r, column=6, value=ex)
        for j in range(1, 7):
            cell = ws.cell(row=r, column=j)
            cell.border = BOX
            cell.alignment = Alignment(wrap_text=(j in (5, 6)), vertical="center")
            if not cell.font.bold:
                cell.font = Font(name=FONT, size=10)
            if not from_l1:
                cell.fill = PatternFill("solid", fgColor=OKBG)
        r += 1

    r += 1
    ws.cell(row=r, column=1,
            value=f"แถวสีเขียว = {n_new} คอลัมน์ที่ระบบสกรีนเติมให้ · "
                  f"แถวขาว = {len(screen_rows) - n_new} คอลัมน์ที่ยกมาจากชั้นที่ 1").font = Font(
        name=FONT, size=9, italic=True, color="1F6F1F")
    for col, w in zip("ABCDEF", (5, 26, 12, 15, 62, 30)):
        ws.column_dimensions[col].width = w

# ══════════ ชีท 2.6: ค่าที่เป็นไปได้ของชั้นที่ 2 ══════════
VOCABS = [
    ("vocab_match_method", "match_method", "ชั้นการจับคู่ L0–L8"),
    ("vocab_mapping_status", "mapping_status_detail", "สถานะการจับคู่แบบละเอียด"),
    ("vocab_review_reason", "review_reason", "รหัสเหตุผลที่ต้องให้คนตรวจ"),
]
ws = wb.create_sheet("ค่าที่เป็นไปได้ ชั้น 2")
r = title(ws, "ค่าที่เป็นไปได้ของคอลัมน์ที่ระบบสกรีนเติม",
          "อ่านสดจาก DD ของ osuka-sku เช่นกัน")
for i, h in enumerate(["คอลัมน์", "ค่า", "ความหมาย"], start=1):
    ws.cell(row=r, column=i, value=h)
style_header(ws, r, 3)
r += 1
for sheet_name, col_name, _note in VOCABS:
    vals = read_screen_dd(sheet_name)
    for fld, _typ, desc, _ex in vals:
        ws.cell(row=r, column=1, value=col_name)
        ws.cell(row=r, column=2, value=fld).font = Font(name=FONT, bold=True, size=10)
        ws.cell(row=r, column=3, value=desc)
        for j in range(1, 4):
            cell = ws.cell(row=r, column=j)
            cell.border = BOX
            cell.alignment = Alignment(wrap_text=(j == 3), vertical="center")
            if not cell.font.bold:
                cell.font = Font(name=FONT, size=10)
        r += 1
for col, w in zip("ABC", (24, 34, 74)):
    ws.column_dimensions[col].width = w

# ══════════ ชีท 3: ข้อควรระวัง ══════════
ws = wb.create_sheet("ข้อควรระวัง")
r = title(ws, "ข้อควรระวัง — อ่านก่อนออกแบบตารางและก่อนเขียนรายงาน",
          "ทุกข้อเป็นเรื่องที่ถ้าพลาดแล้วตัวเลขจะผิดโดยไม่มีอะไรเตือน")

for i, h in enumerate(["#", "หัวข้อ", "รายละเอียด", "สิ่งที่ต้องทำ"], start=1):
    ws.cell(row=r, column=i, value=h)
style_header(ws, r, 4)
hr = r
r += 1

RISKS = [
    ("1 แถว = 1 สินค้า ไม่ใช่ 1 ออเดอร์",
     "ออเดอร์ที่ซื้อ 3 สินค้าจะมี 3 แถว ใช้ order_id เดียวกัน "
     "คอลัมน์ระดับออเดอร์ (total_amount, shipping_fee, payment_method, province, buyer_username) "
     "จะซ้ำในทุกแถวของออเดอร์นั้น",
     "ห้าม SUM(total_amount) ตรง ๆ จะได้ยอดเกินจริงตามจำนวนสินค้า "
     "ต้องยุบเป็นระดับออเดอร์ก่อน หรือใช้ค่าเฉพาะแถวแรกของแต่ละ order_id"),
    ("order_id ต้องเป็น TEXT",
     "TikTok ใช้เลข 19 หลัก ซึ่งเกินช่วงของจำนวนเต็ม 64 บิตที่ระบบส่วนใหญ่ใช้ปลอดภัย "
     "ถ้ารับเป็นตัวเลข หลักท้ายจะถูกปัดทิ้งโดยไม่มี error",
     "ประกาศเป็น VARCHAR ทั้ง order_id, sku, tracking_no และห้ามให้ตัวนำเข้าเดาชนิดเอง"),
    ("ข้อความ 'Null' ไม่ใช่ NULL จริง",
     "ในไฟล์ Excel ค่าว่างบางช่องถูกเขียนเป็นข้อความ Null (4 ตัวอักษร)",
     "ตอน import ต้องแปลงข้อความ 'Null' เป็น NULL จริง ไม่งั้น IS NULL ใช้ไม่ได้"),
    ("1 ชื่อร้าน มีได้หลาย shop_id",
     "ร้านเดียวกันขายหลายแพลตฟอร์ม ระบบแปลงชื่อให้เป็นชื่อมาตรฐานเดียวกันแล้ว "
     "เช่น powerstool (TikTok) กับ Powerstools (Shopee) กลายเป็น Powerstools ทั้งคู่",
     "GROUP BY shop_name = ยอดของร้านจริง / GROUP BY shop_id = ยอดของแต่ละหน้าร้าน "
     "เลือกให้ตรงกับคำถามที่จะตอบ"),
    ("Lazada ไม่มีจังหวัด",
     "province ถูกแพลตฟอร์มปิดบังมาตั้งแต่ต้นทาง ดึงมาไม่ได้",
     "อย่าทำรายงานยอดขายรายจังหวัดจากข้อมูลทั้งหมด ต้องกรอง platform <> 'lazada' "
     "แล้วระบุในรายงานว่าไม่รวม Lazada"),
    ("ยังไม่มีค่าธรรมเนียมและ settlement",
     "commission_fee, transaction_fee, service_fee, settlement_amount ว่าง 0% ทุกแพลตฟอร์ม "
     "ข้อมูลอยู่ในเมนูการเงินซึ่งยังไม่ได้รับอนุญาตให้เข้าถึง",
     "สร้างคอลัมน์ไว้ได้เพื่อไม่ต้องแก้ schema ภายหลัง แต่ห้ามนำไปคำนวณกำไรจนกว่าจะมีข้อมูลจริง"),
    ("สถานะออเดอร์เปลี่ยนได้หลังดึง",
     "ออเดอร์ที่ดึงวันนี้อาจถูกยกเลิกหรือเปลี่ยนเป็นส่งสำเร็จในภายหลัง "
     "ข้อมูลที่ดึงไปแล้วจะไม่ถูกอัปเดตย้อนหลังอัตโนมัติ",
     "ทำ upsert ด้วย (platform, order_id, sku) ไม่ใช่ insert อย่างเดียว "
     "จะได้รองรับการดึงซ้ำเพื่ออัปเดตสถานะโดยไม่เกิดข้อมูลซ้ำ"),
    ("ข้อมูลผู้ซื้อถูกปกปิดแล้ว",
     "buyer_username เก็บแค่ตัวแรกกับตัวท้าย เช่น somchai123 → s********3 "
     "ชื่อ-นามสกุล เบอร์โทร ที่อยู่ เลขบัตรประชาชน ถูกตัดทิ้งก่อนออกไฟล์",
     "ใช้จับลูกค้าซ้ำได้ แต่ระบุตัวตนไม่ได้ ถ้าต้องการข้อมูลเต็มต้องขออนุมัติเรื่อง PDPA ก่อน"),
]
for i, (topic, detail, action) in enumerate(RISKS, start=1):
    ws.cell(row=r, column=1, value=i)
    ws.cell(row=r, column=2, value=topic).font = Font(name=FONT, bold=True, size=10)
    ws.cell(row=r, column=3, value=detail)
    ws.cell(row=r, column=4, value=action)
    for j in range(1, 5):
        c = ws.cell(row=r, column=j)
        c.border = BOX
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if not c.font.bold:
            c.font = Font(name=FONT, size=10)
    ws.row_dimensions[r].height = 58
    r += 1

for col, w in zip("ABCD", (5, 30, 62, 62)):
    ws.column_dimensions[col].width = w
ws.freeze_panes = f"A{hr + 1}"

# ══════════ ชีท 3.5: 2 ระบบต่อกันยังไง ══════════
ws = wb.create_sheet("2 ระบบต่อกันยังไง")
r = title(ws, "สองระบบต่อกันยังไง — และอะไรคือจุดที่กระทบกันได้",
          "ระบบดึงยอด (Dealer MKP Platform) + ระบบสกรีน SKU (osuka-sku) · เชื่อมอัตโนมัติ 2026-08-10")

FLOW = [
    ("ชั้น 1", "Dealer MKP Platform", "ดึงคำสั่งซื้อจากหลังบ้าน 15 ร้าน",
     "Excel 32 คอลัมน์ · schema กลาง", "output/<วันที่>/"),
    ("ชั้น 2", "osuka-sku", "เติม osuka_sml_id / osuka_model_code + 29 คอลัมน์ประกอบ",
     "Excel 63 คอลัมน์ ← ของส่งมอบ", "output/<วันที่>/screened/"),
    ("ชั้น 3", "OSUKA Super Intelligence", "ฐานข้อมูลปลายทาง schema intel",
     "mp_orders_raw · mp_sku_mapping", "รอรายละเอียดการเชื่อมต่อ"),
]
for i, h in enumerate(["ชั้น", "ระบบ", "ทำอะไร", "ผลลัพธ์", "เก็บที่"], start=1):
    ws.cell(row=r, column=i, value=h)
style_header(ws, r, 5)
r += 1
for row_vals in FLOW:
    for j, v in enumerate(row_vals, start=1):
        c = ws.cell(row=r, column=j, value=v)
        c.border = BOX
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.font = Font(name=FONT, size=10, bold=(j <= 2))
    ws.row_dimensions[r].height = 34
    r += 1

r += 1
ws.cell(row=r, column=1, value="สิ่งที่ทำให้ 2 ระบบไม่ชนกัน").font = Font(
    name=FONT, bold=True, size=12, color=NAVY)
r += 1
for i, h in enumerate(["หลักการ", "ทำอย่างไร", "ถ้าไม่ทำจะเกิดอะไร"], start=1):
    ws.cell(row=r, column=i, value=h)
style_header(ws, r, 3)
r += 1

RULES = [
    ("แยกโฟลเดอร์ แยก repo",
     "ระบบดึงก๊อปไฟล์เข้า input/ ของระบบสกรีน แล้วเรียกสคริปต์ของเขา "
     "ไม่แก้โค้ดเขาแม้แต่บรรทัดเดียว ผลลัพธ์ก๊อปกลับมาเก็บฝั่งเรา",
     "ถ้ารวมโฟลเดอร์ตอนนี้ ต้องสร้าง .venv ใหม่ ติดตั้งงานตั้งเวลาใหม่ "
     "และเสี่ยงต้องล็อกอินใหม่ทั้ง 15 ร้าน"),
    ("ที่อยู่ปรับได้ ไม่ฝังตาย",
     "ระบบดึงหาระบบสกรีนจากตัวแปร OSUKA_SKU_DIR (มีค่าเริ่มต้นให้) "
     "ย้ายโฟลเดอร์เมื่อไหร่แค่เปลี่ยนค่านี้",
     "ฝัง path ตายตัวแล้วย้ายโฟลเดอร์ = พังเงียบ ๆ ตอนตี 8 ครึ่ง"),
    ("เอกสารอ่านจากต้นทาง ไม่คัดลอก",
     "ชีท 'หลังสกรีน 63 คอลัมน์' อ่านคำอธิบายสดจาก DD ของ osuka-sku ตอนสร้างไฟล์นี้",
     "คัดลอกมาเก็บซ้ำ = วันหนึ่งเขาแก้ของเขา ของเราค้างอยู่กับข้อมูลเก่า "
     "กลายเป็น 2 แหล่งความจริงที่ขัดกันเอง"),
    ("ขั้นสกรีนล้มไม่ได้ทำให้รอบดึงล้ม",
     "run_daily.ps1 ครอบ try/catch และไม่สนใจ exit code ของขั้นสกรีน "
     "อีเมลถอยไปแนบไฟล์ดิบให้อัตโนมัติ",
     "ดึงสำเร็จ 15 ร้านแล้วไม่ได้อะไรเลยเพราะตัวสกรีนสะดุด"),
    ("ใช้ Python ตัวเดียวกัน",
     "ระบบสกรีนไม่มี .venv ของตัวเอง จึงเรียกด้วย .venv ของระบบดึง "
     "ซึ่งมี pandas + openpyxl ครบตามที่ match_sku.py ต้องใช้",
     "ถ้าต่างคนต่างมี venv แล้วเวอร์ชันไม่ตรง จะได้ผลลัพธ์ต่างกันโดยไม่มีใครรู้"),
]
for who, how, why in RULES:
    ws.cell(row=r, column=1, value=who).font = Font(name=FONT, bold=True, size=10)
    ws.cell(row=r, column=2, value=how)
    ws.cell(row=r, column=3, value=why)
    for j in range(1, 4):
        c = ws.cell(row=r, column=j)
        c.border = BOX
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if not c.font.bold:
            c.font = Font(name=FONT, size=10)
    ws.row_dimensions[r].height = 54
    r += 1

r += 1
ws.cell(row=r, column=1, value="⚠️ จุดเดียวที่ 2 ระบบผูกกันจริง").font = Font(
    name=FONT, bold=True, size=12, color="C00000")
r += 1
COUPLE = [
    "schema 32 คอลัมน์ของชั้นที่ 1 คือ 'สัญญา' ระหว่าง 2 ระบบ",
    "DD ของ osuka-sku เขียนไว้เองว่า \"จำนวนคอลัมน์ดิบเปลี่ยนตาม schema ของไฟล์ต้นทาง\"",
    "แปลว่า ถ้าเราเพิ่ม/ลบ/เปลี่ยนชื่อคอลัมน์ในชั้นที่ 1 ไฟล์ 63 คอลัมน์จะเปลี่ยนตามทันที",
    "ตัวสกรีนต้องการจริง ๆ แค่ 3 คอลัมน์: product_name · sku · variation",
    "→ ถ้าจะแก้ schema ชั้นที่ 1 ต้องแจ้งฝั่ง osuka-sku ก่อน โดยเฉพาะ 3 คอลัมน์นี้",
]
for line in COUPLE:
    ws.cell(row=r, column=1, value=line).font = Font(name=FONT, size=10)
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True)
    r += 1

r += 1
ws.cell(row=r, column=1, value="ข้อที่ DD ของ osuka-sku เตือนไว้ แต่ไม่ใช้กับท่อนี้").font = Font(
    name=FONT, bold=True, size=11, color="806000")
r += 1
ws.cell(row=r, column=1,
        value="เขาเขียนว่า \"buyer_username ยกมาจากไฟล์ต้นทางแบบไม่ mask — ต้องจัดการก่อนเข้าฐานข้อมูล\"  "
              "ข้อนี้เป็นจริงเฉพาะตอนเขารับไฟล์ดิบจากแพลตฟอร์มโดยตรง  "
              "แต่ท่อของเราป้อนไฟล์ที่ mask มาแล้ว (include_pii = false)  "
              "ตรวจจริงกับผลลัพธ์วันที่ 2026-08-10 ทั้ง 7 ร้านที่มีค่า: mask ครบ 100% "
              "เช่น n****0 · f**********m  →  ข้อมูลที่ส่งเข้าฐานไม่มี PII").font = Font(
    name=FONT, size=10)
ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=WARN)
ws.row_dimensions[r].height = 60
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)

for col, w in zip("ABCDE", (30, 52, 52, 26, 24)):
    ws.column_dimensions[col].width = w

# ══════════ ชีท 4: รายชื่อร้าน ══════════
ws = wb.create_sheet("รายชื่อร้าน")
r = title(ws, "รายชื่อร้าน — 9 ชื่อร้าน จาก 13 หน้าร้าน",
          "ชื่อจริงบนแพลตฟอร์มต่างจากชื่อในไฟล์ เพราะร้านเดียวกันตั้งชื่อไม่ตรงกันในแต่ละเจ้า")
for i, h in enumerate(["shop_name (ในไฟล์)", "shop_id", "แพลตฟอร์ม", "ชื่อจริงบนแพลตฟอร์ม"], start=1):
    ws.cell(row=r, column=i, value=h)
style_header(ws, r, 4)
r += 1

SHOPS = [
    ("กัปตัน เอกสตีล", "lazada_01", "Lazada", "กัปตัน เอกสตีล"),
    ("กัปตัน เอกสตีล", "shopee_06", "Shopee", "กัปตัน เอกสตีล"),
    ("Powerstools", "tiktok_01", "TikTok", "powerstool  ← ไม่มี s"),
    ("Powerstools", "shopee_04", "Shopee", "Powerstools"),
    ("เฮียเก๋า เครื่องมือช่างราคาถูก", "tiktok_05", "TikTok", "เฮียเก๋าเครื่องมือช่าง ราคาถูก  ← เว้นวรรคคนละที่"),
    ("เฮียเก๋า เครื่องมือช่างราคาถูก", "shopee_05", "Shopee", "เฮียเก๋า เครื่องมือช่างราคาถูก"),
    ("TNLTOOLSTORE", "shopee_03", "Shopee", "Toolspartner  ← คนละชื่อเลย"),
    ("TNLTOOLSTORE", "shopee_08", "Shopee", "TNLTOOLSTORE"),
    ("toolsdee1", "tiktok_02", "TikTok", "toolsdee1"),
    ("ฝ้ายการช่าง", "tiktok_03", "TikTok", "ฝ้ายการช่าง"),
    ("100อัน1000อย่าง", "tiktok_04", "TikTok", "100อัน1000อย่าง88  ← มี 88 ต่อท้าย"),
    ("เฮียคิมคลองถม", "shopee_01", "Shopee", "เฮียคิมคลองถม"),
    ("Smarttooltech", "shopee_02", "Shopee", "Smarttooltech"),
]
prev = None
for name, sid, plat, real in SHOPS:
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value=sid)
    ws.cell(row=r, column=3, value=plat)
    ws.cell(row=r, column=4, value=real)
    diff = "←" in real
    for j in range(1, 5):
        c = ws.cell(row=r, column=j)
        c.border = BOX
        c.font = Font(name=FONT, size=10, bold=(j == 1 and name != prev))
        if diff:
            c.fill = PatternFill("solid", fgColor=WARN)
    prev = name
    r += 1
r += 1
ws.cell(row=r, column=1,
        value="แถวสีเหลือง = ชื่อจริงบนแพลตฟอร์มไม่ตรงกับ shop_name ในไฟล์").font = Font(
    name=FONT, size=9, italic=True, color="806000")
for col, w in zip("ABCD", (34, 14, 12, 46)):
    ws.column_dimensions[col].width = w

# ══════════ ชีท 5: ค่าที่เป็นไปได้ ══════════
ws = wb.create_sheet("ค่าที่เป็นไปได้")
r = title(ws, "ค่าที่เป็นไปได้ของคอลัมน์ประเภทรหัส")
for i, h in enumerate(["คอลัมน์", "ค่า", "ความหมาย"], start=1):
    ws.cell(row=r, column=i, value=h)
style_header(ws, r, 3)
r += 1

ENUMS = [
    ("platform", "lazada", "Lazada Seller Center"),
    ("platform", "tiktok", "TikTok Shop Seller Center"),
    ("platform", "shopee", "Shopee Seller Centre"),
    ("order_status", OrderStatus.PENDING.value, "รอดำเนินการ / ยังไม่ชำระ"),
    ("order_status", OrderStatus.READY_TO_SHIP.value, "พร้อมจัดส่ง"),
    ("order_status", OrderStatus.SHIPPED.value, "จัดส่งแล้ว"),
    ("order_status", OrderStatus.DELIVERED.value, "ส่งถึงผู้รับแล้ว"),
    ("order_status", OrderStatus.CANCELLED.value, "ยกเลิก"),
    ("order_status", OrderStatus.RETURNED.value, "คืนสินค้า"),
    ("order_status", OrderStatus.UNKNOWN.value,
     "แปลงสถานะไม่ได้ — ดูค่าดิบที่คอลัมน์ status_raw"),
]
for colname, val, mean in ENUMS:
    ws.cell(row=r, column=1, value=colname)
    ws.cell(row=r, column=2, value=val).font = Font(name=FONT, size=10, bold=True)
    ws.cell(row=r, column=3, value=mean)
    for j in range(1, 4):
        c = ws.cell(row=r, column=j)
        c.border = BOX
        if not c.font.bold:
            c.font = Font(name=FONT, size=10)
    r += 1
for col, w in zip("ABC", (18, 20, 56)):
    ws.column_dimensions[col].width = w

# ══════════ ชีท 6: คีย์และ schema ══════════
ws = wb.create_sheet("คีย์และ schema")
r = title(ws, "คีย์ — ผลทดสอบกับข้อมูลจริง",
          "ทดสอบครบทุกแถว 660,181 แถว (ม.ค.–ก.ค. 2026) ไม่ได้สุ่มตัวอย่าง")
for i, h in enumerate(["ผู้สมัครเป็นคีย์", "ผล", "หมายเหตุ"], start=1):
    ws.cell(row=r, column=i, value=h)
style_header(ws, r, 3)
r += 1
KEYS = [
    ("order_id", "❌ ใช้ไม่ได้",
     "ซ้ำ 73,470 แถว — 1 ออเดอร์มีหลายรายการสินค้าตาม grain"),
    ("platform + order_id", "❌ ใช้ไม่ได้",
     "ซ้ำ 73,470 แถว เท่ากับข้างบน (ยืนยันว่า order_id ไม่ชนกันข้ามแพลตฟอร์ม)"),
    ("platform + order_id + sku", "✅ ใช้ได้",
     "ไม่ซ้ำเลยทั้ง 660,181 แถว · ในกลุ่มที่ sku ว่าง 93,486 แถว ก็ไม่ชนกันสักคู่"),
    ("shop_id + order_id + sku", "✅ ใช้ได้",
     "ผลเท่ากับข้างบน เลือกใช้ชุดใดชุดหนึ่งได้"),
]
for k, res, note in KEYS:
    ws.cell(row=r, column=1, value=k).font = Font(name=FONT, size=10, bold=True)
    ws.cell(row=r, column=2, value=res)
    ws.cell(row=r, column=3, value=note)
    for j in range(1, 4):
        c = ws.cell(row=r, column=j)
        c.border = BOX
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if not c.font.bold:
            c.font = Font(name=FONT, size=10)
    ws.row_dimensions[r].height = 32
    r += 1

r += 1
ws.cell(row=r, column=1, value="โครงตาราง Postgres — แยก 2 ระดับตามที่ฝั่งฐานข้อมูลขอ").font = Font(
    name=FONT, bold=True, size=12, color=NAVY)
r += 1
ws.cell(row=r, column=1,
        value="แยกแล้วหมดปัญหายอดซ้ำจาก grain — SUM(total_amount) บน order_header "
              "ได้ยอดขายจริงทันที ไม่ต้องระวัง DISTINCT").font = Font(
    name=FONT, size=9, italic=True, color="595959")
r += 1
DDL = """-- ═══ ตารางที่ 1: ระดับออเดอร์ (1 แถว = 1 ออเดอร์) ═══
CREATE TABLE order_header (
    order_id          VARCHAR(64)  NOT NULL,   -- TEXT เท่านั้น ห้าม BIGINT
    platform          VARCHAR(16)  NOT NULL,   -- lazada / tiktok / shopee
    shop_id           VARCHAR(32)  NOT NULL,   -- หน้าร้าน เช่น shopee_09
    shop_name         VARCHAR(128) NOT NULL,   -- ชื่อมาตรฐาน 1 ชื่อมีได้หลาย shop_id
    order_created_at  TIMESTAMP    NOT NULL,   -- Asia/Bangkok (ไม่มี offset)
    order_updated_at  TIMESTAMP,
    paid_at           TIMESTAMP,               -- Lazada ไม่มี
    order_status      VARCHAR(16)  NOT NULL,   -- 7 ค่า ดูชีท 'ค่าที่เป็นไปได้'
    status_raw        VARCHAR(64),             -- ค่าดิบไว้ย้อนตรวจการ map
    payment_method    VARCHAR(64),
    shipping_fee      DECIMAL(12,2),
    shipping_carrier  VARCHAR(64),
    tracking_no       VARCHAR(64),             -- TEXT เท่านั้น
    total_amount      DECIMAL(12,2),           -- ยอดที่ลูกค้าจ่ายทั้งออเดอร์
    buyer_username    VARCHAR(64),             -- ถูก mask ตาม PDPA
    province          VARCHAR(64),             -- Lazada ไม่มี
    cancel_reason     VARCHAR(256),
    return_status     VARCHAR(64),
    fetched_at        TIMESTAMP    NOT NULL,
    PRIMARY KEY (platform, order_id)
);

-- ═══ ตารางที่ 2: ระดับสินค้า (1 แถว = 1 รายการในออเดอร์) ═══
CREATE TABLE order_line (
    id                BIGSERIAL PRIMARY KEY,
    order_id          VARCHAR(64)  NOT NULL,
    platform          VARCHAR(16)  NOT NULL,
    sku               VARCHAR(128),            -- ว่างได้ ~3% (93,486 จาก 660,181)
    product_name      VARCHAR(512),
    variation         VARCHAR(256),
    quantity          INT,
    item_price        DECIMAL(12,2),
    seller_discount   DECIMAL(12,2),
    platform_discount DECIMAL(12,2),
    notes             VARCHAR(512),            -- เหตุผลที่บางคอลัมน์ว่าง
    FOREIGN KEY (platform, order_id) REFERENCES order_header (platform, order_id),
    UNIQUE (platform, order_id, sku)           -- ทดสอบแล้วไม่ซ้ำทั้ง 660,181 แถว
);

CREATE INDEX idx_hdr_created ON order_header (order_created_at);
CREATE INDEX idx_hdr_shop    ON order_header (shop_id, order_created_at);
CREATE INDEX idx_hdr_name    ON order_header (shop_name, order_created_at);
CREATE INDEX idx_line_order  ON order_line   (platform, order_id);

-- ⚠️ ไม่ใส่ commission_fee / transaction_fee / service_fee / settlement_amount
--    ฝั่งฐานข้อมูลตอบว่า No need และเรายังไม่มีข้อมูลจริงอยู่ดี (ว่าง 0% ทุกแพลตฟอร์ม)
--    ถ้าวันหนึ่งเปิดใช้ ให้เพิ่มเข้า order_header

-- ⚠️ นำเข้าด้วย upsert ไม่ใช่ insert อย่างเดียว
--    สถานะออเดอร์เปลี่ยนได้หลังดึงไปแล้ว (SHIPPED -> DELIVERED / ถูกยกเลิกภายหลัง)
--    ON CONFLICT (platform, order_id) DO UPDATE ...
"""
for line in DDL.split("\n"):
    c = ws.cell(row=r, column=1, value=line)
    c.font = Font(name="Consolas", size=9)
    c.fill = PatternFill("solid", fgColor=GREY)
    r += 1
for col, w in zip("ABC", (78, 16, 56)):
    ws.column_dimensions[col].width = w

# ══════════ ชีท 7: คำถามที่ต้องตัดสินใจ ══════════
ws = wb.create_sheet("คำตอบจากฝั่งฐานข้อมูล")
r = title(ws, "คำตอบจากฝั่งฐานข้อมูล — ได้รับ 2026-08-10",
          "5 ข้อนี้ตอบครบแล้ว ใช้เป็นข้อกำหนดในการส่งมอบต่อจากนี้")
for i, h in enumerate(["#", "คำถาม", "คำตอบที่ได้รับ", "สิ่งที่ต้องทำต่อ"], start=1):
    ws.cell(row=r, column=i, value=h)
style_header(ws, r, 4)
r += 1
QS = [
    ("ต้องการข้อมูลย้อนหลังถึงเมื่อไหร่",
     "From 1 Jan 2026 till latest date D-1",
     "ตรงกับที่มีอยู่แล้ว — ม.ค. 2026 ถึงเมื่อวาน ไม่ต้องดึงเพิ่ม "
     "ระบบดึงของ D-1 ทุกเช้าอยู่แล้ว จึงต่อเนื่องได้เอง"),
    ("จะรับข้อมูลอย่างไร",
     "Export ลง Postgres โดยตรง (รายละเอียดการเชื่อมต่อจะให้ในรอบถัดไป)",
     "รอ host / database / schema / user จากฝั่งฐานข้อมูล "
     "แล้วเราเขียนตัวส่งเข้า Postgres — ดูโครงตารางที่ชีท 'คีย์และ schema'"),
    ("ต้องการตารางระดับออเดอร์แยกอีกตารางไหม",
     "Yes — ให้แยกตารางออเดอร์ออกมา",
     "ส่ง 2 ตาราง: order_header (1 แถว = 1 ออเดอร์) + order_line (1 แถว = 1 สินค้า) "
     "แก้ปัญหายอดซ้ำจาก grain ได้ที่ต้นทาง"),
    ("ต้องการค่าธรรมเนียมและ settlement ไหม",
     "No need",
     "ไม่ต้องขออนุญาตเข้าเมนูการเงิน — 4 คอลัมน์นั้นยังคงไว้ในตารางแต่จะเป็น NULL เสมอ "
     "หรือจะตัดออกก็ได้ถ้าฝั่งฐานข้อมูลต้องการ"),
    ("ต้องการข้อมูลผู้ซื้อแบบไม่ปกปิดไหม",
     "No need — ให้ทำตาม PDPA ปกปิด PII ต่อไป",
     "คงค่า include_pii = false ไว้เหมือนเดิม "
     "buyer_username ยังถูก mask · ไม่มีชื่อ-เบอร์-ที่อยู่ในข้อมูลที่ส่ง"),
]
for i, (q, ans, todo) in enumerate(QS, start=1):
    ws.cell(row=r, column=1, value=i)
    ws.cell(row=r, column=2, value=q).font = Font(name=FONT, bold=True, size=10)
    c_ans = ws.cell(row=r, column=3, value=ans)
    c_ans.font = Font(name=FONT, bold=True, size=10, color="1F6F1F")
    ws.cell(row=r, column=4, value=todo)
    for j in range(1, 5):
        c = ws.cell(row=r, column=j)
        c.border = BOX
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if not c.font.bold:
            c.font = Font(name=FONT, size=10)
        if j == 3:
            c.fill = PatternFill("solid", fgColor=OKBG)
    ws.row_dimensions[r].height = 50
    r += 1
r += 1
ws.cell(row=r, column=1,
        value="ช่องสีเขียว = คำตอบที่ได้รับจากฝั่งฐานข้อมูล 2026-08-10").font = Font(
    name=FONT, size=9, italic=True, color="1F6F1F")
r += 2
ws.cell(row=r, column=1, value="ยังต้องรออีก 1 อย่าง").font = Font(
    name=FONT, bold=True, size=11, color="C00000")
r += 1
ws.cell(row=r, column=1,
        value="รายละเอียดการเชื่อมต่อ Postgres (host / port / database / schema / user / วิธียืนยันตัวตน) "
              "— ยังส่งข้อมูลเข้าฐานไม่ได้จนกว่าจะได้ครบ").font = Font(name=FONT, size=10)
ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
for col, w in zip("ABCD", (5, 34, 40, 52)):
    ws.column_dimensions[col].width = w

OUT.parent.mkdir(exist_ok=True)
wb.save(OUT)
print(f"บันทึกแล้ว: {OUT}")
print(f"ชีททั้งหมด: {', '.join(wb.sheetnames)}")
