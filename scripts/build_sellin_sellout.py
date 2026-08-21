r"""Sell-in vs Sell-out ต่อ dealer — ไฟล์ที่ทีมขอ (2026-08-13)

    .\.venv\Scripts\python.exe -u scripts\build_sellin_sellout.py

⚠️ ทำไมไม่ใช้ intel.v_mp_sellin_sellout ที่มีอยู่แล้ว
   view นั้นผูก sell-out เข้ากับ dealer ผ่าน intel.mp_dealer_shops.vendor_code
   ซึ่งตอนนี้เติมไว้แค่ 2 ร้านจาก 16 → ผูกสำเร็จแค่ 2 dealer จาก 18
   อีก 16 ราย view จะขึ้นว่า "sell-in only - no marketplace access yet"
   ทั้งที่เราดึงข้อมูลร้านเขาได้อยู่ทุกวัน
   สคริปต์นี้ผูกเองจาก config/dealers.yaml แล้วออกไฟล์ให้ทีมได้เลย
   ส่วน view จะถูกต้องเมื่อเจ้าของ Dashboard เติม vendor_code ครบ (ดู FIX_vendor_code.sql)

⚠️ อัตราส่วนที่เป็น "มูลค่า" จะเกิน 100% เป็นเรื่องปกติ ไม่ใช่ข้อมูลผิด
   sell-in คือราคาที่ dealer ซื้อจาก SML / sell-out คือราคาที่ผู้บริโภคจ่ายบนแพลตฟอร์ม
   ส่วนต่างคือกำไรของ dealer → ต้องดู "จำนวนชิ้น" ถึงจะเป็น sell-through จริง
"""
from __future__ import annotations

import os
import subprocess
import sys
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from src.core.naming import canonical_name          # noqa: E402

PSQL = r"C:\Program Files\PostgreSQL\18\bin\psql.exe"
PGPASS = r"C:\Users\tada.p\Postgres\pgpass.conf"
DEALERS = PROJECT_ROOT / "config" / "dealers.yaml"

# เดือนที่ทั้งสองฝั่งมีข้อมูล — sell-in มีถึง 202607 เท่านั้น
# ถ้าเอา sell-out ถึง ส.ค. มาหารด้วย sell-in ที่จบแค่ ก.ค. อัตราส่วนจะเฟ้อทันที
RATIO_MONTHS = ["202601", "202602", "202603", "202604", "202605", "202606", "202607"]

HDR_FILL = PatternFill("solid", fgColor="1F4E79")


def run_sql(sql: str) -> list[list[str]]:
    """ส่งผ่านไฟล์เสมอ — SQL มีภาษาไทย และ path ต้องเป็น ASCII ล้วน"""
    env = dict(os.environ, PGPASSFILE=PGPASS, PGCLIENTENCODING="UTF8")
    tmp = PROJECT_ROOT / "output" / "_sellout_extract" / "q.sql"
    tmp.parent.mkdir(parents=True, exist_ok=True)
    tmp.write_text(sql, encoding="utf-8")
    p = subprocess.run([PSQL, "service=osuka", "-w", "-A", "-t", "-F", "\x1f",
                        "-f", str(tmp)], capture_output=True, env=env)
    if p.returncode != 0:
        print(p.stdout.decode("utf-8", "replace"), p.stderr.decode("utf-8", "replace"))
        raise SystemExit("❌ psql ไม่ผ่าน")
    return [ln.split("\x1f") for ln in
            p.stdout.decode("utf-8", "replace").splitlines() if ln.strip()]


def num(v: str) -> float:
    return float(v) if v not in ("", None) else 0.0


def header(ws, cols: list[tuple[str, int]], row: int = 1) -> None:
    for i, (name, w) in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 32


def main() -> int:
    d = yaml.safe_load(DEALERS.read_text(encoding="utf-8"))
    shop2ar: dict[str, str] = {}
    ar_info: dict[str, dict] = {}
    for x in d.get("dealers", []):
        ar_info[x["ar_code"]] = {"group": x["group"], "juristic": x.get("juristic", ""),
                                 "shops": x["shops"]}
        for s in x["shops"]:
            shop2ar[s] = x["ar_code"]
    for x in d.get("no_shop_yet", []):
        ar_info.setdefault(x["ar_code"], {"group": x["group"], "juristic": "", "shops": []})

    # ── sell-in จาก SML ────────────────────────────────────────────
    sell_in = run_sql("""
        SELECT vendor_code, year_month, max(dealer_legal_name),
               sum(sell_in_thb_incvat), sum(sell_in_thb_exvat), sum(sell_in_units)
        FROM   intel.mp_dealer_sellin GROUP BY 1,2 ORDER BY 1,2
    """)
    si: dict[tuple[str, str], tuple[float, float, float]] = {}
    si_name: dict[str, str] = {}
    for ar, ym, name, inc, exv, units in sell_in:
        si[(ar, ym)] = (num(inc), num(exv), num(units))
        si_name[ar] = name

    # ── sell-out ของเรา รายร้าน รายเดือน ───────────────────────────
    sell_out = run_sql("""
        SELECT shop_id, max(shop_name), max(platform), to_char(ordered_at,'YYYYMM'),
               coalesce(sum(revenue_thb) FILTER (WHERE counts_as_sale
                        AND product_brand ILIKE '%osuka%'),0),
               coalesce(sum(quantity)    FILTER (WHERE counts_as_sale
                        AND product_brand ILIKE '%osuka%'),0),
               coalesce(sum(revenue_thb) FILTER (WHERE counts_as_sale),0),
               count(*)
        FROM   intel.mp_order_line GROUP BY shop_id, 4 ORDER BY 1,4
    """)
    so_shop: dict[str, dict] = {}
    so_ar: dict[str, dict] = {}
    for shop_id, sname, plat, ym, o_thb, o_u, all_thb, lines in sell_out:
        ar = shop2ar.get(shop_id)
        in_ratio = ym in RATIO_MONTHS
        for bucket, key in ((so_shop, shop_id), (so_ar, ar)):
            if key is None:
                continue
            b = bucket.setdefault(key, {"osuka_thb": 0.0, "osuka_u": 0.0, "all_thb": 0.0,
                                        "lines": 0, "shops": set(), "months": set(),
                                        "name": "", "plat": set()})
            b["osuka_thb"] += num(o_thb) * in_ratio
            b["osuka_u"] += num(o_u) * in_ratio
            b["all_thb"] += num(all_thb) * in_ratio
            b["lines"] += int(lines)
            b["shops"].add(shop_id)
            b["months"].add(ym)
            b["plat"].add(plat)
            b["name"] = b["name"] or canonical_name(shop_id, sname)

    wb = Workbook()

    # ══════════ ชีต 1: Sell-in vs Sell-out ต่อ dealer ══════════
    ws = wb.active
    ws.title = "Sell-in vs Sell-out"
    ws["A1"] = ("Sell-in vs Sell-out ต่อ dealer · ม.ค.–ก.ค. 2026 "
                "(sell-in จาก SML มีถึง ก.ค. เท่านั้น จึงตัดที่ ก.ค. ทั้งสองฝั่ง)")
    ws["A1"].font = Font(bold=True, size=12)
    cols = [("AR code", 12), ("Dealer (นิติบุคคลตามฐาน SML)", 36), ("กลุ่มร้าน", 20),
            ("ร้านที่ระบบดึงได้", 15), ("แพลตฟอร์ม", 16),
            ("Sell-in incVAT", 17), ("Sell-in ชิ้น", 13),
            ("Sell-out OSUKA incVAT", 19), ("Sell-out OSUKA ชิ้น", 15),
            ("% ชิ้น (sell-through)", 15), ("% มูลค่า", 12),
            ("Sell-out ทุกแบรนด์", 17), ("สถานะการเชื่อมข้อมูล", 42)]
    header(ws, cols, row=3)

    all_ar = sorted(set(si_name) | set(ar_info),
                    key=lambda a: -sum(si[(a, m)][0] for m in RATIO_MONTHS
                                       if (a, m) in si))
    r = 4
    tot = [0.0] * 5
    unmatched_in = [0.0, 0.0]    # sell-in ของ dealer ที่เรายังไม่มี sell-out
    unmatched_out = [0.0, 0.0]   # sell-out ที่จับคู่ AR code ไม่ได้
    for ar in all_ar:
        inc = sum(si[(ar, m)][0] for m in RATIO_MONTHS if (ar, m) in si)
        u_in = sum(si[(ar, m)][2] for m in RATIO_MONTHS if (ar, m) in si)
        info = ar_info.get(ar, {})
        b = so_ar.get(ar)
        n_shops = len(b["shops"]) if b else 0
        n_conf = len(info.get("shops", []))

        # ร้านที่มีแถวในฐาน แต่ไม่มีเดือนไหนอยู่ในช่วงเทียบเลย = เพิ่งเข้าระบบทีหลัง
        in_window = bool(b) and any(m in RATIO_MONTHS for m in b["months"])
        no_ar = ar not in si_name

        if no_ar:
            note = ("⚠️ AR code นี้ไม่มีในฐาน sell-in ของ SML — เทียบอัตราส่วนไม่ได้ "
                    "ต้องให้ทีมยืนยันรหัสก่อน (ไม่ถูกนับในแถวรวม)")
        elif n_conf == 0:
            note = "ยังไม่ได้เก็บข้อมูลร้าน — ต้องเปิดร้านเข้าระบบก่อน"
        elif b is None:
            note = "ผูกร้านไว้แล้วแต่ยังไม่มีแถวในฐาน"
        elif not in_window:
            first = min(b["months"])
            note = (f"ร้านเพิ่งเข้าระบบเดือน {first[:4]}-{first[4:]} "
                    f"ยังไม่มีข้อมูลช่วง ม.ค.–ก.ค. จึงเทียบไม่ได้ (ไม่ถูกนับในแถวรวม)")
        elif n_shops < n_conf:
            note = f"เก็บได้ {n_shops} จาก {n_conf} ร้านที่ผูกไว้ ยอดจึงเป็นขั้นต่ำ"
        else:
            note = "ครบทุกร้านที่ผูกไว้ (ไม่รวมช่องทางออฟไลน์)"

        vals = [ar, si_name.get(ar, info.get("juristic", "")), info.get("group", ""),
                n_shops or "", " + ".join(sorted(b["plat"])) if b else "",
                round(inc) or "", round(u_in) or "",
                round(b["osuka_thb"]) if in_window else "",
                round(b["osuka_u"]) if in_window else "",
                round(b["osuka_u"] / u_in * 100, 1) if in_window and u_in else "",
                round(b["osuka_thb"] / inc * 100, 1) if in_window and inc else "",
                round(b["all_thb"]) if in_window else "", note]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            if i in (6, 7, 8, 9, 12):
                c.number_format = "#,##0"
            if i in (10, 11):
                c.number_format = '0.0"%"'
        # เหลือง = เทียบอัตราส่วนไม่ได้ ไม่ใช่ยอดเป็นศูนย์
        if no_ar or not in_window:
            for i in range(1, len(cols) + 1):
                ws.cell(row=r, column=i).fill = PatternFill("solid", fgColor="FFF2CC")

        # ⚠️ แถวรวมนับเฉพาะ dealer ที่ "มีข้อมูลทั้งสองฝั่ง" เท่านั้น
        #    เอา sell-out ของ dealer ที่ไม่มี sell-in มารวม -> อัตราส่วนเฟ้อ
        #    เอา sell-in ของ dealer ที่ไม่มี sell-out มารวม -> อัตราส่วนต่ำเกินจริง
        #    ทั้งสองแบบผิด จึงแยกเป็น 3 แถว ให้เห็นว่าอะไรจับคู่ไม่ได้บ้าง
        if not no_ar and in_window:
            tot[0] += inc
            tot[1] += u_in
            tot[2] += b["osuka_thb"]
            tot[3] += b["osuka_u"]
            tot[4] += b["all_thb"]
        elif no_ar:
            unmatched_out[0] += b["osuka_thb"] if in_window else 0
            unmatched_out[1] += b["osuka_u"] if in_window else 0
        else:
            unmatched_in[0] += inc
            unmatched_in[1] += u_in
        r += 1

    ws.cell(row=r, column=1, value="รวม (dealer ที่มีข้อมูลครบทั้งสองฝั่ง)").font = Font(bold=True)
    for i, v in zip((6, 7, 8, 9, 12), tot):
        c = ws.cell(row=r, column=i, value=round(v))
        c.number_format = "#,##0"
        c.font = Font(bold=True)
    c = ws.cell(row=r, column=10, value=round(tot[3] / tot[1] * 100, 1) if tot[1] else "")
    c.number_format = '0.0"%"'
    c.font = Font(bold=True)
    c = ws.cell(row=r, column=11, value=round(tot[2] / tot[0] * 100, 1) if tot[0] else "")
    c.number_format = '0.0"%"'
    c.font = Font(bold=True)
    ws.cell(row=r, column=13, value="อัตราส่วนที่ใช้อ้างอิงได้ ใช้แถวนี้")

    r += 1
    ws.cell(row=r, column=1, value="sell-in ที่ยังไม่มี sell-out")
    ws.cell(row=r, column=6, value=round(unmatched_in[0])).number_format = "#,##0"
    ws.cell(row=r, column=7, value=round(unmatched_in[1])).number_format = "#,##0"
    ws.cell(row=r, column=13, value="dealer ที่ระบบยังไม่ได้เก็บข้อมูลร้าน")

    r += 1
    ws.cell(row=r, column=1, value="sell-out ที่ยังจับคู่ AR ไม่ได้")
    ws.cell(row=r, column=8, value=round(unmatched_out[0])).number_format = "#,##0"
    ws.cell(row=r, column=9, value=round(unmatched_out[1])).number_format = "#,##0"
    ws.cell(row=r, column=13, value="เอกสตีล — รอทีมยืนยัน AR code")
    r += 2
    ws.cell(row=r, column=1, value="แถวสีเหลือง = เทียบอัตราส่วนไม่ได้ ไม่ได้แปลว่ายอดเป็นศูนย์"
            ).font = Font(bold=True)
    ws.cell(row=r + 1, column=1,
            value="  · เอกสตีล (AR-0402) มี sell-out ในระบบเรา แต่ AR code ไม่ตรงกับที่อยู่ในฐาน "
                  "sell-in ของ SML ซึ่งมี 'บริษัท เอกสตีล จำกัด' รหัส AR-1050 — ต้องให้ทีมยืนยันว่า"
                  "เป็นนิติบุคคลเดียวกันหรือคนละราย ก่อนจะรวมยอดเข้าด้วยกัน")
    ws.cell(row=r + 2, column=1,
            value="  · นาดา / DIY Tool / JumboA เพิ่งเปิดให้ระบบดึงเมื่อ 1 ส.ค. 2026 "
                  "จึงยังไม่มีข้อมูลช่วง ม.ค.–ก.ค. ให้เทียบ")
    ws.freeze_panes = "C4"

    # ══════════ ชีต 2: รายร้าน ══════════
    w2 = wb.create_sheet("รายร้าน")
    w2["A1"] = "Sell-out รายร้าน · ม.ค.–ก.ค. 2026 (ช่วงเดียวกับชีตแรก)"
    w2["A1"].font = Font(bold=True, size=12)
    cols2 = [("AR code", 12), ("Dealer group", 22), ("ร้าน", 28), ("shop_id", 12),
             ("แพลตฟอร์ม", 12), ("เดือนที่มีข้อมูล", 14), ("บรรทัดรวมทุกเดือน", 15),
             ("Sell-out OSUKA incVAT", 19), ("Sell-out OSUKA ชิ้น", 15),
             ("Sell-out ทุกแบรนด์", 17), ("% OSUKA", 10)]
    header(w2, cols2, row=3)
    r = 4
    for shop_id, b in sorted(so_shop.items(), key=lambda kv: -kv[1]["osuka_thb"]):
        ar = shop2ar.get(shop_id, "ยังไม่ผูก")
        months = sorted(m for m in b["months"] if m in RATIO_MONTHS)
        vals = [ar, ar_info.get(ar, {}).get("group", ""), b["name"], shop_id,
                " + ".join(sorted(b["plat"])), len(months), b["lines"],
                round(b["osuka_thb"]), round(b["osuka_u"]), round(b["all_thb"]),
                round(b["osuka_thb"] / b["all_thb"] * 100, 1) if b["all_thb"] else ""]
        for i, v in enumerate(vals, start=1):
            c = w2.cell(row=r, column=i, value=v)
            if i in (7, 8, 9, 10):
                c.number_format = "#,##0"
            if i == 11:
                c.number_format = '0.0"%"'
        r += 1
    w2.freeze_panes = "C4"

    # ══════════ ชีต 3: รายเดือน ══════════
    w3 = wb.create_sheet("รายเดือน")
    cols3 = [("เดือน", 10), ("Sell-in incVAT", 17), ("Sell-in ชิ้น", 13),
             ("Sell-out OSUKA incVAT", 19), ("Sell-out OSUKA ชิ้น", 15),
             ("% ชิ้น", 11), ("% มูลค่า", 11), ("ร้านที่มีข้อมูล", 13)]
    header(w3, cols3)
    month_so = run_sql("""
        SELECT to_char(ordered_at,'YYYYMM'), count(DISTINCT shop_id),
               coalesce(sum(revenue_thb) FILTER (WHERE counts_as_sale
                        AND product_brand ILIKE '%osuka%'),0),
               coalesce(sum(quantity)    FILTER (WHERE counts_as_sale
                        AND product_brand ILIKE '%osuka%'),0)
        FROM   intel.mp_order_line GROUP BY 1 ORDER BY 1
    """)
    mso = {ym: (int(n), num(thb), num(u)) for ym, n, thb, u in month_so}
    for i, ym in enumerate(sorted(set(mso) | {m for _, m in si}), start=2):
        inc = sum(v[0] for (a, m), v in si.items() if m == ym)
        u_in = sum(v[2] for (a, m), v in si.items() if m == ym)
        n, thb, u = mso.get(ym, (0, 0.0, 0.0))
        vals = [ym, round(inc) or "", round(u_in) or "", round(thb), round(u),
                round(u / u_in * 100, 1) if u_in else "",
                round(thb / inc * 100, 1) if inc else "", n]
        for j, v in enumerate(vals, start=1):
            c = w3.cell(row=i, column=j, value=v)
            if j in (2, 3, 4, 5):
                c.number_format = "#,##0"
            if j in (6, 7):
                c.number_format = '0.0"%"'
        if ym not in RATIO_MONTHS:
            for j in range(1, len(cols3) + 1):
                w3.cell(row=i, column=j).fill = PatternFill("solid", fgColor="FFF2CC")
    w3.cell(row=len(mso) + 4, column=1,
            value="แถวสีเหลือง = เดือนที่ยังไม่มี sell-in จาก SML เทียบอัตราส่วนไม่ได้")

    # ══════════ ชีต 4: นิยาม ══════════
    w4 = wb.create_sheet("นิยามที่ใช้")
    notes = [
        "นิยามของตัวเลขในไฟล์นี้ — ต้องตรงกันก่อนเอาไปเทียบกับตัวเลขจากที่อื่น",
        "",
        "ช่วงเวลา",
        "  ม.ค.–ก.ค. 2026 ทั้งสองฝั่ง",
        "  sell-in จาก SML โหลดไว้ถึง 202607 เท่านั้น ถ้าเอา sell-out ถึง ส.ค. มาหาร",
        "  อัตราส่วนจะเฟ้อทันที จึงตัดที่ ก.ค. ให้ตรงกัน (ยอด ส.ค. ดูได้ที่ชีตรายเดือน)",
        "",
        "⚠️ % มูลค่า ที่เกิน 100% ไม่ใช่ข้อมูลผิด",
        "  sell-in = ราคาที่ dealer ซื้อจาก SML",
        "  sell-out = ราคาที่ผู้บริโภคจ่ายบนแพลตฟอร์ม",
        "  ส่วนต่างคือกำไรของ dealer เอง ตัวเลขนี้จึงบอก markup ไม่ใช่ sell-through",
        "  ถ้าต้องการ sell-through จริง ให้ดูคอลัมน์ % ชิ้น",
        "",
        "Sell-out ของเรา",
        "  ผลรวม revenue_thb เฉพาะบรรทัดที่ counts_as_sale = true",
        "  counts_as_sale = order_status อยู่ใน DELIVERED / SHIPPED / READY TO SHIP",
        "  ไม่นับ UNPAID / CANCELLED / RETURNED / LOST BY 3PL",
        "  revenue_thb คือยอดที่ผู้ซื้อจ่ายหลังหักส่วนลด และรวม VAT แล้ว",
        "  จึงเทียบกับ sell_in_thb_incvat ไม่ใช่ exvat",
        "",
        "  คอลัมน์ OSUKA กรองด้วย product_brand ที่มีคำว่า osuka",
        "  คอลัมน์ 'ทุกแบรนด์' รวมสินค้ายี่ห้ออื่นที่ร้านเดียวกันขาย ใช้ดูสัดส่วนของเราในร้าน",
        "",
        "⚠️ ตัวเลข sell-out เป็น 'ขั้นต่ำ' เสมอ ไม่ใช่ยอดขายทั้งหมดของ dealer",
        "  1. เก็บเฉพาะร้านที่ dealer เปิดให้ระบบเข้าถึง ร้านอื่นของเขาไม่ถูกนับ",
        "  2. ไม่รวมการขายหน้าร้าน / ขายส่ง / ช่องทางออฟไลน์ ซึ่งอยู่ใน sell-in ด้วย",
        "  → อัตราส่วนที่ได้จึงเป็นพื้นที่ที่ 'มองเห็น' ไม่ใช่ sell-through ทั้งหมดของ dealer",
        "",
        "⚠️ ข้อมูลชุดนี้ผ่านการแก้ 3 อย่างเมื่อ 2026-08-11 ยอดจึงต่างจากรายงานรุ่นก่อน",
        "  1. แก้ Lazada 13,633 บรรทัดที่เคยถูกตีเป็น ADD TO CART ทั้งที่ส่งของแล้ว",
        "     ทำให้ยอด Lazada เพิ่มขึ้นราวเท่าตัว",
        "  2. เปลี่ยนชื่อสถานะ ADD TO CART เป็น UNPAID (ไม่กระทบยอด แค่เปลี่ยนชื่อ)",
        "  3. เติมคำสถานะที่ระบบเดิมแปลไม่ได้ ทำให้บางบรรทัดถูกนับเป็นยอดขายที่ควรนับ",
        "",
        "แหล่งข้อมูล",
        "  sell-in   intel.mp_dealer_sellin (SML f_daily_sale_by_items โหลด 2026-08-02)",
        "  sell-out  intel.mp_order_line (ดึงจากหลังบ้านร้านทุกวัน ล่าสุด 12 ส.ค. 2026)",
    ]
    for i, line in enumerate(notes, start=1):
        c = w4.cell(row=i, column=1, value=line)
        if line and not line.startswith(" "):
            c.font = Font(bold=True)
    w4.column_dimensions["A"].width = 100

    out_dir = PROJECT_ROOT / "output" / "_sellout_extract"
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / "Sell-in vs Sell-out ต่อ dealer_ม.ค.-ก.ค. 2026.xlsx"
    wb.save(out)

    print(f"✅ {out.name}")
    print(f"   dealer ทั้งหมด {len(all_ar)} · ผูกร้านได้ {len(so_ar)}")
    print("   --- dealer ที่มีข้อมูลครบทั้งสองฝั่ง (ใช้อ้างอิงได้) ---")
    print(f"   sell-in  {tot[0]:>14,.0f} บาท · {tot[1]:>9,.0f} ชิ้น")
    print(f"   sell-out {tot[2]:>14,.0f} บาท · {tot[3]:>9,.0f} ชิ้น")
    print(f"   % ชิ้น {tot[3]/tot[1]*100:.1f}%  ·  % มูลค่า {tot[2]/tot[0]*100:.1f}%")
    print("   --- จับคู่ไม่ได้ แยกไว้ไม่ให้ปนอัตราส่วน ---")
    print(f"   sell-in ที่ยังไม่มี sell-out   {unmatched_in[0]:>14,.0f} บาท")
    print(f"   sell-out ที่ยังไม่มี AR code   {unmatched_out[0]:>14,.0f} บาท")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
