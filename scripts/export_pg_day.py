r"""เตรียมข้อมูล "ทั้งวัน ทุกร้าน ทุกแพลตฟอร์ม" ส่งให้ฝั่ง Postgres โหลด

ต่างจาก `export_pg_handoff.py` ที่ทำทีละร้าน ตัวนี้รวมทุกร้านของวันนั้นเป็นไฟล์เดียว
เพื่อให้ Dashboard โชว์ได้ทั้งวันจริง ไม่ใช่ร้านเดียวโดด ๆ

⚠️ ตัวที่กัดแน่ถ้าไม่ระวัง — `intel.mp_order_state()` ฉบับที่อยู่ในฐานตอนนี้
   รู้จักเฉพาะคำของ Shopee พอเจอคำของ Lazada/TikTok จะ RAISE แล้วทั้งทรานแซกชันล้ม
   สคริปต์นี้จึงไล่เช็คคำสถานะให้ก่อน แล้วบอกว่าตัวไหนยังไม่มีในบันได

    .\.venv\Scripts\python.exe -u scripts\export_pg_day.py --date 2026-08-01
"""
from __future__ import annotations

import argparse
import csv
import os
import re
import subprocess
import sys
from collections import Counter, defaultdict
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

BLANKS = {"null", "none", "nan", "n/a", "na", "-", "nil", ""}

# คอลัมน์ที่ส่งให้ปลายทาง — ตรงกับที่ LOAD_revised.sql ฝั่ง Postgres เขียนจริง
# เพิ่ม shop_name จากของเดิม เพราะทำหลายร้านจะ hardcode ชื่อเดียวไม่ได้
COLUMNS = [
    "platform", "order_id", "shop_id", "shop_name",
    "sku", "variation", "product_name",
    "quantity", "item_price", "revenue_thb", "total_amount",
    "ordered_at", "order_created_at", "paid_at", "status_raw",
    "osuka_sml_id", "osuka_model_code", "product_brand", "mapping_status",
    "match_method", "match_confidence", "needs_review",
]

# ── คอลัมน์ที่มีก็เอา ไม่มีก็ปล่อยว่าง ─────────────────────────
# แยกจาก COLUMNS เพราะไฟล์สกรีนรุ่นเก่ายังไม่มีคอลัมน์การเงินชุดใหม่
# ถ้าใส่รวมใน COLUMNS สคริปต์จะ error แล้วโหลดข้อมูลย้อนหลังไม่ได้เลย
#
# ⚠️ เดิม CSV ที่ส่งเข้าฐานไม่มีคอลัมน์การเงินสักตัว แม้แต่ของ Shopee
#    ที่เปิดใช้ไปแล้ว ข้อมูลจึงตกหล่นตั้งแต่ขั้นส่งเข้าฐาน ไม่ใช่ขั้นดึง
OPTIONAL_COLUMNS = [
    "item_discount", "seller_discount", "platform_discount",
    "shipping_fee", "commission_fee", "transaction_fee", "service_fee",
    "settlement_amount", "payment_method", "province",
    # 5 ตัวใหม่ของ TikTok เปิดใช้ 2026-08-11
    "item_subtotal_before_discount", "payment_discount", "tax_amount",
    "shipping_fee_seller_discount", "shipping_fee_platform_discount",
    # ── คอลัมน์ Shopee ที่เพิ่งเปิดใช้ 2026-08-13 ──────────────
    "deal_price", "net_price", "parent_sku", "returned_qty",
    "seller_voucher", "seller_coin_cashback", "seller_bundle_discount",
    "seller_tradein_bonus", "platform_voucher", "platform_bundle_discount",
    "coin_discount", "tradein_discount", "tradein_bonus", "voucher_total",
    "item_paid_by_buyer", "estimated_shipping_fee", "return_shipping_fee",
    "installation_fee_buyer", "installation_fee_actual",
    "shipping_method", "promised_ship_at", "shipped_at", "delivered_at",
    "completed_at", "cancelled_at", "settlement_date",
    # ⚠️ ไม่มี postcode โดยตั้งใจ — privacy.py ลบทิ้งทุกรอบ (อยู่ใน PII_FIELDS)
    #    รหัสไปรษณีย์ + เขต/อำเภอ ชี้ตำแหน่งได้แคบเกินไป ตรงกับที่สั่งว่าไม่เอาที่อยู่
    #    ใส่ไว้ก็ได้ช่องว่างเปล่าทุกแถว หลอกให้คนอ่านคิดว่าข้อมูลหาย
    "district", "country", "order_type",
    "fulfilled_by_platform", "owned_by_platform", "in_bundle_deal",
    "hot_listing", "tax_invoice_requested", "tax_invoice_type", "seller_note",
    # ⚠️ 6 ตัวนี้มีคอลัมน์ในฐานมาตลอด และมีค่าในไฟล์สกรีนมาตลอด
    #    แต่ไม่เคยถูกใส่ในรายการส่งออก จึงไม่เคยถึงฐานเลยสักรอบ
    "shipping_carrier", "tracking_no", "buyer_username",
    "cancel_reason", "return_status", "notes",
    # ⚠️ fetched_at = เวลาที่ "ดึงจริง" ไม่ใช่เวลาที่โหลดเข้าฐาน
    #    Dashboard วัดความสดของแต่ละร้านจากคอลัมน์นี้ (v_mp_shop_freshness)
    #    เดิมไม่ได้ส่งออกมา ทุกแถวที่โหลดหลัง 2026-08-07 จึงเป็นค่าว่าง
    #    ทำให้หน้าเว็บขึ้นว่าร้านค้างมา 6-8 วัน ทั้งที่ข้อมูลถึงเมื่อวาน
    "fetched_at",
    # ── คอลัมน์ผลการจับคู่สินค้า เพิ่ม 2026-08-19 ──────────────────
    #
    # ⚠️ ทั้งชุดนี้มีคอลัมน์ในฐานมาตลอด และมีค่าอยู่ในไฟล์สกรีนมาตลอด
    #    แต่ไม่เคยอยู่ในรายการส่งออก จึงว่างเปล่าใน 445,126 บรรทัด
    #    ที่โหลดผ่านสคริปต์นี้ (ข้อมูลชุดแรกที่โหลดด้วยวิธีอื่นมีครบ)
    #
    #    เจอตอนเจ้าของงานถามว่าทำไม mapping_status ของ Metool ไม่ตรงกับตอนสกรีน
    #    ความจริง mapping_status ตรงอยู่แล้ว — ตัวที่หายคือ match_status
    #    ซึ่งเป็นคนละคอลัมน์ และหายไปพร้อมอีก 18 ตัว
    #
    # ⚠️ ไม่ใส่ is_osuka_brand โดยตั้งใจ — ธงในฐานถูกเติมย้อนหลังแล้ว
    #    (FIX_osuka_flag_backfill.sql) ถ้าโหลดทับด้วยค่าว่างจากไฟล์จะลบทิ้ง
    "match_status", "mapping_status_detail",
    "matched_by", "match_rule",
    "brand_status", "brand_raw",
    "accuracy_matching_pct", "order_mapping_accuracy_pct",
    "variation_name", "review_question", "review_reason",
    "osuka_product_name", "candidates_if_ambiguous",
    "product_key", "name_variants_seen", "renamed_listing", "row_in_source",
    "order_updated_at", "extracted_at",
]

# ── ชื่อคอลัมน์ในไฟล์สกรีนที่ไม่ตรงกับชื่อในฐาน ──────────────────
# ไฟล์ใช้ % แต่ชื่อคอลัมน์ในฐานใช้ _pct (% ใช้เป็นชื่อคอลัมน์ SQL ไม่ได้)
HEADER_ALIAS = {
    "accuracy_matching_pct": "accuracy_matching_%",
    "order_mapping_accuracy_pct": "order_mapping_accuracy_%",
}

PSQL = r"C:\Program Files\PostgreSQL\18\bin\psql.exe"
PGPASS = r"C:\Users\tada.p\Postgres\pgpass.conf"


def ask_db_uncovered(pairs: set[tuple[str, str]]) -> set[tuple[str, str]] | None:
    """ถามฐานตรง ๆ ว่าคู่ (platform, status_raw) ไหนที่ mp_order_state_v2 แปลงไม่ได้

    ⚠️ ทำไมต้องถามฐาน ไม่เก็บลิสต์ไว้ในไฟล์นี้
       เดิมมีลิสต์คำสถานะฝังไว้ตรงนี้ แต่รู้จักแค่ของ Shopee พอฐานอัปเกรดเป็น
       mp_order_state_v2 ที่ครอบ TikTok/Lazada ด้วย ลิสต์ในนี้ไม่ได้ตามไป
       กลายเป็นแจ้งเตือนผิด 891 แถวแล้วบล็อกทั้งไปป์ไลน์ทั้งที่ฐานแปลงได้หมด
       ถามฐานทุกครั้งจึงไม่มีวันหลุดกัน — ฐานเป็นเจ้าของความจริงเรื่องนี้

    คืน None ถ้าต่อฐานไม่ได้ (ตอนนั้นข้ามการตรวจไป ไม่ใช่ถือว่าผ่าน)
    """
    if not Path(PSQL).exists() or not pairs:
        return None

    # ⚠️ ต้องใช้ DO block ที่ดักข้อผิดพลาดทีละแถว ห้ามยิง SELECT ตรง ๆ
    #    mp_order_state_v2 ใช้ RAISE EXCEPTION เมื่อเจอคำที่ไม่รู้จัก
    #    ถ้ายิง SELECT รวม แถวแรกที่ไม่รู้จักจะทำให้ทั้ง query ล้ม แล้วไม่รู้ว่าเหลืออีกกี่คำ
    #    ที่แย่กว่านั้น: psql ไม่ตั้ง ON_ERROR_STOP จะคืน exit 0 ทั้งที่คำสั่งล้ม
    #    ของเดิมจึงอ่านผลว่าง แล้วสรุปว่า "ผ่านหมด" ทั้งที่ไม่ผ่าน — พลาดเงียบสนิท
    values = ",".join(
        "(" + ",".join("$q$" + x + "$q$" for x in p) + ")" for p in sorted(pairs))
    sql = f"""\\set ON_ERROR_STOP on
DO $do$
DECLARE r record; s text;
BEGIN
  FOR r IN SELECT * FROM (VALUES {values}) AS v(platform, status_raw) LOOP
    BEGIN
      s := intel.mp_order_state_v2(r.platform, r.status_raw, NULL);
    EXCEPTION WHEN OTHERS THEN
      RAISE NOTICE 'UNKNOWN|%|%', r.platform, r.status_raw;
    END;
  END LOOP;
END $do$;"""
    tmp = PROJECT_ROOT / "output" / "_pg_check_status.sql"
    tmp.parent.mkdir(parents=True, exist_ok=True)
    tmp.write_text(sql, encoding="utf-8")
    env = dict(os.environ, PGPASSFILE=PGPASS, PGCLIENTENCODING="UTF8")
    p = subprocess.run([PSQL, "service=osuka", "-w", "-f", str(tmp)],
                       capture_output=True, env=env)
    if p.returncode != 0:
        return None
    out = (p.stdout.decode("utf-8", "replace") + p.stderr.decode("utf-8", "replace"))
    bad = set()
    for ln in out.splitlines():
        if "UNKNOWN|" in ln:
            _, plat, status = ln.split("UNKNOWN|", 1)[0], *ln.split("UNKNOWN|", 1)[1].split("|", 1)
            bad.add((plat, status))
    return bad


ALL_COLUMNS = COLUMNS + OPTIONAL_COLUMNS

I_SKU, I_VAR, I_NAME = (ALL_COLUMNS.index("sku"), ALL_COLUMNS.index("variation"),
                        ALL_COLUMNS.index("product_name"))
I_QTY, I_REV = ALL_COLUMNS.index("quantity"), ALL_COLUMNS.index("revenue_thb")
I_PRICE = ALL_COLUMNS.index("item_price")
I_ORDERED, I_STATUS = ALL_COLUMNS.index("ordered_at"), ALL_COLUMNS.index("status_raw")
I_PLAT, I_OID = ALL_COLUMNS.index("platform"), ALL_COLUMNS.index("order_id")
I_TOTAL, I_NET = ALL_COLUMNS.index("total_amount"), ALL_COLUMNS.index("net_price")


def reallocate_revenue(staged: dict) -> tuple[int, float, float]:
    """เฉลี่ยยอดออเดอร์ลงแต่ละบรรทัดตามสัดส่วน แทนการเขียนยอดเต็มทุกบรรทัด

    ⚠️ ทำไมต้องมี — ตัวสกรีนตั้ง revenue_thb = total_amount ซึ่งเป็นค่าระดับ "ออเดอร์"
       แล้วค่านั้นถูกเขียนซ้ำทุกบรรทัดของออเดอร์เดียวกัน ออเดอร์ที่มี 3 SKU
       จึงถูกนับยอดเต็ม 3 รอบ ยอดในฐานเฟ้อ 20.5% (฿159 ล้าน ณ 2026-08-13)

       ยืนยันกับไฟล์จริง 2026-08-13:
         - ออเดอร์หลายบรรทัด 710/710 มี total_amount ซ้ำทุกบรรทัด
         - 'ราคาสินค้าที่ชำระโดยผู้ซื้อ' + 'ค่าจัดส่ง' = 'จำนวนเงินทั้งหมด' ตรง 514/514
         - ไฟล์ Shopee ไม่มียอด "ที่จ่ายจริงรายบรรทัด" ให้เลย ส่วนลดคิดที่ระดับออเดอร์

    น้ำหนักที่ใช้คือ net_price (ราคาขายสุทธิ = ยอดทั้งบรรทัดก่อนหักส่วนลด)
    ซึ่งเป็นค่ารายบรรทัดจริงตัวเดียวที่ไฟล์ให้มา
    ถ้าไม่มี net_price (Lazada/TikTok) ถอยไปใช้ item_price x quantity
    ถ้ายังไม่ได้อีก หารเท่ากันทุกบรรทัด — ผลรวมต่อออเดอร์เท่าเดิมเสมอ

    การเฉลี่ยไม่ได้สร้างยอดขึ้นใหม่ ผลรวมต่อออเดอร์เท่ากับที่แพลตฟอร์มแจ้งเป๊ะ
    แค่กระจายให้ถูกบรรทัด เพื่อให้แยกยอดราย SKU / ราย brand ได้
    (ออเดอร์เดียวมีทั้งสินค้า OSUKA และแบรนด์อื่นปนกัน นับครั้งเดียวต่อออเดอร์จึงใช้ไม่ได้)
    """
    by_order: dict[tuple, list] = defaultdict(list)
    for vals in staged.values():
        by_order[(vals[I_PLAT], vals[I_OID])].append(vals)

    touched = before = after = 0
    for rows in by_order.values():
        before += sum(num(r[I_REV]) for r in rows)

        # ยอดออเดอร์: ค่าที่ซ้ำอยู่ทุกบรรทัด จึงหยิบมาตัวเดียว
        total = max(num(r[I_TOTAL]) for r in rows) or max(num(r[I_REV]) for r in rows)

        # ออเดอร์บรรทัดเดียว: ยอดบรรทัด = ยอดออเดอร์ ตรง ๆ
        # ต้องเซ็ตด้วย ไม่ใช่ปล่อยผ่าน เพราะบรรทัดที่ถูก merge มาก่อนหน้านี้
        # เคยพา revenue เดิมที่บวกซ้ำติดมา
        if len(rows) == 1:
            rows[0][I_REV] = fmt(total)
            after += total
            continue
        weights = [num(r[I_NET]) for r in rows]
        if sum(weights) <= 0:
            weights = [num(r[I_PRICE]) * num(r[I_QTY]) for r in rows]
        if sum(weights) <= 0:
            weights = [1.0] * len(rows)

        w_sum = sum(weights)
        # ปัดเศษให้บรรทัดสุดท้ายรับส่วนต่าง ผลรวมจะเท่ายอดออเดอร์เป๊ะ ไม่หายเศษสตางค์
        running = 0.0
        for i, (r, w) in enumerate(zip(rows, weights)):
            share = total - running if i == len(rows) - 1 else round(total * w / w_sum, 2)
            running += share
            r[I_REV] = fmt(share)
        touched += len(rows)
        after += total

    return touched, before, after


# ยุบช่องว่างทุกชนิดที่ติดกันให้เหลือช่องว่างเดียว
#
# ⚠️ ทำไมต้องยุบ: sku / variation / product_name เป็น 3 ใน 5 คอลัมน์ของคีย์
#    ไฟล์ Export ส่งค่าเดิมมาไม่เหมือนกันทุกครั้ง เจอทั้งอักขระขึ้นบรรทัดใหม่
#    ฝังกลาง sku และช่องว่างซ้อน เช่น 'OCMC537-M1 +  OCMC2536'
#    ถ้าไม่ยุบ ครั้งหน้าที่ต้นทางส่งมาต่างไปนิดเดียว คีย์จะเปลี่ยน แล้ว
#    ON CONFLICT DO UPDATE จะหาแถวเดิมไม่เจอ กลายเป็นแถวใหม่ซ้อนเข้ามาแทน
#
# กฎนี้ต้องตรงกับฝั่งฐานเป๊ะ: btrim(regexp_replace(col, '\s+', ' ', 'g'))
# \s ของ Python ครอบ \xa0 (non-breaking space) ด้วย ส่วนของ Postgres ไม่ครอบ
# จึงแปลง \xa0 เป็นช่องว่างธรรมดาก่อน ทั้งสองฝั่งจะได้ผลลัพธ์เดียวกัน
_WS = re.compile(r"\s+")


def clean(v: object) -> str:
    if v is None:
        return ""
    s = _WS.sub(" ", str(v).replace("\xa0", " ")).strip()
    return "" if s.lower() in BLANKS else s


def num(s: str) -> float:
    try:
        return float(s or 0)
    except ValueError:
        return 0.0


def fmt(x: float) -> str:
    """คืนเป็นจำนวนเต็มถ้าลงตัว — กัน quantity กลายเป็น 2.0 ใน CSV"""
    return str(int(x)) if x == int(x) else repr(x)


def merge(keep: list[str], extra: list[str]) -> None:
    """รวม 2 บรรทัดที่มีคีย์ 5 คอลัมน์เดียวกันให้เหลือบรรทัดเดียว

    บวกเฉพาะ quantity กับ net_price — สองตัวนี้เป็นค่าราย "บรรทัด" จริง
    ห้ามบวก total_amount เด็ดขาด เพราะเป็นค่าระดับ "ออเดอร์" ที่ซ้ำอยู่ทุกบรรทัด
    ของตะกร้าเดียวกัน บวกแล้วยอดจะพองตามจำนวนสินค้าในตะกร้า
    item_price เป็นราคาต่อชิ้น ก็ไม่บวกเช่นกัน

    ⚠️ ไม่บวก revenue_thb ที่นี่แล้ว (เดิมบวก) — revenue_thb ที่ตัวสกรีนให้มา
       คือ total_amount ซึ่งเป็นค่าระดับออเดอร์ บวกแล้วได้ 2 เท่าของยอดจริง
       ตอนนี้ reallocate_revenue() เป็นคนกำหนดค่านี้ให้ทุกบรรทัดแทน
    """
    keep[I_QTY] = fmt(num(keep[I_QTY]) + num(extra[I_QTY]))
    keep[I_NET] = fmt(num(keep[I_NET]) + num(extra[I_NET]))




def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--date", default="2026-08-01")
    # ⚠️ โหมดช่วงวัน ใช้กับงานย้อนหลัง — ทำทีละวันเปลืองโดยไม่จำเป็น
    #    212 วันต้องเปิด/ปิด transaction 424 รอบ (ซ้อม+เขียนจริง วันละ 2)
    #    ทั้งที่ข้อมูลจริงแค่ ~60,000 แถว งานที่หนักคือ overhead ไม่ใช่การเขียน
    ap.add_argument("--from", dest="d_from", help="โหมดช่วง: วันเริ่ม")
    ap.add_argument("--to", dest="d_to", help="โหมดช่วง: วันสุดท้าย")
    # ⚠️ ใช้ตอน "เพิ่มร้านทีหลัง" เท่านั้น — ร้านที่เพิ่งเปิดให้ดึงย้อนหลัง
    #    ปลอดภัยเพราะร้านนั้นยังไม่มีแถวในช่วงนั้นเลย จึงเป็นการ INSERT ล้วน
    #    ไม่ไปแตะร้านอื่น และประหยัดเวลาอ่าน Excel ของอีก 15 ร้านที่ไม่ได้เปลี่ยน
    #    (shopee_09 / 10 / 11 เปิดให้ดึงทีหลัง เกิดขึ้นแล้ว 3 ครั้งในเดือนเดียว)
    ap.add_argument("--shop", default="", help="ทำเฉพาะร้านนี้ เช่น shopee_10")
    # ⚠️ ไฟล์จากงานย้อนหลังไม่ได้อยู่ในโฟลเดอร์รายวัน (output/<วันที่>/screened)
    #    แต่อยู่รวมเป็นไฟล์เดียวใน output/_backfill_* ตัวหาไฟล์ปกติจึงไม่เจอ
    #    ตัวเลือกนี้ให้ระบุไฟล์สกรีนเองตรง ๆ ตรรกะที่เหลือใช้ของเดิมทั้งหมด
    #    รวมทั้งด่านกรองวันที่ — แถวที่อยู่นอกช่วง --from/--to ยังถูกทิ้งเหมือนเดิม
    ap.add_argument("--files", nargs="+", default=None,
                    help="ระบุไฟล์ *_matched.xlsx เอง (ใช้กับงานย้อนหลัง)")
    args = ap.parse_args()

    if bool(args.d_from) != bool(args.d_to):
        print("❌ โหมดช่วงต้องใส่ทั้ง --from และ --to")
        return 1

    d = date.fromisoformat(args.d_from or args.date)
    d_last = date.fromisoformat(args.d_to) if args.d_to else d
    if d_last < d:
        print("❌ --from ต้องไม่หลัง --to")
        return 1

    days = [d + timedelta(days=i) for i in range((d_last - d).days + 1)]
    want_days = {x.isoformat() for x in days}

    files: list[Path] = []
    missing: list[str] = []
    if args.files:
        for f in args.files:
            p = Path(f)
            if not p.exists():
                print(f"❌ ไม่พบไฟล์ {p}")
                return 1
            files.append(p)
        print(f"ใช้ไฟล์ที่ระบุเอง {len(files)} ไฟล์")
    for x in ([] if args.files else days):
        # โฟลเดอร์ตั้งชื่อตาม "วันรัน" ซึ่งคือวันถัดจากวันของข้อมูล
        sub = PROJECT_ROOT / "output" / (x + timedelta(days=1)).isoformat() / "screened"
        got = sorted(sub.glob("*_matched.xlsx"))
        if args.shop:
            # ประกบขีดล่างทั้งสองข้าง ไม่งั้น shopee_1 จะไปโดน shopee_10/11 ด้วย
            got = [g for g in got if f"_{args.shop}_" in g.name]
        if got:
            files.extend(got)
        else:
            missing.append(x.isoformat())
    if not files:
        print(f"❌ ไม่พบไฟล์สกรีนของช่วง {d} ถึง {d_last}")
        return 1
    if missing:
        print(f"⚠️ ไม่มีไฟล์สกรีน {len(missing)} วัน: "
              f"{', '.join(missing[:6])}{' ...' if len(missing) > 6 else ''}")
    print(f"ไฟล์ต้นทาง {len(files)} ไฟล์ · ช่วง {d} ถึง {d_last} ({len(days)} วัน)\n")

    # ⚠️ ชื่อโฟลเดอร์/ไฟล์ต้องเป็น ASCII ล้วน — psql เปิดไฟล์ที่ path มีอักษรไทยไม่ได้
    #    (เจอจริง 2026-08-13: ใช้ "_ถึง_" แล้ว psql ฟ้อง No such file or directory)
    tag = f"{d}" if len(days) == 1 else f"{d}_to_{d_last}"
    if args.shop:
        tag = f"{tag}_{args.shop}"
    out_dir = PROJECT_ROOT / "output" / f"_pg_day_{tag}"
    out_dir.mkdir(parents=True, exist_ok=True)
    csv_path = out_dir / f"all_shops_{tag}.csv"

    staged: dict[tuple, list[str]] = {}
    keys: Counter = Counter()
    seen_status: Counter = Counter()
    off_day: Counter = Counter()
    merged_rows = 0
    raw_qty = 0

    for src in files:
        wb = load_workbook(src, read_only=True, data_only=True)
        try:
            ws = wb["data"]
            it = ws.iter_rows(values_only=True)
            hdr = [str(c) if c is not None else "" for c in next(it)]
            missing = [c for c in COLUMNS if c not in hdr]
            if missing:
                print(f"❌ {src.name} ขาดคอลัมน์ {missing}")
                return 1
            # คอลัมน์ทางเลือก: ไม่มีก็ไม่เป็นไร ปล่อยเป็นค่าว่าง ไม่ทำให้ทั้งรอบล้ม
            idx = {c: hdr.index(c) for c in COLUMNS}
            # คอลัมน์ทางเลือก: ไม่มีก็ปล่อยว่าง และรองรับชื่อในไฟล์ที่ไม่ตรงกับชื่อในฐาน
            for c in OPTIONAL_COLUMNS:
                h = HEADER_ALIAS.get(c, c)
                if h in hdr:
                    idx[c] = hdr.index(h)

            for r in it:
                if not any(v is not None for v in r):
                    continue
                vals = [clean(r[idx[c]]) if c in idx else ""
                        for c in ALL_COLUMNS]
                vals[0] = vals[0].lower()

                # กันไฟล์ปนวัน — เคยโดนมาแล้วตอนสกรีน ไฟล์ค้างทำยอดเฟ้อ 18%
                # กันไฟล์ปนวัน — เก็บเฉพาะวันที่อยู่ในช่วงที่สั่ง
                day = vals[I_ORDERED][:10]
                if day not in want_days:
                    off_day[day] += 1
                    continue

                raw_qty += num(vals[I_QTY])

                k = (vals[0], vals[1], vals[I_SKU], vals[I_VAR], vals[I_NAME])
                keys[k] += 1
                if k in staged:
                    # ⚠️ ยุบรวม ห้ามทิ้ง — Shopee ส่งสินค้าเดียวกันในออเดอร์เดียว
                    #    มาเป็น 2 บรรทัดได้จริง ถ้าปล่อยไปชนกันที่ปลายทาง
                    #    ON CONFLICT DO UPDATE จะทับทิ้ง ทำให้จำนวนชิ้นกับยอดเงินหายเงียบ ๆ
                    #    (Lazada ยุบแบบนี้อยู่แล้วตั้งแต่ต้นทาง Shopee เพิ่งเจอ)
                    merge(staged[k], vals)
                    merged_rows += 1
                    continue
                staged[k] = vals
        finally:
            wb.close()

    touched, rev_before, rev_after = reallocate_revenue(staged)
    print("=== เฉลี่ยยอดออเดอร์ลงรายบรรทัด ===")
    print(f"  บรรทัดที่ถูกเฉลี่ยใหม่  {touched:,}")
    print(f"  ยอดแบบเดิม (นับซ้ำ)   {rev_before:>16,.2f}")
    print(f"  ยอดหลังเฉลี่ย          {rev_after:>16,.2f}")
    if rev_before:
        print(f"  ตัดยอดที่นับซ้ำออก      {rev_before - rev_after:>16,.2f} "
              f"({(1 - rev_after / rev_before) * 100:.1f}%)\n")

    with csv_path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, quoting=csv.QUOTE_ALL, lineterminator="\n")
        w.writerow(ALL_COLUMNS)
        for vals in staged.values():
            w.writerow(vals)

    per_shop: dict[str, list] = defaultdict(lambda: [0, set()])
    per_platform: Counter = Counter()
    orders: set[str] = set()
    qty = 0
    for vals in staged.values():
        orders.add(vals[1])
        qty += num(vals[I_QTY])
        per_shop[vals[2]][0] += 1
        per_shop[vals[2]][1].add(vals[1])
        per_platform[vals[0]] += 1
        seen_status[(vals[0], vals[I_STATUS])] += 1

    n = len(staged)
    dup_keys = {k: c for k, c in keys.items() if c > 1}

    print("=== ตัวเลขจากไฟล์ต้นทาง (ใช้กระทบยอด ไม่ได้นับจากฐาน) ===")
    print(f"  บรรทัดสินค้า   {n:,}")
    print(f"  ออเดอร์ไม่ซ้ำ  {len(orders):,}")
    print(f"  ผลรวม quantity {qty:,.0f}")
    if merged_rows:
        print(f"  ยุบรวม {len(dup_keys)} คีย์ · {merged_rows} แถว "
              f"— quantity ก่อนยุบ {raw_qty:,.0f} หลังยุบ {qty:,.0f} "
              f"({'ครบ ไม่หาย' if raw_qty == qty else f'⚠️ หาย {raw_qty - qty:,.0f}'})")
    if off_day:
        print(f"  ⚠️ ตัดแถวที่อยู่นอกช่วง {d}..{d_last} ออก {sum(off_day.values()):,} แถว "
              f"({dict(list(off_day.items())[:4])})")

    print("\n=== แยกตามแพลตฟอร์ม ===")
    for p, c in per_platform.most_common():
        print(f"  {p:<10} {c:>7,}")

    print("\n=== แยกตามร้าน ===")
    for shop in sorted(per_shop):
        rows, ords = per_shop[shop]
        print(f"  {shop:<12} {rows:>6,} บรรทัด · {len(ords):>5,} ออเดอร์")

    print(f"\n✅ {csv_path.relative_to(PROJECT_ROOT)}  ({csv_path.stat().st_size/1024:,.0f} KB)")

    bad = ask_db_uncovered({p for p in seen_status if p[1]})
    if bad is None:
        print("\n⚠️ ต่อฐานไม่ได้ ข้ามการตรวจคำสถานะ — ตอนโหลดจริงจะรู้อยู่ดี "
              "เพราะ LOAD_day.sql ซ้อมด้วย ROLLBACK ก่อนเสมอ")
    elif bad:
        n_rows = sum(c for p, c in seen_status.items() if p in bad)
        print(f"\n⛔ คำสถานะที่ intel.mp_order_state_v2 แปลงไม่ได้ — รวม {n_rows:,} แถว")
        print("   ถ้าโหลดตอนนี้ ฟังก์ชันจะ RAISE แล้วทรานแซกชันล้มทั้งก้อน")
        print("   ต้องเติมคำเหล่านี้เข้าฟังก์ชันในฐานก่อน (ดู FIX_v2_add_pending.sql เป็นตัวอย่าง)")
        for p in sorted(bad, key=lambda x: -seen_status.get(x, 0)):
            print(f"   {seen_status.get(p, 0):>6,}  {p[0]} · {p[1]}")
        return 1
    else:
        print(f"\n✅ คำสถานะทั้ง {len(seen_status)} แบบ ฐานแปลงได้หมด (ถามฐานตรง ๆ ไม่ใช่ลิสต์ฝังไว้)")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
