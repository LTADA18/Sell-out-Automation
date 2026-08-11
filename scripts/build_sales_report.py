r"""ไฟล์รวมยอดขายช่วงวันที่กำหนด ดึงจาก Postgres โดยตรง

ดึงจากฐานไม่ใช่จากไฟล์ Excel เพราะฐานผ่านการแก้มาแล้วหลายรอบ
ตัวเลขจึงตรงกับ Dashboard เป๊ะ ต่างจากไฟล์ Excel เดิมที่ยังมีค่าเก่าค้างอยู่
  order_status ในไฟล์ Excel ยังเป็นชุดคำเก่า UNKNOWN/PENDING/READY_TO_SHIP
  ส่วนในฐานเป็นชุด 7 คำที่ถูกต้อง รวม UNPAID ที่เพิ่งเปลี่ยนจาก ADD TO CART

    .\.venv\Scripts\python.exe -u scripts\build_sales_report.py --from 2026-08-01 --to 2026-08-10
"""
from __future__ import annotations

import argparse
import os
import subprocess
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
PSQL = r"C:\Program Files\PostgreSQL\18\bin\psql.exe"
PGPASS = r"C:\Users\tada.p\Postgres\pgpass.conf"

# ⚠️ ไม่ฮาร์ดโค้ดรายชื่อคอลัมน์ — อ่านจากฐานตอนรัน
#    เคยเลือกมาเองแค่ 29 คอลัมน์แล้วของหาย เจ้าของงานต้องการครบทุกคอลัมน์
#    อ่านจาก information_schema แปลว่าถ้าฐานเพิ่มคอลัมน์ ไฟล์นี้ได้ตามอัตโนมัติ
COLUMNS_SQL = """
SELECT column_name FROM information_schema.columns
WHERE table_schema = 'intel' AND table_name = 'mp_order_line'
ORDER BY ordinal_position
"""

# บังคับรูปแบบข้อความ — เลข 19 หลักของ TikTok ถ้าหลุดเป็น int จะโดนปัดหลักท้ายทิ้ง
TEXT_COLS = {"order_id", "sku", "tracking_no", "osuka_sml_id", "osuka_model_code",
             "shop_id", "order_month", "product_key", "parent_sku", "variation",
             "buyer_username", "line_id"}
NUM_COLS = {"quantity", "item_price", "item_discount", "seller_discount",
            "platform_discount", "shipping_fee", "commission_fee", "transaction_fee",
            "service_fee", "total_amount", "settlement_amount", "revenue_thb",
            "accuracy_matching_pct", "order_mapping_accuracy_pct", "name_variants_seen",
            "row_in_source"}


def run_sql(sql: str) -> list[list[str]]:
    """ส่งผ่านไฟล์เสมอ และ path ต้องเป็น ASCII ล้วน"""
    env = dict(os.environ, PGPASSFILE=PGPASS, PGCLIENTENCODING="UTF8")
    tmp = PROJECT_ROOT / "output" / "_sku_review_tmp" / "sales_query.sql"
    tmp.parent.mkdir(parents=True, exist_ok=True)
    tmp.write_text(sql, encoding="utf-8")
    p = subprocess.run(
        [PSQL, "service=osuka", "-w", "-A", "-t", "-F", "\x1f", "-f", str(tmp)],
        capture_output=True, env=env,
    )
    if p.returncode != 0:
        print(p.stdout.decode("utf-8", "replace"), p.stderr.decode("utf-8", "replace"))
        raise SystemExit("❌ psql ไม่ผ่าน")
    return [ln.split("\x1f") for ln in
            p.stdout.decode("utf-8", "replace").splitlines() if ln.strip()]


def num(s: str):
    if s is None or s == "":
        return None
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except ValueError:
        return s


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--from", dest="d_from", default="2026-08-01")
    ap.add_argument("--to", dest="d_to", default="2026-08-10")
    args = ap.parse_args()

    COLUMNS = [r[0] for r in run_sql(COLUMNS_SQL)]
    print(f"คอลัมน์ในตาราง {len(COLUMNS)} คอลัมน์ — เอาครบทุกตัว")

    cols = ", ".join(COLUMNS)
    sql = (f"SELECT {cols} FROM intel.mp_order_line "
           f"WHERE ordered_at >= DATE '{args.d_from}' "
           f"  AND ordered_at < DATE '{args.d_to}' + 1 "
           f"ORDER BY ordered_at, shop_id, order_id")
    print(f"ดึง {args.d_from} ถึง {args.d_to} จาก Postgres ...")
    rows = run_sql(sql)
    print(f"  {len(rows):,} บรรทัด")

    i_day = COLUMNS.index("ordered_at")
    i_shop = COLUMNS.index("shop_name")
    i_plat = COLUMNS.index("platform")
    i_st = COLUMNS.index("order_status")
    i_sale = COLUMNS.index("counts_as_sale")
    i_rev = COLUMNS.index("revenue_thb")
    i_qty = COLUMNS.index("quantity")
    i_ord = COLUMNS.index("order_id")

    wb = Workbook()

    # ---------------- ชีทข้อมูล ----------------
    ws = wb.active
    ws.title = "ข้อมูล"
    for i, name in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = \
            34 if name == "product_name" else (18 if "_at" in name else 15)
    i_paid = COLUMNS.index("paid_at")
    i_status = COLUMNS.index("order_status")
    n_unpaid = n_nodata = 0

    for r_i, r in enumerate(rows, start=2):
        for c_i, name in enumerate(COLUMNS, start=1):
            v = r[c_i - 1]
            # ── paid_at ที่ว่าง ต้องบอกให้ชัดว่าว่างเพราะอะไร ────────────
            # ⚠️ ห้ามเติม Unpaid ลงทุกแถวที่ว่าง ผูกป้ายกับ order_status เท่านั้น
            #    paid_at ว่างมี 94,824 แถว แต่ในนั้น 16,720 แถวส่งของไปแล้ว
            #    lazada READY TO SHIP 11,688 แถว มูลค่า 22.15 ล้านบาท นับเป็นยอดขายอยู่
            #    เพราะ export ของ lazada ไม่มีคอลัมน์เวลาชำระเงินมาแต่ต้นทาง
            #    เขียนว่า Unpaid ตรงนั้น = บอกว่าลูกค้าไม่จ่ายทั้งที่จ่ายแล้ว
            if name == "paid_at" and not v:
                if r[i_status] == "UNPAID":
                    v = "Unpaid"
                    n_unpaid += 1
                else:
                    v = "ไม่มีข้อมูล"
                    n_nodata += 1
            cell = ws.cell(row=r_i, column=c_i,
                           value=(num(v) if name in NUM_COLS else v))
            if name in TEXT_COLS:
                cell.number_format = "@"     # เลข 19 หลักของ TikTok ห้ามหลุดเป็น int
            elif name in NUM_COLS:
                cell.number_format = "#,##0.00" if name != "quantity" else "#,##0"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows)+1}"

    # ---------------- สรุป ----------------
    def agg(key_fn):
        d = defaultdict(lambda: [0, set(), 0, 0.0])   # บรรทัด, ออเดอร์, ชิ้น, ยอดขาย
        for r in rows:
            k = key_fn(r)
            d[k][0] += 1
            d[k][1].add(r[i_ord])
            d[k][2] += int(float(r[i_qty] or 0))
            if r[i_sale] == "t":
                d[k][3] += float(r[i_rev] or 0)
        return d

    def sheet(title, data, head):
        s = wb.create_sheet(title)
        for i, h in enumerate([head, "บรรทัด", "ออเดอร์", "ชิ้น", "ยอดขาย"], start=1):
            c = s.cell(row=1, column=i, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E79")
        for i, (k, v) in enumerate(sorted(data.items(), key=lambda kv: -kv[1][3]), start=2):
            s.cell(row=i, column=1, value=k)
            s.cell(row=i, column=2, value=v[0]).number_format = "#,##0"
            s.cell(row=i, column=3, value=len(v[1])).number_format = "#,##0"
            s.cell(row=i, column=4, value=v[2]).number_format = "#,##0"
            s.cell(row=i, column=5, value=round(v[3], 2)).number_format = "#,##0.00"
        n = len(data) + 2
        s.cell(row=n, column=1, value="รวม").font = Font(bold=True)
        for col, idx in ((2, 0), (4, 2)):
            s.cell(row=n, column=col, value=sum(v[idx] for v in data.values())).number_format = "#,##0"
        s.cell(row=n, column=3, value=len({o for v in data.values() for o in v[1]})).number_format = "#,##0"
        s.cell(row=n, column=5, value=round(sum(v[3] for v in data.values()), 2)).number_format = "#,##0.00"
        for col, w in (("A", 30), ("B", 12), ("C", 12), ("D", 12), ("E", 18)):
            s.column_dimensions[col].width = w
        return s

    by_day = agg(lambda r: str(r[i_day])[:10])
    sheet("สรุปรายวัน", by_day, "วันที่")
    sheet("สรุปตามร้าน", agg(lambda r: r[i_shop]), "ร้าน")
    sheet("สรุปตามแพลตฟอร์ม", agg(lambda r: r[i_plat]), "แพลตฟอร์ม")
    sheet("สรุปตามสถานะ", agg(lambda r: r[i_st]), "สถานะ")

    out_dir = PROJECT_ROOT / "output" / f"_report_{args.d_from}_{args.d_to}"
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / f"ยอดขายรวม_{args.d_from}_ถึง_{args.d_to}.xlsx"
    wb.save(out)

    total_sale = sum(v[3] for v in by_day.values())
    total_qty = sum(v[2] for v in by_day.values())
    orders = len({r[i_ord] for r in rows})
    print(f"\n✅ {out.relative_to(PROJECT_ROOT)}  ({out.stat().st_size/1024/1024:,.1f} MB)")
    print(f"   {len(rows):,} บรรทัด · {orders:,} ออเดอร์ · {total_qty:,} ชิ้น")
    print(f"   ยอดขาย (นับเฉพาะ counts_as_sale) {total_sale:,.0f} บาท")
    print(f"\n   paid_at ที่ว่าง เติมป้ายแล้ว")
    print(f"     Unpaid       {n_unpaid:>7,} แถว  (สถานะเป็น UNPAID จริง)")
    print(f"     ไม่มีข้อมูล    {n_nodata:>7,} แถว  (ส่งของแล้วแต่ต้นทางไม่ให้เวลาชำระมา)")
    print("\n   วันที่        บรรทัด    ออเดอร์      ยอดขาย")
    for d in sorted(by_day):
        v = by_day[d]
        print(f"   {d}  {v[0]:>7,}  {len(v[1]):>8,}  {v[3]:>14,.0f}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
