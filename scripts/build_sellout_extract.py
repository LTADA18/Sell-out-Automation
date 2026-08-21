r"""ตาราง sell-out ต่อร้าน ส่งให้ทีม — ดึงจาก Postgres

⚠️ รวมด้วย shop_id ไม่ใช่ shop_name
   ร้านเดียวกันมีหลายชื่อในฐาน เพราะข้อมูล 7 เดือนแรกโหลดด้วยชื่อดิบจากแพลตฟอร์ม
   ส่วนที่โหลดหลัง 2026-08-11 ใช้ชื่อมาตรฐาน รวมด้วยชื่อแล้วยอดร้านเดียวจะถูกหั่นเป็นสองก้อน

    .\.venv\Scripts\python.exe -u scripts\build_sellout_extract.py --from 2026-01-01 --to 2026-08-10
"""
from __future__ import annotations

import argparse
import os
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

import yaml                                          # noqa: E402

from src.core.naming import canonical_name          # noqa: E402

PSQL = r"C:\Program Files\PostgreSQL\18\bin\psql.exe"
PGPASS = r"C:\Users\tada.p\Postgres\pgpass.conf"
DEALERS = PROJECT_ROOT / "config" / "dealers.yaml"


def load_dealers() -> tuple[dict[str, dict], list[dict]]:
    """shop_id -> {ar_code, group, juristic} และรายชื่อกลุ่มที่ยังไม่มีร้าน"""
    d = yaml.safe_load(DEALERS.read_text(encoding="utf-8"))
    by_shop: dict[str, dict] = {}
    for x in d.get("dealers", []):
        for s in x["shops"]:
            by_shop[s] = {"ar_code": x["ar_code"], "group": x["group"],
                          "juristic": x.get("juristic", "")}
    return by_shop, d.get("no_shop_yet", [])


def run_sql(sql: str) -> list[list[str]]:
    """ส่งผ่านไฟล์เสมอ — SQL มีภาษาไทย และ path ต้องเป็น ASCII ล้วน"""
    env = dict(os.environ, PGPASSFILE=PGPASS, PGCLIENTENCODING="UTF8")
    tmp = PROJECT_ROOT / "output" / "_sku_review_tmp" / "sellout.sql"
    tmp.parent.mkdir(parents=True, exist_ok=True)
    tmp.write_text(sql, encoding="utf-8")
    p = subprocess.run([PSQL, "service=osuka", "-w", "-A", "-t", "-F", "\x1f",
                        "-f", str(tmp)], capture_output=True, env=env)
    if p.returncode != 0:
        print(p.stdout.decode("utf-8", "replace"), p.stderr.decode("utf-8", "replace"))
        raise SystemExit("❌ psql ไม่ผ่าน")
    return [ln.split("\x1f") for ln in
            p.stdout.decode("utf-8", "replace").splitlines() if ln.strip()]


COLS = [
    ("AR code", 13), ("Dealer group", 24), ("นิติบุคคล (SML)", 34),
    ("ร้าน (ชื่อมาตรฐาน)", 28), ("shop_id", 12), ("แพลตฟอร์ม", 12),
    ("บรรทัด", 11), ("ออเดอร์", 11),
    ("Sell-out รวม (บาท)", 20), ("Sell-out OSUKA (บาท)", 21), ("% OSUKA", 10),
]


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--from", dest="d_from", default="2026-01-01")
    ap.add_argument("--to", dest="d_to", default="2026-08-10")
    args = ap.parse_args()

    sql = f"""
    SELECT shop_id,
           max(platform),
           string_agg(DISTINCT shop_name, ' / ' ORDER BY shop_name),
           count(*),
           count(DISTINCT order_id),
           round(coalesce(sum(revenue_thb) FILTER (WHERE counts_as_sale),0)),
           round(coalesce(sum(revenue_thb) FILTER (WHERE counts_as_sale
                 AND product_brand ILIKE '%osuka%'),0))
    FROM   intel.mp_order_line
    WHERE  ordered_at >= DATE '{args.d_from}'
      AND  ordered_at <  DATE '{args.d_to}' + 1
    GROUP  BY shop_id
    ORDER  BY 6 DESC NULLS LAST
    """
    rows = run_sql(sql)
    print(f"ดึง {args.d_from} ถึง {args.d_to} · {len(rows)} ร้าน")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sell-out ต่อร้าน"
    for i, (name, w) in enumerate(COLS, start=1):
        c = ws.cell(row=1, column=i, value=name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30

    by_shop, no_shop = load_dealers()
    tot_all = tot_osk = 0
    for n, (shop_id, plat, names, lines, orders, total, osuka) in enumerate(rows, start=2):
        total_i, osuka_i = int(total or 0), int(osuka or 0)
        tot_all += total_i
        tot_osk += osuka_i
        dl = by_shop.get(shop_id, {})
        vals = [dl.get("ar_code", "ยังไม่ผูก"), dl.get("group", "ยังไม่ผูก"),
                dl.get("juristic", ""),
                canonical_name(shop_id, names.split(" / ")[0]), shop_id, plat,
                int(lines or 0), int(orders or 0), total_i, osuka_i,
                round(osuka_i / total_i * 100) if total_i else 0]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=n, column=i, value=v)
            if i in (7, 8, 9, 10):
                c.number_format = "#,##0"
            if i == 11:
                c.number_format = '0"%"'

    r = len(rows) + 2
    ws.cell(row=r, column=1, value="รวมทุกร้าน").font = Font(bold=True)
    ws.cell(row=r, column=9, value=tot_all).number_format = "#,##0"
    ws.cell(row=r, column=10, value=tot_osk).number_format = "#,##0"
    ws.cell(row=r, column=11,
            value=round(tot_osk / tot_all * 100) if tot_all else 0).number_format = '0"%"'
    for col in (1, 9, 10, 11):
        ws.cell(row=r, column=col).font = Font(bold=True)
    ws.freeze_panes = "D2"

    # ---- dealer group ที่ยังไม่มีร้านในระบบ ----
    g = wb.create_sheet("ยังไม่มีข้อมูล")
    g["A1"] = "Dealer group ที่ทีมมี แต่ระบบยังไม่ได้เก็บข้อมูลร้าน"
    g["A1"].font = Font(bold=True)
    g["A2"] = "ไม่ใช่ข้อมูลหาย — เป็นร้านที่ยังไม่ได้เปิดให้ระบบดึง ต้องเพิ่มร้านแล้วล็อกอินก่อน"
    for i, h in enumerate(["AR code", "Dealer group"], start=1):
        c = g.cell(row=4, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
    for i, x in enumerate(no_shop, start=5):
        g.cell(row=i, column=1, value=x.get("ar_code"))
        g.cell(row=i, column=2, value=x.get("group"))
    g.column_dimensions["A"].width = 14
    g.column_dimensions["B"].width = 70

    # ---- นิยาม ----
    d = wb.create_sheet("นิยามที่ใช้")
    notes = [
        "นิยามของตัวเลขในไฟล์นี้ — ต้องตรงกันก่อนเอาไปเทียบกับตัวเลขจากที่อื่น",
        "",
        f"ช่วงข้อมูล   {args.d_from} ถึง {args.d_to} (นับตาม ordered_at)",
        "",
        "Sell-out รวม",
        "  ผลรวม revenue_thb เฉพาะบรรทัดที่ counts_as_sale = true",
        "  counts_as_sale = order_status อยู่ใน DELIVERED / SHIPPED / READY TO SHIP",
        "  ไม่นับ UNPAID / CANCELLED / RETURNED / LOST BY 3PL",
        "",
        "revenue_thb คือยอดที่ผู้ซื้อจ่ายหลังหักส่วนลดแล้ว และ **รวม VAT**",
        "  ถ้าจะเทียบกับ sell-in ซึ่งเป็นราคาไม่รวม VAT ต้องหารด้วย 1.07 ก่อน",
        "",
        "Sell-out OSUKA",
        "  กรองด้วย product_brand ที่มีคำว่า osuka",
        "",
        "⚠️ รวมยอดด้วย shop_id ไม่ใช่ชื่อร้าน",
        "  ร้านเดียวกันมีหลายชื่อในฐาน เพราะข้อมูล 7 เดือนแรกโหลดด้วยชื่อดิบจากแพลตฟอร์ม",
        "  ส่วนที่โหลดหลัง 2026-08-11 ใช้ชื่อมาตรฐาน",
        "  คอลัมน์ 'ชื่อที่พบในฐาน' แสดงทุกชื่อที่เจอ เพื่อให้ตรวจย้อนได้",
        "",
        "⚠️ ข้อมูลชุดนี้ผ่านการแก้ 3 อย่างเมื่อ 2026-08-11 ซึ่งทำให้ยอดต่างจากรายงานรุ่นก่อน",
        "  1. แก้ Lazada 13,633 บรรทัดที่เคยถูกตีเป็น ADD TO CART ทั้งที่ส่งของแล้ว",
        "     ทำให้ยอด Lazada เพิ่มขึ้นราวเท่าตัว",
        "  2. เปลี่ยนชื่อสถานะ ADD TO CART เป็น UNPAID (ไม่กระทบยอด แค่เปลี่ยนชื่อ)",
        "  3. เติมคำสถานะที่ระบบเดิมแปลไม่ได้ ทำให้บางบรรทัดถูกนับเป็นยอดขายที่ควรนับ",
    ]
    for i, line in enumerate(notes, start=1):
        c = d.cell(row=i, column=1, value=line)
        if line and not line.startswith(" "):
            c.font = Font(bold=True)
    d.column_dimensions["A"].width = 96

    out_dir = PROJECT_ROOT / "output" / "_sellout_extract"
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / f"Sell-out ต่อร้าน_{args.d_from}_ถึง_{args.d_to}.xlsx"
    wb.save(out)
    print(f"✅ {out.relative_to(PROJECT_ROOT)}")
    print(f"   รวมทุกร้าน {tot_all:,.0f} บาท · OSUKA {tot_osk:,.0f} บาท "
          f"({tot_osk/tot_all*100:.0f}%)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
