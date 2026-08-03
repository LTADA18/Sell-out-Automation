from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.core.exporter import export_shop
from src.core.models import NULL, OrderStatus
from tests.conftest import make_order


def _export(orders, settings, run_date="2026-08-03") -> Path:
    return export_shop(
        orders,
        shop_id="lazada_01", platform="lazada", shop_name="ร้านทดสอบ",
        run_date=run_date, date_from="2026-08-02", date_to="2026-08-02",
        output_dir=Path(settings.paths.output_dir),
        archive_dir=Path(settings.paths.archive_dir),
    )


def test_creates_three_sheets(settings):
    wb = load_workbook(_export([make_order("100")], settings))
    assert wb.sheetnames == ["Orders", "Summary", "Meta"]


def test_long_ids_stay_text(settings):
    """เลข 19 หลักของ TikTok ถ้าหลุดเป็นตัวเลข Excel จะปัดหลักท้ายทิ้ง"""
    long_id = "1731354623761352306"
    wb = load_workbook(_export([make_order(long_id, sku="9876543210123456789")], settings))
    ws = wb["Orders"]
    hdr = [c.value for c in ws[1]]

    for field, expected in (("order_id", long_id), ("sku", "9876543210123456789")):
        cell = ws.cell(row=2, column=hdr.index(field) + 1)
        assert isinstance(cell.value, str), f"{field} ต้องเป็น str ไม่ใช่ {type(cell.value)}"
        assert cell.value == expected
        assert cell.number_format == "@"


def test_int_id_coerced_to_text(settings):
    """กันกรณี adapter เผลอส่ง int มา — validator ต้องแปลงให้เป็น str"""
    order = make_order(1118857732989544)
    assert order.order_id == "1118857732989544"
    assert isinstance(order.order_id, str)


def test_missing_value_becomes_null_not_blank(settings):
    """กฎเหล็ก: ไม่มีข้อมูล = "Null" ไม่ใช่ช่องว่าง จะได้แยกออกจาก 'ลืมดึง'"""
    wb = load_workbook(_export([make_order("100", tracking_no=None)], settings))
    ws = wb["Orders"]
    hdr = [c.value for c in ws[1]]
    assert ws.cell(row=2, column=hdr.index("tracking_no") + 1).value == NULL


def test_order_level_amount_not_double_counted(settings):
    """1 ออเดอร์ 3 สินค้า = 3 แถว แต่ total_amount ต้องนับครั้งเดียว"""
    orders = [make_order("500", sku=f"SKU-{i}", total_amount=1000.0) for i in range(3)]
    wb = load_workbook(_export(orders, settings))
    rows = list(wb["Summary"].iter_rows(min_row=2, values_only=True))
    assert rows[0][1] == 1, "ต้องนับเป็น 1 ออเดอร์ ไม่ใช่ 3"
    assert rows[0][3] == 1000.0, "ยอดขายต้องเป็น 1000 ไม่ใช่ 3000"


def test_aov_row_matches_total_row(settings):
    """AOV แถวรวมต้องคิดฐานเดียวกับแถวรายวัน (ไม่รวมออเดอร์ที่ยกเลิก)"""
    orders = [
        make_order("1", total_amount=1000.0),
        make_order("2", total_amount=3000.0),
        make_order("3", total_amount=500.0, order_status=OrderStatus.CANCELLED),
    ]
    ws = load_workbook(_export(orders, settings))["Summary"]
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
    daily, total = rows[0], rows[-1]
    assert daily[5] == total[5] == 2000.0        # (1000+3000)/2 ออเดอร์ที่ขายจริง
    assert total[4] == 1, "ต้องนับออเดอร์ยกเลิก 1 รายการ"


def test_rerun_archives_previous_file(settings):
    first = _export([make_order("1")], settings)
    assert first.exists()
    second = _export([make_order("1"), make_order("2")], settings)

    assert second == first, "ชื่อไฟล์ต้องเหมือนเดิม (idempotent)"
    archived = list(Path(settings.paths.archive_dir).rglob("*.xlsx"))
    assert len(archived) == 1, "ไฟล์เดิมต้องถูกย้ายไป archive ก่อนเขียนทับ"


def test_empty_orders_still_produces_file(settings):
    """ร้านที่ไม่มีออเดอร์ต้องยังได้ไฟล์ ไม่งั้นชุดรายวันขาดไป"""
    wb = load_workbook(_export([], settings))
    assert wb.sheetnames == ["Orders", "Summary", "Meta"]
    assert wb["Orders"].max_row == 1              # มีแต่หัวตาราง
