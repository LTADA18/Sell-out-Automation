"""เขียน Excel 1 ร้าน = 1 ไฟล์ 3 sheet (Orders / Summary / Meta)"""

from __future__ import annotations

import shutil
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.core.models import (
    EXCEL_COLUMNS,
    MONEY_COLUMNS,
    NULL,
    TEXT_COLUMNS,
    Order,
    OrderStatus,
)

SCRIPT_VERSION = "0.1.0"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MONEY_FMT = "#,##0.00"
DATE_FMT = "yyyy-mm-dd hh:mm:ss"

# สถานะที่ถือว่า "ไม่เป็นยอดขาย" ตอนคิด Summary
NON_SALE = frozenset({OrderStatus.CANCELLED, OrderStatus.RETURNED})


def _cell_value(order: Order, field: str) -> object:
    """แปลงค่า 1 field ให้พร้อมเขียนลง cell

    กฎเหล็ก: ไม่มีข้อมูล = "Null" (string) ไม่ใช่ปล่อยว่าง
    เพื่อให้แยกออกว่า 'ดึงมาแล้วไม่มีค่า' ต่างจาก 'ลืมดึง/คอลัมน์หาย'
    """
    val = getattr(order, field, None)
    if val is None:
        return NULL
    if field in TEXT_COLUMNS:
        return str(val)
    if isinstance(val, OrderStatus):
        return val.value
    return val


def _style_header(ws: Worksheet, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _autofit(ws: Worksheet, max_width: int = 42) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        widest = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(widest + 2, 10), max_width)


#  อักขระที่ Windows ห้ามใช้ในชื่อไฟล์ — ชื่อร้านจริงมีทั้งวงเล็บ เว้นวรรค และภาษาไทย
_BAD_CHARS = r'\/:*?"<>|'


def safe_name(name: str, max_len: int = 30) -> str:
    """ทำชื่อร้านให้ใช้เป็นชื่อไฟล์ได้ — คงภาษาไทยไว้ แต่ตัดอักขระต้องห้ามทิ้ง

    เว้นวรรคเปลี่ยนเป็น "-" เพื่อให้ก๊อปชื่อไฟล์ไปวางในคำสั่งได้โดยไม่ต้องใส่เครื่องหมายคำพูด
    ตัดความยาวกันชนเพดาน path ของ Windows (ชื่อร้านบางร้านยาวมาก)
    """
    cleaned = "".join(" " if c in _BAD_CHARS else c for c in (name or ""))
    cleaned = "-".join(cleaned.split())          # ยุบช่องว่างซ้ำ + เปลี่ยนเป็น -
    cleaned = cleaned.strip("-.")                 # Windows ไม่ชอบชื่อลงท้ายด้วยจุด
    return cleaned[:max_len].strip("-.") or "ไม่ระบุชื่อ"


def _day_of(order: Order) -> str:
    return order.order_created_at.strftime("%Y-%m-%d") if order.order_created_at else NULL


def _write_orders_sheet(wb: Workbook, orders: list[Order], title: str = "Orders") -> None:
    ws = wb.create_sheet(title)
    ws.append(list(EXCEL_COLUMNS))

    for order in orders:
        ws.append([_cell_value(order, f) for f in EXCEL_COLUMNS])

    for idx, field in enumerate(EXCEL_COLUMNS, start=1):
        letter = get_column_letter(idx)
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=idx)
            if field in TEXT_COLUMNS:
                # '@' = บังคับ text ไม่งั้น Excel แปลงเลข 19 หลักเป็น 1.23457E+18 แล้วหลักท้ายหาย
                cell.number_format = "@"
            elif field in MONEY_COLUMNS and isinstance(cell.value, (int, float)):
                cell.number_format = MONEY_FMT
            elif isinstance(cell.value, datetime):
                cell.number_format = DATE_FMT
        ws.column_dimensions[letter].width = 18

    _style_header(ws, len(EXCEL_COLUMNS))
    _autofit(ws)


def _write_day_sheets(wb: Workbook, orders: list[Order]) -> list[str]:
    """แตกชีทเพิ่ม 1 วัน = 1 ชีท ชื่อชีทเป็น YYYY-MM-DD

    ทำเฉพาะตอนข้อมูลคร่อมหลายวัน (เช่น backfill / --from --to)
    รอบรายวันปกติมีวันเดียว ชีท Orders ก็คือวันนั้นอยู่แล้ว แตกไปก็ซ้ำเปล่า ๆ
    """
    by_day: dict[str, list[Order]] = defaultdict(list)
    for o in orders:
        by_day[_day_of(o)].append(o)

    if len(by_day) <= 1:
        return []

    for day in sorted(by_day):
        _write_orders_sheet(wb, by_day[day], title=day)
    return sorted(by_day)


def _summarize(orders: list[Order]) -> tuple[list[list], dict]:
    """สรุปรายวัน — นับที่ระดับ 'ออเดอร์' ไม่ใช่ 'แถว'

    1 ออเดอร์ที่มี 3 สินค้า = 3 แถว ถ้านับแถวจะกลายเป็น 3 ออเดอร์ ยอดเพี้ยนทันที
    """
    by_day: dict[str, list[Order]] = defaultdict(list)
    for o in orders:
        by_day[_day_of(o)].append(o)

    rows: list[list] = []
    # sale_orders นับแยกจาก orders เพราะ AOV ต้องหารด้วย "ออเดอร์ที่เป็นยอดขายจริง"
    # ถ้าหารด้วยออเดอร์ทั้งหมด แถวรวมจะได้ค่าไม่ตรงกับแถวรายวัน
    totals = {"orders": 0, "sale_orders": 0, "units": 0, "sales": 0.0, "cancelled": 0}

    for day in sorted(by_day):
        day_orders = by_day[day]
        order_ids = {o.order_id for o in day_orders}
        # total_amount อยู่ระดับออเดอร์ — ต้องนับครั้งเดียวต่อออเดอร์ ไม่งั้นบวกซ้ำตามจำนวนสินค้า
        amount_by_order = {
            o.order_id: o.total_amount for o in day_orders if o.total_amount is not None
        }
        status_by_order = {o.order_id: o.order_status for o in day_orders}

        sale_ids = {oid for oid, st in status_by_order.items() if st not in NON_SALE}
        cancelled_ids = {oid for oid, st in status_by_order.items() if st in NON_SALE}

        gross = sum(amount_by_order.get(oid, 0.0) or 0.0 for oid in sale_ids)
        units = sum(o.quantity or 0 for o in day_orders if o.order_status not in NON_SALE)
        aov = round(gross / len(sale_ids), 2) if sale_ids else NULL

        status_counts = Counter(st.value for st in status_by_order.values())
        rows.append([
            day,
            len(order_ids),
            units,
            round(gross, 2),
            len(cancelled_ids),
            aov,
            ", ".join(f"{k}={v}" for k, v in sorted(status_counts.items())) or NULL,
        ])

        totals["orders"] += len(order_ids)
        totals["sale_orders"] += len(sale_ids)
        totals["units"] += units
        totals["sales"] += gross
        totals["cancelled"] += len(cancelled_ids)

    return rows, totals


def _write_summary_sheet(wb: Workbook, orders: list[Order]) -> None:
    ws = wb.create_sheet("Summary")
    headers = ["วันที่", "จำนวนออเดอร์", "จำนวนชิ้น", "ยอดขายรวม",
               "ออเดอร์ยกเลิก/คืน", "AOV", "แยกตามสถานะ"]
    ws.append(headers)

    rows, totals = _summarize(orders)
    for r in rows:
        ws.append(r)

    if rows:
        ws.append([])
        ws.append([
            "รวม", totals["orders"], totals["units"], round(totals["sales"], 2),
            totals["cancelled"],
            round(totals["sales"] / totals["sale_orders"], 2) if totals["sale_orders"] else NULL,
            "",
        ])
        for c in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
    else:
        ws.append([NULL] * len(headers))

    for row in range(2, ws.max_row + 1):
        for col in (4, 6):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = MONEY_FMT

    _style_header(ws, len(headers))
    _autofit(ws)


def _write_meta_sheet(wb: Workbook, meta: dict) -> None:
    ws = wb.create_sheet("Meta")
    ws.append(["หัวข้อ", "ค่า"])
    for k, v in meta.items():
        ws.append([k, NULL if v is None else str(v)])
    _style_header(ws, 2)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 60


def export_shop(
    orders: list[Order],
    *,
    shop_id: str,
    platform: str,
    shop_name: str,
    run_date: str,
    date_from: str,
    date_to: str,
    output_dir: Path,
    archive_dir: Path,
    status: str = "SUCCESS",
    notes: str | None = None,
) -> Path:
    """เขียน Excel 1 ไฟล์ คืน path — รันซ้ำวันเดิมได้ (idempotent) ไฟล์เดิมย้ายไป archive"""
    day_dir = Path(output_dir) / run_date
    day_dir.mkdir(parents=True, exist_ok=True)
    # ใส่ชื่อร้านในชื่อไฟล์ด้วย — เปิดโฟลเดอร์แล้วรู้เลยว่าไฟล์ไหนของร้านไหน
    # ยังคง shop_id ไว้เพราะเป็นคีย์ที่ตรงกับ Dashboard/log และไม่เปลี่ยนแม้ร้านจะเปลี่ยนชื่อ
    out_path = day_dir / f"{platform}_{shop_id}_{safe_name(shop_name)}_{run_date}.xlsx"

    # ย้ายของเดิมก่อนทับ — ถ้ารอบใหม่ดึงได้น้อยกว่าเดิม จะได้ย้อนกลับไปดูได้
    # ใช้ glob เพราะไฟล์เก่าอาจเป็นชื่อรูปแบบก่อนหน้า (ยังไม่มีชื่อร้าน) หรือร้านเปลี่ยนชื่อ
    old = [p for p in day_dir.glob(f"{platform}_{shop_id}_*{run_date}.xlsx")]
    old += [p for p in day_dir.glob(f"{platform}_{shop_id}_{run_date}.xlsx")]
    if old:
        arc_dir = Path(archive_dir) / run_date
        arc_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%H%M%S")
        for p in dict.fromkeys(old):              # กันซ้ำถ้า glob ทั้งสองอันชนกัน
            shutil.move(str(p), str(arc_dir / f"{p.stem}_{stamp}.xlsx"))

    wb = Workbook()
    wb.remove(wb.active)                     # ตัด sheet เปล่าที่ openpyxl แถมมา

    _write_orders_sheet(wb, orders)
    day_sheets = _write_day_sheets(wb, orders)
    _write_summary_sheet(wb, orders)
    _write_meta_sheet(wb, {
        "ร้าน": f"{shop_name} ({shop_id})",
        "แพลตฟอร์ม": platform,
        "วันที่ของรอบ": run_date,
        "ช่วงข้อมูล": f"{date_from} ถึง {date_to}",
        "จำนวนแถว (order line)": len(orders),
        "จำนวนออเดอร์": len({o.order_id for o in orders}),
        "ชีทรายวัน": ", ".join(day_sheets) if day_sheets else "ไม่มี (ข้อมูลวันเดียว)",
        "เวลาที่ดึง": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "สถานะ": status,
        "เวอร์ชันสคริปต์": SCRIPT_VERSION,
        "หมายเหตุ": notes,
    })

    wb.save(out_path)
    return out_path
