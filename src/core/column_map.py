"""อ่านไฟล์ Export ของแต่ละแพลตฟอร์มตามกติกาใน config/column_maps/{platform}.yaml

โค้ดตรงนี้ไม่รู้จักชื่อคอลัมน์ของใครเลย — ทุกอย่างมาจาก YAML
หลังบ้านเปลี่ยนชื่อคอลัมน์เมื่อไหร่ แก้ YAML อย่างเดียว ไม่ต้องแตะ .py
"""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

from src.core.config import PROJECT_ROOT
from src.core.models import OrderStatus

MAP_DIR = PROJECT_ROOT / "config" / "column_maps"


class PlatformMap:
    def __init__(self, platform: str) -> None:
        path = MAP_DIR / f"{platform}.yaml"
        if not path.exists():
            raise FileNotFoundError(
                f"ไม่มี column map ของ '{platform}' ที่ {path} — "
                f"ต้องมีไฟล์ Export ตัวอย่างก่อนถึงจะทำ map ได้"
            )
        self.platform = platform
        self.raw: dict = yaml.safe_load(path.read_text(encoding="utf-8"))
        self.fields: dict[str, list[str]] = self.raw.get("fields", {})
        self.quirks: dict = self.raw.get("quirks", {})
        self.flow: dict = self.raw.get("export_flow", {})
        self._file: dict = self.raw.get("file", {})
        self._status_map: dict = self.raw.get("status_map", {})
        self._dt_fmt: str = self.raw.get("datetime_format", "")
        self.pii_columns: set[str] = set(self.raw.get("pii_columns", []))

    # ── อ่านไฟล์ ─────────────────────────────────────────────

    @property
    def data_start_row(self) -> int:
        return int(self._file.get("data_start_row", 2))

    @property
    def header_row(self) -> int:
        return int(self._file.get("header_row", 1))

    def read_export(self, path: Path) -> list[dict[str, Any]]:
        """คืนแถวข้อมูลเป็น dict {ชื่อคอลัมน์ดิบ: ค่า} — ยังไม่แปลงเป็น Order"""
        if path.suffix.lower() == ".csv":
            return self._read_csv(path)
        return self._read_xlsx(path)

    def _read_xlsx(self, path: Path) -> list[dict[str, Any]]:
        """อ่าน xlsx โดยลองโหมดเร็วก่อน แล้ว fallback ถ้าอ่านคอลัมน์ได้ไม่ครบ

        ไฟล์ Export ของทั้ง Lazada และ TikTok ไม่ประกาศ dimension ให้ถูกต้อง
        ผลคือ openpyxl โหมด read_only มองเห็นแค่คอลัมน์เดียว (`Order ID` ตัวเดียว)
        แล้วข้อมูลที่เหลืออีก 60-70 คอลัมน์หายเงียบ ๆ โดยไม่มี error
        reset_dimensions() ช่วยได้กับ Lazada แต่ไม่ช่วยกับ TikTok
        จึงต้องเช็คความกว้างที่อ่านได้จริง แล้วโหลดใหม่แบบเต็มถ้าน้อยผิดปกติ
        """
        rows = self._rows_from(path, read_only=True)
        if rows and len(rows[0]) <= 1:
            rows = self._rows_from(path, read_only=False)
        return rows

    def _rows_from(self, path: Path, read_only: bool) -> list[dict[str, Any]]:
        wb = load_workbook(path, read_only=read_only, data_only=True)
        try:
            sheet = self._file.get("sheet")
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
            if read_only:
                ws.reset_dimensions()

            rows: list[dict[str, Any]] = []
            header: list[str] = []
            for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if idx == self.header_row:
                    header = [str(c).strip() if c is not None else "" for c in row]
                elif idx >= self.data_start_row and header:
                    if all(c is None for c in row):
                        continue
                    padded = list(row) + [None] * (len(header) - len(row))
                    rows.append(self._dedupe(header, padded))
            return rows
        finally:
            wb.close()

    def _read_csv(self, path: Path) -> list[dict[str, Any]]:
        with path.open(encoding="utf-8-sig", newline="") as fh:
            reader = list(csv.reader(fh))
        header = [c.strip() for c in reader[self.header_row - 1]]
        out = []
        for row in reader[self.data_start_row - 1:]:
            if not any(row):
                continue
            out.append(self._dedupe(header, row + [None] * (len(header) - len(row))))
        return out

    @staticmethod
    def _dedupe(header: list[str], values: list) -> dict[str, Any]:
        """จับคู่หัวตารางกับค่า โดย **เก็บคอลัมน์แรกไว้** ถ้าชื่อซ้ำ

        ⚠️ dict(zip(...)) เฉย ๆ จะให้คอลัมน์หลังทับคอลัมน์แรก
        ไฟล์ Shopee มี "จังหวัด" 2 คอลัมน์ (ที่อยู่จัดส่ง กับ ที่อยู่ใบกำกับภาษี)
        ตัวหลังว่างเกือบทุกแถว ถ้าปล่อยให้ทับ province จะกลายเป็น Null 445/484

        คอลัมน์ที่ซ้ำยังเข้าถึงได้ด้วยชื่อ "<ชื่อ>__2", "<ชื่อ>__3" เผื่อวันหลังต้องใช้
        """
        out: dict[str, Any] = {}
        seen: dict[str, int] = {}
        for name, value in zip(header, values):
            if not name:
                continue
            if name not in out:
                out[name] = value
                seen[name] = 1
            else:
                seen[name] += 1
                out[f"{name}__{seen[name]}"] = value
        return out

    # ── แปลงค่า ──────────────────────────────────────────────

    def get(self, row: dict, field: str) -> Any:
        """ดึงค่าตามชื่อ field กลาง โดยไล่หาชื่อคอลัมน์ดิบตามลำดับใน YAML"""
        for candidate in self.fields.get(field, []):
            if candidate in row:
                val = row[candidate]
                if val is None:
                    return None
                if isinstance(val, str):
                    val = val.strip()
                    return val or None
                return val
        return None

    def parse_dt(self, value: Any) -> datetime | None:
        if value is None or value == "":
            return None
        if isinstance(value, datetime):
            return value
        text = str(value).strip()
        for fmt in (self._dt_fmt, "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%d %b %Y %H:%M"):
            if not fmt:
                continue
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
        return None      # แปลงไม่ได้ = None ไม่เดา (กฎเหล็กข้อ 1)

    def map_status(self, raw: Any) -> OrderStatus:
        if raw is None:
            return OrderStatus.UNKNOWN
        text = str(raw).strip()
        for key in (text, text.lower()):
            if key in self._status_map:
                return OrderStatus(self._status_map[key])
        return OrderStatus.UNKNOWN

    @staticmethod
    def to_float(value: Any) -> float | None:
        if value is None or value == "":
            return None
        try:
            return float(str(value).replace(",", "").replace("฿", "").strip())
        except ValueError:
            return None

    @staticmethod
    def to_text(value: Any) -> str | None:
        """id ทุกชนิดต้องเป็น string — เลข 19 หลักถ้าเป็น float จะเสียหลักท้าย"""
        if value is None or value == "":
            return None
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip() or None
